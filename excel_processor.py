#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel表格处理工具 - 调用Dify API匹配答案并填充渠道信息
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl import load_workbook
import requests
import json
import threading
import difflib
import time
import os
import sys
import re
import ast
import uuid
import shutil
import subprocess
import io
import contextlib
import tempfile
from datetime import datetime


# ---- API Key 查看保护 ----
APIKEY_VIEW_PASSWORD = "rmeet2026.."   # 查看 API Key 明文所需的校验密码
APIKEY_UNLOCK_SECONDS = 60             # 解锁后自动重新掩码的时长（秒）
APIKEY_CLICK_REQUIRED = 5              # 触发查看所需的连续点击次数
APIKEY_CLICK_WINDOW = 2.0              # 连续点击判定窗口（秒），超过则重新计数

SPLASH_MS = 1500                       # 启动画面展示时长（毫秒）


def _apikey_click_tick(state, now):
    """API Key 连续点击计数：窗口内达到 APIKEY_CLICK_REQUIRED 次返回 True 并清零，否则返回 False。
    state: {"clicks": int, "last": float}，now: time.monotonic() 值"""
    if now - state["last"] > APIKEY_CLICK_WINDOW:
        state["clicks"] = 0
    state["clicks"] += 1
    state["last"] = now
    if state["clicks"] >= APIKEY_CLICK_REQUIRED:
        state["clicks"] = 0
        return True
    return False


def _apikey_should_protect(value, default, unlocked):
    """是否需要掩码保护：值为默认值且当前未解锁（未通过密码查看）"""
    return value == default and not unlocked


def _deep_find(obj, key, max_depth=6):
    """递归搜索嵌套 dict/list 中指定 key 的值；找不到返回 None"""
    if max_depth < 0:
        return None
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        for v in obj.values():
            found = _deep_find(v, key, max_depth - 1)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = _deep_find(item, key, max_depth - 1)
            if found is not None:
                return found
    return None


def _extract_systemfrom_from_text(text):
    """从回答文本中提取系统来源：优先机器可读标记，其次常见来源表述；找不到返回 ''"""
    if not text:
        return ''
    patterns = [
        r'【系统来源】\s*([^\n]+)',
        r'systemfrom\s*[：:]\s*([^\n，。,;；]+)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            if val:
                return val
    return ''


def _extract_json_substring(text):
    """提取文本中第一个平衡的 {...} 或 [...] 块（正确处理字符串内括号与转义）；找不到返回 None"""
    if not text:
        return None
    start = -1
    for i, ch in enumerate(text):
        if ch in '{[':
            start = i
            break
    if start < 0:
        return None
    open_ch = text[start]
    close_ch = '}' if open_ch == '{' else ']'
    depth = 0
    in_str = False
    esc = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if esc:
                esc = False
            elif ch == '\\':
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def _parse_json_loose(text):
    """宽松解析 JSON 文本：围栏剥离 → json.loads → ast.literal_eval（单引号/True/False/None）→ 平衡子串提取 → None"""
    if not text:
        return None
    stripped = text.strip()
    if stripped.startswith('```'):
        stripped = stripped.strip('`').strip()
        if stripped.lower().startswith('json'):
            stripped = stripped[4:].strip()
    candidates = [stripped]
    sub = _extract_json_substring(stripped)
    if sub and sub != stripped:
        candidates.append(sub)
    for cand in candidates:
        try:
            return json.loads(cand)
        except (json.JSONDecodeError, ValueError):
            pass
        try:
            val = ast.literal_eval(cand)
            if isinstance(val, (dict, list)):
                return json.loads(json.dumps(val))
        except (ValueError, SyntaxError):
            pass
    return None


# ---------- 文件处理 Agent（Phase 1）确定性工具 ----------

# Excel 列号常量（1-based）
COL_MSG_ID = 1
COL_QUESTION = 8
COL_ANSWER = 9
COL_HIT_TYPE = 12
COL_CHANNEL = 15
COL_FAIL = 16
COL_CLEAN_ANSWER = 17
COL_ARR = 18
COL_MATCH_TYPE = 19
COL_SIM_DETAIL = 20

# 文件列结构说明（写入意图识别 prompt）
COLUMN_STRUCTURE = (
    "1列:消息ID, 8列(H):问题, 9列(I):答案, 12列(L):命中类型, "
    "15列(O):渠道(含(反向)/(正向)标识), 16列(P):失败标记(=\"失败\"), "
    "17列(Q):对比原始问题, 18列(R):arr数组数据, "
    "19列(S):匹配方式(关键字匹配成功/相似度匹配成功/匹配失败/正向工作流获取成功/正向失败:原因), "
    "20列(T):相似度接口详情"
)

# Agent 任务白名单
AGENT_TASK_WHITELIST = {'statistics', 'export_failed', 'retry', 'diff', 'export_csv'}

INTENT_PROMPT_TEMPLATE = """你是一个Excel文件处理需求识别器。以下是选中Excel文件的列结构说明：
{column_structure}
{history_section}
运维人员的需求描述：
{request}

请从以下任务中选择最匹配的一个并提取参数，只输出JSON对象（不要输出任何其他内容），格式为：
{{
  "task": "任务名",
  "params": {{}},
  "confidence": 0到1的置信度,
  "understanding": "用1-2句通俗口语（像客服聊天一样、有温度）以第一人称复述你打算为用户做什么，例如'我来帮您把O列里出现的『(反向)』这几个字去掉，改完保存成一个新文件'；不要用'用户希望…'这类书面语，不要出现文件序号[0]这类内部术语（供理解确认展示）",
  "friendly_summary": "用3-5句客服式口语整体说明你将如何处理：涉及哪个文件、具体做什么操作、结果保存在哪里。像微信聊天一样自然、有温度、通俗易懂，不要出现参数JSON、不要用'用户希望…'书面语（供执行方案预览展示）",
  "clarify_questions": ["需求中含糊或缺失的关键信息问题；明确则为空数组"],
  "can_use_builtin_tool": true或false
}}

可选任务：
1. {{"task":"statistics","params":{{}}}}  统计报表：关键字/相似度匹配占比、成功率、渠道分布、命中类型、失败原因
2. {{"task":"export_failed","params":{{"scope":"reverse|forward|all","format":"excel|csv"}}}}  导出失败项为新文件
3. {{"task":"retry","params":{{"mode":"reverse|forward"}}}}  对失败行重新调用接口处理
4. {{"task":"diff","params":{{"file_index":0,"other_file_index":1}}}}  对比两个选中文件的O列渠道差异（file_index为文件序号，从0开始）
5. {{"task":"export_csv","params":{{"columns":"1,8,15","filter":"关键词"}}}}  通用导出指定列为CSV
6. 原子工具也可直接选用：
   filter_excel(column,operator,value) / sort_excel(column,desc) / group_excel(column) /
   merge_excel / split_excel(rows_per_sheet) / modify_excel(ops) / export_excel /
   inspect_excel / read_excel(limit) / read_text / search_text(keyword) /
   replace_text(find,replace) / copy_file / validate_output
   例如：{{"task":"filter_excel","params":{{"column":"P","operator":"equals","value":"失败"}}}}

工具适用范围：
- read_text / search_text / replace_text 仅适用于 txt/csv 等文本文件；
- Excel（xlsx/xls）文件的查找替换、读取请使用 modify_excel / read_excel / filter_excel 等 Excel 工具，不要选用文本工具。
- modify_excel 参数模板：{{"task":"modify_excel","params":{{"ops":[{{"type":"replace","column":"O","find":"(反向)","replace":""}}],"save_mode":"new|overwrite"}}}}
  （用于 Excel 单元格内容查找替换/删除行等；Excel 文件请优先用它而非 replace_text）
  save_mode：当用户要求"保存回原文件/覆盖原文件/直接修改原文件/保留在该文件"时填 "overwrite"（覆盖不可逆）；未明确要求改原文件时省略或填 "new"（另存新文件）。

如果需求无法用上述任务完成，输出：{{"task":"custom","params":{{}},"reason":"简要说明","can_use_builtin_tool":false}}

澄清规则：
- 当需求存在歧义、缺少关键参数（统计口径/列名/范围/输出格式等）、或引用了文件列结构中不存在的列时，必须在 clarify_questions 中逐条列出需要用户补充的问题。
- 当需求明确无歧义时，clarify_questions 必须为空数组 []。
- 提取 find/equals 等查找关键字时，若文件列结构或数据样本中的实际写法与用户描述存在符号差异（如全角/半角括号、中英文标点），优先采用文件中实际出现的符号形式；无法确定时放入 clarify_questions 请用户确认写法。

注意：绝对不要输出本地文件的绝对路径，文件引用一律使用序号索引（file_index）。"""


def _format_history_section(conversation):
    """将多轮对话历史格式化为 prompt 片段（无历史时返回空串）"""
    conv = conversation or []
    if not conv:
        return ''
    lines = ['此前对话背景：']
    for item in conv[-6:]:  # 截断，仅保留最近 6 条
        role = '用户' if item.get('role') == 'user' else 'Agent'
        lines.append(f"{role}: {item.get('content', '')}")
    return '\n'.join(lines)


REFLECTION_PROMPT_TEMPLATE = """你是文件处理结果质检员。以下是：
1. 用户需求：
{request}

2. 输入文件结构：
{input_meta}

3. 输出文件结构：
{output_meta}

4. 输出摘要：
{output_summary}

5. 输入内容样本（前3行）：
{input_samples}

6. 输出内容样本（前3行）：
{output_samples}

请判断输出结果是否真正满足用户需求。注意：
- 结合内容样本对比输入与输出：若需求要求修改单元格内容（如去除某字样），但输出内容与输入完全一致，视为未满足；
- 若内容已按要求变化（即使文件结构/行列数不变），视为满足。
只输出JSON对象（不要输出任何其他内容）：
{{
  "satisfied": true或false,
  "reason": "判断依据（简短）",
  "suggestions": "若不满足，给出具体改进建议（一句话）；满足则为空字符串"
}}"""


def _cell_str(row, idx):
    """安全读取行中第 idx 列（1-based）的字符串值；越界或空返回 ''"""
    if idx < 1 or idx > len(row):
        return ''
    v = row[idx - 1].value
    return '' if v is None else str(v).strip()


def scan_worksheet(ws):
    """逐行归入「状态×匹配方式」桶，返回统计 dict（供统计/报表工具使用）"""
    stats = {
        'total': 0,
        'reverse_success': 0,
        'forward_success': 0,
        'keyword_match': 0,
        'similarity_match': 0,
        'reverse_only': 0,
        'forward_only': 0,
        'both': 0,
        'both_fail': 0,
        'channels': {},
        'hit_types': {},
        'fail_reasons': {},
    }
    for row in ws.iter_rows(min_row=2):
        o_val = _cell_str(row, COL_CHANNEL)
        p_val = _cell_str(row, COL_FAIL)
        s_val = _cell_str(row, COL_MATCH_TYPE)
        l_val = _cell_str(row, COL_HIT_TYPE)
        if not o_val and not p_val and not s_val:
            continue  # 空行跳过
        stats['total'] += 1

        is_fail = p_val.startswith('失败')

        # 匹配方式
        if s_val == '关键字匹配成功':
            stats['keyword_match'] += 1
            stats['reverse_success'] += 1
        elif s_val == '相似度匹配成功':
            stats['similarity_match'] += 1
            stats['reverse_success'] += 1
        elif '正向工作流获取成功' in s_val:
            stats['forward_success'] += 1

        # 渠道正反向命中情况
        has_rev = '(反向)' in o_val
        has_fwd = '(正向)' in o_val
        if has_rev and has_fwd:
            stats['both'] += 1
        elif has_rev:
            stats['reverse_only'] += 1
        elif has_fwd:
            stats['forward_only'] += 1
        elif is_fail:
            stats['both_fail'] += 1

        # 渠道分布（剥离 (反向)/(正向) 标识）
        for part in o_val.split(','):
            part = part.strip().replace('(反向)', '').replace('(正向)', '').strip()
            if part:
                stats['channels'][part] = stats['channels'].get(part, 0) + 1

        # 命中类型分布
        if l_val:
            stats['hit_types'][l_val] = stats['hit_types'].get(l_val, 0) + 1

        # 失败原因分组
        if is_fail or '失败' in s_val or '异常' in s_val:
            if '正向失败' in s_val:
                key = '正向失败'
            elif '异常' in s_val:
                key = '异常'
            elif '匹配失败' in s_val:
                key = '匹配失败'
            else:
                key = s_val[:20] if s_val else '未知原因'
            stats['fail_reasons'][key] = stats['fail_reasons'].get(key, 0) + 1
    return stats


def generate_statistics_report(ws):
    """生成多行文本统计报告"""
    s = scan_worksheet(ws)
    lines = ['=' * 52, f"文件统计报表（数据行数: {s['total']}）", '-' * 52]
    lines.append(f"反向匹配成功: {s['reverse_success']}")
    lines.append(f"  ├─ 关键字匹配成功: {s['keyword_match']}")
    lines.append(f"  └─ 相似度匹配成功: {s['similarity_match']}")
    lines.append(f"正向匹配成功: {s['forward_success']}")
    lines.append(f"渠道命中情况: 仅反向 {s['reverse_only']} | 仅正向 {s['forward_only']} | 双命中 {s['both']} | 均失败 {s['both_fail']}")
    if s['channels']:
        lines.append('-' * 52)
        lines.append("渠道来源分布:")
        for name, cnt in sorted(s['channels'].items(), key=lambda x: -x[1]):
            lines.append(f"  {name}: {cnt}")
    if s['hit_types']:
        lines.append('-' * 52)
        lines.append("命中类型分布:")
        for t, cnt in sorted(s['hit_types'].items(), key=lambda x: -x[1]):
            lines.append(f"  {t}: {cnt}")
    if s['fail_reasons']:
        lines.append('-' * 52)
        lines.append("失败原因分组:")
        for r, cnt in sorted(s['fail_reasons'].items(), key=lambda x: -x[1]):
            lines.append(f"  {r}: {cnt}")
    lines.append('=' * 52)
    return '\n'.join(lines)


def export_failed_rows(ws, out_dir, source_name, scope='all', fmt='excel'):
    """导出失败行（P列=失败）为新文件；返回 (输出路径, 条数)"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    import csv as _csv

    header = [c.value for c in ws[1]]
    failed = []
    for row in ws.iter_rows(min_row=2):
        p_val = _cell_str(row, COL_FAIL)
        if not p_val.startswith('失败'):
            continue
        s_val = _cell_str(row, COL_MATCH_TYPE)
        is_forward_fail = s_val.startswith('正向失败')
        if scope == 'reverse' and is_forward_fail:
            continue
        if scope == 'forward' and not is_forward_fail:
            continue
        failed.append([c.value for c in row])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"{source_name}_agent_export_failed_{timestamp}"
    if fmt == 'csv':
        path = os.path.join(out_dir, base + '.csv')
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = _csv.writer(f)
            w.writerow(header)
            w.writerows(failed)
        return path, len(failed)

    path = os.path.join(out_dir, base + '.xlsx')
    wb = Workbook()
    new_ws = wb.active
    new_ws.append(header)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for r in failed:
        new_ws.append(r)
    for row_idx in range(2, new_ws.max_row + 1):
        new_ws.cell(row=row_idx, column=COL_FAIL).fill = red_fill
    wb.save(path)
    return path, len(failed)


def diff_worksheets(ws_a, ws_b, out_dir, name_a, name_b):
    """对比两个工作表的 O 列渠道差异；返回 (输出CSV路径, 文本报告)"""
    import csv as _csv

    def collect(ws):
        d = {}
        for row in ws.iter_rows(min_row=2):
            row_idx = row[0].row
            msg_id = _cell_str(row, COL_MSG_ID)
            o_val = _cell_str(row, COL_CHANNEL)
            if o_val:
                d[row_idx] = (msg_id, o_val)
        return d

    da, db = collect(ws_a), collect(ws_b)
    only_a = {k: v for k, v in da.items() if k not in db}
    only_b = {k: v for k, v in db.items() if k not in da}
    both = [(k, da[k], db[k]) for k in da if k in db]

    lines = ['=' * 52, f"O列渠道差异对比: {name_a} vs {name_b}", '-' * 52]
    lines.append(f"仅 {name_a} 有: {len(only_a)} 行 | 仅 {name_b} 有: {len(only_b)} 行 | 共并行: {len(both)} 行")
    if only_a:
        lines.append('-' * 52)
        lines.append(f"仅 {name_a} 有（前20行）:")
        for k, (mid, o) in list(only_a.items())[:20]:
            lines.append(f"  行{k} (ID:{mid}): {o}")
    if only_b:
        lines.append('-' * 52)
        lines.append(f"仅 {name_b} 有（前20行）:")
        for k, (mid, o) in list(only_b.items())[:20]:
            lines.append(f"  行{k} (ID:{mid}): {o}")
    lines.append('=' * 52)
    report = '\n'.join(lines)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"{name_a}_agent_diff_{timestamp}.csv")
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = _csv.writer(f)
        w.writerow(['行号', '消息ID', f'{name_a}渠道', f'{name_b}渠道', '差异类型'])
        for k, (mid, o) in only_a.items():
            w.writerow([k, mid, o, '', '仅A有'])
        for k, (mid, o) in only_b.items():
            w.writerow([k, mid, '', o, '仅B有'])
        for k, (mid_a, oa), (mid_b, ob) in both:
            if oa != ob:
                w.writerow([k, mid_a, oa, ob, '渠道不同'])
    return path, report


def export_columns_csv(ws, out_dir, source_name, columns='', keyword=''):
    """通用导出指定列为CSV；columns为1-based列号逗号分隔（如 '1,8,15'）或表头名，keyword过滤行"""
    import csv as _csv

    header = [c.value for c in ws[1]]
    idxs = []
    if columns:
        for part in str(columns).split(','):
            part = part.strip()
            if not part:
                continue
            if part.isdigit():
                n = int(part)
                if 1 <= n <= len(header):
                    idxs.append(n)
            else:
                for i, h in enumerate(header, start=1):
                    if h is not None and part.lower() == str(h).strip().lower():
                        idxs.append(i)
                        break
    if not idxs:
        idxs = list(range(1, len(header) + 1))
    idxs = sorted(set(idxs))

    rows = []
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        if keyword:
            kw = str(keyword)
            if not any(v is not None and kw in str(v) for v in vals):
                continue
        rows.append([vals[i - 1] for i in idxs])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"{source_name}_agent_export_csv_{timestamp}.csv")
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = _csv.writer(f)
        w.writerow([header[i - 1] for i in idxs])
        w.writerows(rows)
    return path, len(rows)


# ---------- 文件处理 Agent（Phase 2）原子工具集 ----------

def _ts():
    """时间戳字符串（输出文件命名用）"""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _new_path(out_dir, base, suffix, ext):
    """生成不覆盖原文件的新输出路径"""
    return os.path.join(out_dir, f"{base}_agent_{suffix}_{_ts()}{ext}")


def detect_file_type(path):
    """识别文件类型（扩展名映射）"""
    ext = os.path.splitext(path)[1].lower()
    mapping = {
        '.xlsx': 'xlsx', '.xls': 'xls', '.csv': 'csv', '.txt': 'txt',
        '.json': 'json', '.xml': 'xml', '.pdf': 'pdf', '.docx': 'docx',
        '.pptx': 'pptx', '.jpg': 'jpg', '.jpeg': 'jpg', '.png': 'png', '.zip': 'zip',
    }
    return mapping.get(ext, ext.lstrip('.') or 'unknown')


def _resolve_column(ws, column):
    """将列引用（数字/字母/表头名）解析为 1-based 索引"""
    if isinstance(column, int):
        return column
    s = str(column).strip()
    if not s:
        raise ValueError("列引用为空")
    if s.isdigit():
        return int(s)
    if s.isalpha() and len(s) <= 3:
        from openpyxl.utils import column_index_from_string
        try:
            return column_index_from_string(s.upper())
        except Exception:
            pass
    header = [c.value for c in ws[1]]
    for i, h in enumerate(header, start=1):
        if h is not None and str(h).strip() == s:
            return i
    raise ValueError(f"无法解析列引用: {column}")


def inspect_file(files, params, out_dir):
    """文件基础信息（filename/type/size/sheets/rows/columns）"""
    path = files[0]
    t = detect_file_type(path)
    info = {'filename': os.path.basename(path), 'type': t, 'size': os.path.getsize(path)}
    if t in ('xlsx', 'xls'):
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        info['sheets'] = wb.sheetnames
        info['rows'] = ws.max_row
        info['columns'] = ws.max_column
        wb.close()
    elif t == 'csv':
        with open(path, encoding='utf-8-sig', errors='replace') as fh:
            first = fh.readline()
        info['rows'] = None
        info['columns'] = len([c for c in first.split(',')]) if first else 0
    return [], json.dumps(info, ensure_ascii=False)


def read_excel(files, params, out_dir):
    """读取 Excel：表头 + 前 N 行样例"""
    path = files[0]
    limit = int(params.get('limit', 5) or 5)
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    title = ws.title
    max_row = ws.max_row
    max_col = ws.max_column
    header = []
    try:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        pass
    sample = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=limit + 1)):
        sample.append([str(c.value or '')[:50] for c in row])
    wb.close()
    lines = [f"文件: {os.path.basename(path)} | Sheet: {title} | 行: {max_row} | 列: {max_col}"]
    lines.append("表头: " + ', '.join(str(h) if h is not None else '' for h in header))
    for i, r in enumerate(sample, start=2):
        lines.append(f"行{i}: " + ' | '.join(r))
    return [], '\n'.join(lines)


def inspect_excel(files, params, out_dir):
    """Excel 结构分析（sheets/rows/columns/headers 带字母）"""
    path = files[0]
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    info = {'sheets': wb.sheetnames, 'rows': ws.max_row, 'columns': ws.max_column}
    header = []
    try:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        pass
    wb.close()
    headers = [str(h) if h is not None else '' for h in header]
    col_lines = ', '.join(f"{get_column_letter(i + 1)}:{headers[i]}" for i in range(len(headers)) if headers[i])
    return [], (f"Sheet: {info['sheets']} | 行: {info['rows']} | 列: {info['columns']}\n列结构: {col_lines}")


def filter_excel(files, params, out_dir):
    """按列过滤行 → 新 Excel（operator: equals/not_equals/contains/gte/lte/gt/lt）"""
    from openpyxl import Workbook
    path = files[0]
    column = params.get('column', 1)
    operator = str(params.get('operator', 'equals') or 'equals')
    value = params.get('value', '')
    wb = load_workbook(path)
    ws = wb.active
    col_idx = _resolve_column(ws, column)
    out = Workbook()
    ows = out.active
    ows.append([c.value for c in ws[1]])
    cnt = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx - 1].value
        cell_s = '' if cell is None else str(cell)
        val_s = '' if value is None else str(value)
        match = False
        try:
            if operator in ('equals', '=='):
                match = cell_s == val_s
            elif operator in ('not_equals', '!='):
                match = cell_s != val_s
            elif operator in ('contains', 'like'):
                match = val_s in cell_s
            elif operator == 'gte':
                match = float(cell_s) >= float(val_s)
            elif operator == 'lte':
                match = float(cell_s) <= float(val_s)
            elif operator == 'gt':
                match = float(cell_s) > float(val_s)
            elif operator == 'lt':
                match = float(cell_s) < float(val_s)
            else:
                match = cell_s == val_s
        except (ValueError, TypeError):
            match = False
        if match:
            ows.append([c.value for c in row])
            cnt += 1
    base = os.path.splitext(os.path.basename(path))[0]
    dest = _new_path(out_dir, base, 'filter', '.xlsx')
    out.save(dest)
    wb.close()
    return [dest], f"过滤完成: {cnt} 行（列{column} {operator} {value}）-> {os.path.basename(dest)}"


def sort_excel(files, params, out_dir):
    """按列排序 → 新 Excel"""
    from openpyxl import Workbook
    path = files[0]
    column = params.get('column', 1)
    desc = str(params.get('desc', 'false')).lower() in ('true', '1', 'yes')
    wb = load_workbook(path)
    ws = wb.active
    col_idx = _resolve_column(ws, column)
    header = [c.value for c in ws[1]]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    try:
        rows.sort(key=lambda r: (r[col_idx - 1] is None, str(r[col_idx - 1] or '')), reverse=desc)
    except Exception:
        pass
    out = Workbook()
    ows = out.active
    ows.append(header)
    for r in rows:
        ows.append(r)
    base = os.path.splitext(os.path.basename(path))[0]
    dest = _new_path(out_dir, base, 'sorted', '.xlsx')
    out.save(dest)
    wb.close()
    return [dest], f"排序完成（列{column} {'降序' if desc else '升序'}）-> {os.path.basename(dest)}"


def group_excel(files, params, out_dir):
    """按列分组 → 每组分一个 Sheet"""
    from openpyxl import Workbook
    path = files[0]
    column = params.get('column', 1)
    wb = load_workbook(path)
    ws = wb.active
    col_idx = _resolve_column(ws, column)
    header = [c.value for c in ws[1]]
    groups = {}
    for row in ws.iter_rows(min_row=2):
        key = _cell_str(row, col_idx) or '(空)'
        groups.setdefault(key, []).append([c.value for c in row])
    out = Workbook()
    out.remove(out.active)
    for key, rows in groups.items():
        safe = str(key).replace('/', '_')[:30] or 'empty'
        ows = out.create_sheet(safe)
        ows.append(header)
        for r in rows:
            ows.append(r)
    base = os.path.splitext(os.path.basename(path))[0]
    dest = _new_path(out_dir, base, 'grouped', '.xlsx')
    out.save(dest)
    wb.close()
    return [dest], f"分组完成: {len(groups)} 组 -> {os.path.basename(dest)}"


def merge_excel(files, params, out_dir):
    """多文件合并为一个 Excel（保留首个文件表头）"""
    from openpyxl import Workbook
    merged = Workbook()
    mws = merged.active
    for fi, path in enumerate(files):
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if fi > 0 and ri == 0:
                continue  # 跳过后续文件表头
            if any(v is not None and str(v).strip() for v in row):
                mws.append(list(row))
        wb.close()
    base = os.path.splitext(os.path.basename(files[0]))[0]
    dest = _new_path(out_dir, base, 'merged', '.xlsx')
    merged.save(dest)
    return [dest], f"合并 {len(files)} 个文件 -> {os.path.basename(dest)}"


def split_excel(files, params, out_dir):
    """按行数拆分 → 多个 Sheet"""
    from openpyxl import Workbook
    path = files[0]
    rows_per = int(params.get('rows_per_sheet', 500) or 500)
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)
            if any(c.value is not None and str(c.value).strip() for c in row)]
    out = Workbook()
    out.remove(out.active)
    chunk = 0
    for i in range(0, len(rows), rows_per):
        chunk += 1
        ows = out.create_sheet(f"Part{chunk}")
        ows.append(header)
        for r in rows[i:i + rows_per]:
            ows.append(r)
    base = os.path.splitext(os.path.basename(path))[0]
    dest = _new_path(out_dir, base, 'split', '.xlsx')
    out.save(dest)
    wb.close()
    return [dest], f"拆分为 {chunk} 个Sheet（每{rows_per}行）-> {os.path.basename(dest)}"


# ---- 符号差异检测（find 匹配预检：全角/半角、中英文括号等） ----

_BRACKET_PAIRS = [
    ('（', '）'), ('(', ')'), ('【', '】'), ('[', ']'),
    ('「', '」'), ('『', '』'), ('《', '》'), ('<', '>'),
]

_PUNCT_MAP = {  # 全角 → 半角
    '（': '(', '）': ')', '【': '[', '】': ']',
    '《': '<', '》': '>', '，': ',', '。': '.', '、': ',',
    '：': ':', '；': ';', '！': '!', '？': '?',
    '“': '"', '”': '"', '‘': "'", '’': "'", '　': ' ',
}


def _symbol_variants(text, max_variants=16):
    """生成符号归一化变体（括号对互换、全角↔半角标点），不含原值、去重

    示例：'【反向】' → '(反向)'、'[反向]'、'（反向）' 等；
    最多生成 max_variants 个变体，防止组合爆炸。
    """
    if not text:
        return []
    variants = set()
    # 1) 成对括号整体互换：'【X】' → '(X)' / '[X]' / '（X）' ...
    for open_, close in _BRACKET_PAIRS:
        if open_ not in text and close not in text:
            continue
        for t_open, t_close in _BRACKET_PAIRS:
            if t_open == open_ and t_close == close:
                continue
            v = text.replace(open_, t_open).replace(close, t_close)
            if v != text:
                variants.add(v)
    # 2) 全角↔半角 单字符标点互换
    for c, t in _PUNCT_MAP.items():
        if c in text:
            variants.add(text.replace(c, t))
    for t, c in _PUNCT_MAP.items():
        if t in text:
            variants.add(text.replace(t, c))
    variants.discard(text)
    return sorted(variants)[:max_variants]


def check_find_in_column(path, column, find):
    """检查 find 在指定列是否存在；不存在时探测符号变体并统计次数。

    返回 {'found': bool, 'count': int, 'variants': [{'find': v, 'count': n}, ...]}
    variants 按出现次数降序排列（次数最多的候选优先）。
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    col = _resolve_column(ws, column)
    variants = _symbol_variants(find)
    counts = {find: 0}
    for v in variants:
        counts[v] = 0
    for row in ws.iter_rows(min_row=2):  # read_only 逐行，仅扫目标列
        cell = row[col - 1]
        if cell.value is None:
            continue
        s = str(cell.value)
        for k in list(counts):
            if k in s:
                counts[k] += 1
    wb.close()
    if counts[find] > 0:
        return {'found': True, 'count': counts[find], 'variants': []}
    hits = [{'find': v, 'count': counts[v]} for v in variants if counts[v] > 0]
    hits.sort(key=lambda x: x['count'], reverse=True)
    return {'found': False, 'count': 0, 'variants': hits}


def modify_excel(files, params, out_dir):
    """通用修改（ops: [{type:replace,column,find,replace}, {type:delete_rows,column,equals}]）

    params.save_mode: 'new'（默认，另存新文件）| 'overwrite'（直接覆盖原文件，不可逆）
    """
    path = files[0]
    ops = params.get('ops') or []
    save_mode = str(params.get('save_mode') or 'new').strip().lower()
    wb = load_workbook(path)
    ws = wb.active
    for op in ops:
        if not isinstance(op, dict):
            continue
        op_type = op.get('type')
        if op_type == 'replace':
            col = _resolve_column(ws, op.get('column', 1))
            find = str(op.get('find', ''))
            repl = str(op.get('replace', ''))
            for row in ws.iter_rows(min_row=2):
                cell = row[col - 1]
                if cell.value is not None and find in str(cell.value):
                    cell.value = str(cell.value).replace(find, repl)
        elif op_type == 'delete_rows':
            col = _resolve_column(ws, op.get('column', 1))
            eq = str(op.get('equals', ''))
            for row in list(ws.iter_rows(min_row=2)):
                if _cell_str(row, col) == eq:
                    ws.delete_rows(row[0].row)
    if save_mode == 'overwrite':
        wb.save(path)
        wb.close()
        return [path], f"修改完成 -> {os.path.basename(path)}（已覆盖原文件）"
    base = os.path.splitext(os.path.basename(path))[0]
    dest = _new_path(out_dir, base, 'modified', '.xlsx')
    wb.save(dest)
    wb.close()
    return [dest], f"修改完成 -> {os.path.basename(dest)}"


def export_excel(files, params, out_dir):
    """通用导出（复制为新 Excel，不覆盖源文件）"""
    import shutil
    outputs = []
    for path in files:
        base, ext = os.path.splitext(os.path.basename(path))
        dest = _new_path(out_dir, base, 'export', ext)
        shutil.copy2(path, dest)
        outputs.append(dest)
    return outputs, f"导出 {len(outputs)} 个文件完成"


def read_text(files, params, out_dir):
    """读取文本文件前 N 行"""
    path = files[0]
    limit = int(params.get('limit', 50) or 50)
    lines = []
    with open(path, encoding='utf-8', errors='replace') as fh:
        for i, line in enumerate(fh):
            if i >= limit:
                break
            lines.append(line.rstrip('\n'))
    return [], f"文件 {os.path.basename(path)} 前{len(lines)}行:\n" + '\n'.join(lines)


def search_text(files, params, out_dir):
    """文本搜索（返回命中行数与样例）"""
    path = files[0]
    kw = str(params.get('keyword', '') or params.get('find', '') or '')
    limit = int(params.get('limit', 20) or 20)
    hits = 0
    samples = []
    with open(path, encoding='utf-8', errors='replace') as fh:
        for ln, line in enumerate(fh, start=1):
            if kw and kw in line:
                hits += 1
                if len(samples) < limit:
                    samples.append(f"{ln}: {line.strip()[:120]}")
    report = f"命中 {hits} 行（关键词: {kw}）"
    if samples:
        report += '\n' + '\n'.join(samples)
    return [], report


def replace_text(files, params, out_dir):
    """文本替换 → 新文件（仅支持文本文件；Excel 请用 modify_excel，避免损坏二进制）"""
    path = files[0]
    t = detect_file_type(path)
    if t in ('xlsx', 'xls'):
        raise ValueError("replace_text 仅支持文本文件（txt/csv），Excel 文件请使用 modify_excel")
    find = str(params.get('find', ''))
    repl = str(params.get('replace', ''))
    with open(path, encoding='utf-8', errors='replace') as fh:
        content = fh.read()
    n = content.count(find)
    content = content.replace(find, repl)
    base, ext = os.path.splitext(os.path.basename(path))
    dest = _new_path(out_dir, base, 'replaced', ext)
    with open(dest, 'w', encoding='utf-8') as fh:
        fh.write(content)
    return [dest], f"替换 {n} 处 -> {os.path.basename(dest)}"


def copy_file(files, params, out_dir):
    """复制文件为新名（不覆盖原文件）"""
    import shutil
    path = files[0]
    base, ext = os.path.splitext(os.path.basename(path))
    dest = _new_path(out_dir, base, 'copy', ext)
    shutil.copy2(path, dest)
    return [dest], f"已复制 -> {os.path.basename(dest)}"


def validate_output(files, params, out_dir):
    """输出文件校验（存在/可打开/行数列数/非空）"""
    report_lines = []
    for path in files:
        if not os.path.exists(path):
            report_lines.append(f"✗ 不存在: {path}")
            continue
        size = os.path.getsize(path)
        t = detect_file_type(path)
        try:
            if t in ('xlsx', 'xls'):
                wb = load_workbook(path, read_only=True)
                ws = wb.active
                report_lines.append(f"✓ {os.path.basename(path)}: 可打开, 行{ws.max_row} 列{ws.max_column}, {size}字节")
                wb.close()
            elif t == 'csv':
                with open(path, encoding='utf-8-sig') as fh:
                    rows = sum(1 for _ in fh)
                report_lines.append(f"✓ {os.path.basename(path)}: CSV {rows}行, {size}字节")
            else:
                report_lines.append(f"✓ {os.path.basename(path)}: {size}字节")
        except Exception as e:
            report_lines.append(f"✗ 打不开: {os.path.basename(path)}: {str(e)}")
    return [], '\n'.join(report_lines) if report_lines else "无输出文件可校验"


def statistics_excel(files, params, out_dir):
    """统计报表（高级工具 statistics）"""
    path = files[0]
    wb = load_workbook(path)
    ws = wb.active
    report = generate_statistics_report(ws)
    wb.close()
    base = os.path.splitext(os.path.basename(path))[0]
    txt = _new_path(out_dir, base, 'statistics', '.txt')
    with open(txt, 'w', encoding='utf-8') as fh:
        fh.write(report)
    return [txt], report


def export_failed_tool(files, params, out_dir):
    """导出失败项（高级工具 export_failed）"""
    scope = str(params.get('scope', 'all') or 'all')
    fmt = str(params.get('format', 'excel') or 'excel')
    outputs = []
    for path in files:
        wb = load_workbook(path)
        ws = wb.active
        base = os.path.splitext(os.path.basename(path))[0]
        p, cnt = export_failed_rows(ws, out_dir, base, scope=scope, fmt=fmt)
        wb.close()
        outputs.append(p)
    return outputs, f"导出失败项完成（{len(outputs)} 个文件）"


def export_csv_tool(files, params, out_dir):
    """通用列导出 CSV（高级工具 export_csv）"""
    columns = str(params.get('columns', '') or '')
    keyword = str(params.get('filter', '') or '')
    outputs = []
    for path in files:
        wb = load_workbook(path)
        ws = wb.active
        base = os.path.splitext(os.path.basename(path))[0]
        p, cnt = export_columns_csv(ws, out_dir, base, columns=columns, keyword=keyword)
        wb.close()
        outputs.append(p)
    return outputs, f"CSV导出完成（{len(outputs)} 个文件）"


def compare_excel(files, params, out_dir):
    """两文件 O 列对比（高级工具 diff）"""
    src = int(params.get('file_index', 0) or 0)
    tgt = int(params.get('other_file_index', 1) or 1)
    if len(files) < 2:
        raise ValueError("compare 需要至少两个文件")
    if src >= len(files) or tgt >= len(files) or src == tgt:
        raise ValueError(f"非法文件索引: {src}/{tgt}（共{len(files)}个）")
    wb_a = load_workbook(files[src])
    ws_a = wb_a.active
    wb_b = load_workbook(files[tgt])
    ws_b = wb_b.active
    name_a = os.path.splitext(os.path.basename(files[src]))[0]
    name_b = os.path.splitext(os.path.basename(files[tgt]))[0]
    p, report = diff_worksheets(ws_a, ws_b, out_dir, name_a, name_b)
    wb_a.close()
    wb_b.close()
    return [p], report


# 工具注册表（name → 元数据 + 统一 run 调用 + 参数 schema）
# input_schema: {参数名: {type, enum?, required?, default?, desc?}} —— 供 planner 校验参数完整性，驱动澄清循环
TOOL_REGISTRY = {
    'statistics': {'description': '统计Excel匹配情况（占比/渠道/失败原因）', 'risk_level': 'low', 'deterministic': True,
                   'input_schema': {}, 'run': statistics_excel},
    'export_failed': {'description': '导出失败项为新文件', 'risk_level': 'low', 'deterministic': True,
                      'input_schema': {
                          'scope': {'type': 'str', 'enum': ['reverse', 'forward', 'all'], 'required': False, 'default': 'all', 'desc': '失败范围'},
                          'format': {'type': 'str', 'enum': ['excel', 'csv'], 'required': False, 'default': 'excel', 'desc': '导出格式'},
                      }, 'run': export_failed_tool},
    'diff': {'description': '对比两个Excel的O列渠道差异', 'risk_level': 'low', 'deterministic': True,
             'input_schema': {
                 'file_index': {'type': 'int', 'required': False, 'default': 0, 'desc': '源文件序号（从0开始）'},
                 'other_file_index': {'type': 'int', 'required': False, 'default': 1, 'desc': '对比文件序号（从0开始）'},
             }, 'run': compare_excel},
    'export_csv': {'description': '通用导出指定列为CSV', 'risk_level': 'low', 'deterministic': True,
                   'input_schema': {
                       'columns': {'type': 'str', 'required': True, 'desc': '列号或列名，逗号分隔，如 "1,8,15"'},
                       'filter': {'type': 'str', 'required': False, 'default': '', 'desc': '过滤关键词'},
                   }, 'run': export_csv_tool},
    'retry': {'description': '对失败行重新调用接口处理', 'risk_level': 'high', 'deterministic': False,
              'input_schema': {
                  'mode': {'type': 'str', 'enum': ['reverse', 'forward'], 'required': True, 'desc': '重试方向'},
              }, 'run': None},
    'inspect_file': {'description': '文件基础信息', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': inspect_file},
    'detect_file_type': {'description': '识别文件类型', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': detect_file_type},
    'read_file': {'description': '统一读取文件（按类型分发）', 'risk_level': 'low', 'deterministic': True,
                  'input_schema': {'limit': {'type': 'int', 'required': False, 'default': 5, 'desc': '读取行数'}}, 'run': None},
    'copy_file': {'description': '复制文件为新名', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': copy_file},
    'inspect_excel': {'description': 'Excel结构分析', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': inspect_excel},
    'read_excel': {'description': '读取Excel表头与前N行', 'risk_level': 'low', 'deterministic': True,
                   'input_schema': {'limit': {'type': 'int', 'required': False, 'default': 5, 'desc': '读取行数'}}, 'run': read_excel},
    'filter_excel': {'description': '按列过滤行', 'risk_level': 'low', 'deterministic': True,
                     'input_schema': {
                         'column': {'type': 'str', 'required': True, 'desc': '列号（如 P）或列名'},
                         'operator': {'type': 'str', 'enum': ['equals', 'not_equals', 'contains', 'gte', 'lte', 'gt', 'lt'], 'required': False, 'default': 'equals', 'desc': '过滤操作符'},
                         'value': {'type': 'str', 'required': False, 'default': '', 'desc': '过滤值'},
                     }, 'run': filter_excel},
    'sort_excel': {'description': '按列排序', 'risk_level': 'low', 'deterministic': True,
                   'input_schema': {
                       'column': {'type': 'str', 'required': True, 'desc': '列号（如 P）或列名'},
                       'desc': {'type': 'bool', 'required': False, 'default': False, 'desc': '是否降序'},
                   }, 'run': sort_excel},
    'group_excel': {'description': '按列分组为多个Sheet', 'risk_level': 'low', 'deterministic': True,
                    'input_schema': {
                        'column': {'type': 'str', 'required': True, 'desc': '分组列号（如 O）或列名'},
                    }, 'run': group_excel},
    'merge_excel': {'description': '多文件合并', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': merge_excel},
    'split_excel': {'description': '按行数拆分', 'risk_level': 'low', 'deterministic': True,
                    'input_schema': {
                        'rows_per_sheet': {'type': 'int', 'required': False, 'default': 500, 'desc': '每个Sheet行数'},
                    }, 'run': split_excel},
    'modify_excel': {'description': '通用修改（replace/delete_rows）', 'risk_level': 'medium', 'deterministic': True,
                     'input_schema': {
                         'ops': {'type': 'list', 'required': True, 'desc': '操作列表，如 [{type:replace,column,find,replace}]'},
                     }, 'run': modify_excel},
    'export_excel': {'description': '导出为新Excel（不覆盖源）', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': export_excel},
    'read_text': {'description': '读取文本文件', 'risk_level': 'low', 'deterministic': True,
                  'input_schema': {'limit': {'type': 'int', 'required': False, 'default': 50, 'desc': '读取行数'}},
                  'supports': {'txt', 'csv'}, 'run': read_text},
    'search_text': {'description': '文本关键词搜索', 'risk_level': 'low', 'deterministic': True,
                    'input_schema': {
                        'keyword': {'type': 'str', 'required': True, 'desc': '搜索关键词'},
                        'limit': {'type': 'int', 'required': False, 'default': 20, 'desc': '返回样例行数'},
                    }, 'supports': {'txt', 'csv'}, 'run': search_text},
    'replace_text': {'description': '文本替换为新文件', 'risk_level': 'low', 'deterministic': True,
                     'input_schema': {
                         'find': {'type': 'str', 'required': True, 'desc': '查找内容'},
                         'replace': {'type': 'str', 'required': False, 'default': '', 'desc': '替换为'},
                     }, 'supports': {'txt', 'csv'}, 'run': replace_text},
    'validate_output': {'description': '输出文件校验', 'risk_level': 'low', 'deterministic': True, 'input_schema': {}, 'run': validate_output},
}


def validate_tool_params(tool, params):
    """校验工具参数是否满足 input_schema；返回 (is_ok, missing_questions)

    missing_questions 为可向用户提出的澄清问题列表（缺失/非法参数 → 需要用户补充）。
    """
    schema = TOOL_REGISTRY.get(tool, {}).get('input_schema', {})
    params = params or {}
    questions = []
    for name, spec in schema.items():
        val = params.get(name)
        if spec.get('required') and (val is None or val == '' or val == []):
            enum_hint = f"（可选值: {'/'.join(str(e) for e in spec['enum'])}）" if spec.get('enum') else ''
            desc = spec.get('desc') or name
            questions.append(f"请提供参数「{name}」：{desc}{enum_hint}")
    return (len(questions) == 0), questions


def collect_output_meta(output_files):
    """本地读取输出文件结构 + 无表头列数据示例（供反思质检对比），不含绝对路径"""
    metas = []
    for path in output_files or []:
        try:
            t = detect_file_type(path)
            m = {'basename': os.path.basename(path), 'type': t}
            if t in ('xlsx', 'xls'):
                wb = load_workbook(path, read_only=True)
                ws = wb.active
                m['sheets'] = wb.sheetnames
                m['rows'] = ws.max_row
                m['columns'] = ws.max_column
                try:
                    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    m['headers'] = [str(h) if h is not None else '' for h in header]
                except StopIteration:
                    m['headers'] = []
                try:
                    data_cols, empty_cols, data_samples = _detect_headerless_data_columns(ws, m['headers'])
                    m['data_columns'] = data_cols
                    m['empty_columns'] = empty_cols
                    m['data_samples'] = data_samples
                except Exception:
                    pass
                wb.close()
            elif t == 'csv':
                with open(path, encoding='utf-8-sig', errors='replace') as fh:
                    first = fh.readline()
                    rows = 1 + sum(1 for _ in fh)
                m['headers'] = [c.strip() for c in first.split(',')] if first else []
                m['rows'] = rows
                m['columns'] = len(m['headers'])
            metas.append(m)
        except Exception as e:
            metas.append({'basename': os.path.basename(path), 'type': 'unknown', 'error': str(e)})
    return metas


def _column_structure_text(metas):
    """将文件元数据格式化为列结构文本（供意图识别 prompt / 澄清交互展示）。

    完整展示三类列：有表头列、有数据但无表头的列（data_columns）、空列（empty_columns）。
    避免"有数据无表头"的列被隐藏导致模型误判列不存在。
    """
    from openpyxl.utils import get_column_letter
    if not metas:
        return COLUMN_STRUCTURE
    m = metas[0]
    headers = m.get('headers') or []
    if not headers:
        return COLUMN_STRUCTURE
    parts = [f"文件[{i}]: {mm.get('basename')}" for i, mm in enumerate(metas)]
    cols = ', '.join(f"{get_column_letter(i + 1)}:{h}" for i, h in enumerate(headers) if h)
    parts.append(f"列结构: {cols}")
    data_cols = m.get('data_columns') or []
    if data_cols:
        samples = m.get('data_samples') or {}
        parts.append('数据列(无表头): ' + ', '.join(
            f"{c}(示例: {samples[c]})" if samples.get(c) else c for c in data_cols))
    empty_cols = m.get('empty_columns') or []
    if empty_cols:
        parts.append(f"空列: {', '.join(empty_cols)}")
    return '\n'.join(parts)


def _sample_content(paths, limit=3):
    """读取文件前 N 行全部列内容样本（用于反思内容对比），每列截断防超长"""
    lines = []
    for p in (paths or [])[:2]:
        try:
            t = detect_file_type(p)
            if t in ('xlsx', 'xls'):
                wb = load_workbook(p, read_only=True)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limit + 1, values_only=True)):
                    lines.append(f"[{os.path.basename(p)}] 行{i + 1}: " + ' | '.join(
                        str(v)[:12] if v is not None else '' for v in row))
                wb.close()
            else:
                with open(p, encoding='utf-8-sig', errors='replace') as fh:
                    for i in range(limit + 1):
                        line = fh.readline()
                        if not line:
                            break
                        lines.append(f"[{os.path.basename(p)}] 行{i + 1}: " + line.strip()[:120])
        except Exception as e:
            lines.append(f"[{os.path.basename(p)}] 读取失败: {str(e)}")
    return '\n'.join(lines)


def _detect_headerless_data_columns(ws, headers):
    """检测"表头为空但数据行有值"的列。

    返回 (data_cols, empty_cols, data_samples)：
    - data_cols: 无表头但有数据的列字母（如 ['O','Q','R','S']）
    - empty_cols: 无表头且无数据的列字母
    - data_samples: {列字母: 首行示例值}（供列语义提示，截断 30 字符）
    仅扫描前几行数据（read_only 大文件友好）。
    """
    from openpyxl.utils import get_column_letter
    data_cols = []
    data_samples = {}
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            for i, v in enumerate(row):
                if i < len(headers) and not headers[i] and v is not None and str(v).strip():
                    letter = get_column_letter(i + 1)
                    if letter not in data_cols:
                        data_cols.append(letter)
                        data_samples[letter] = str(v).strip()[:30]
            break
    data_set = set(data_cols)
    empty_cols = [get_column_letter(i + 1) for i, h in enumerate(headers)
                  if not h and get_column_letter(i + 1) not in data_set]
    return data_cols, empty_cols, data_samples


def read_file(files, params, out_dir):
    """read_file 统一入口：按类型分发"""
    path = files[0]
    t = detect_file_type(path)
    if t in ('xlsx', 'xls'):
        return read_excel(files, params, out_dir)
    return read_text(files, params, out_dir)


TOOL_REGISTRY['read_file']['run'] = read_file


def _detect_type_run(files, params, out_dir):
    return [], detect_file_type(files[0])


TOOL_REGISTRY['detect_file_type']['run'] = _detect_type_run


# ---------- 文件处理 Agent（Phase 2）LangGraph 编排 ----------

from typing import TypedDict, Optional


class FileAgentState(TypedDict, total=False):
    user_request: str
    input_files: list
    file_metadata: list
    intent: dict
    plan: dict
    route: str
    generated_code: str
    code_description: str
    code_validation: dict
    task_id: str
    execution_result: dict
    output_files: list
    error: Optional[str]
    logs: list
    status: str
    # ---- 新增：对话与理解（澄清/理解确认循环） ----
    conversation: list          # 多轮对话历史 [{role:'user'|'assistant', content}]
    understanding: str          # Agent 对需求的理解复述（自然语言）
    clarify_questions: list     # 需要澄清的问题列表
    clarify_answers: dict       # 用户对澄清问题的回答 {问题 -> 回答}
    clarify_rounds: int         # 澄清轮次计数（防无限循环，上限 3）
    understanding_ok: bool      # 用户是否确认理解正确
    user_revision: str          # 用户对需求的修正/补充（触发重新理解）
    # ---- 新增：结果反思与用户反馈（反思修复/反馈循环） ----
    verification: dict          # 结果自检 {ok, reason, checks:[{name, passed, detail}]}
    repair_rounds: int          # 已修复轮次计数
    max_repair_rounds: int      # 最大修复轮次（默认 2）
    repair_feedback: str        # 反思意见（供重新生成代码）
    final_feedback: str         # 用户最终反馈（修改意见，触发下一轮）
    finished: bool              # 用户是否确认满意
    input_snapshot: str         # 覆盖原文件前的原始内容快照路径（供自检对比修改前后）


def build_agent_graph(app):
    """构建 LangGraph：inspect → intent → 澄清/理解确认 → planner → route → (execute → validate → verify → summary)"""
    from langgraph.graph import StateGraph, START, END
    from langgraph.checkpoint.memory import InMemorySaver
    from langgraph.types import interrupt, Command

    def inspect_files(state):
        logs = list(state.get('logs') or [])
        metas = []
        for path in state['input_files']:
            try:
                t = detect_file_type(path)
                m = {'path': path, 'basename': os.path.basename(path), 'type': t}
                if t in ('xlsx', 'xls'):
                    wb = load_workbook(path, read_only=True)
                    ws = wb.active
                    m['sheets'] = wb.sheetnames
                    m['rows'] = ws.max_row
                    m['columns'] = ws.max_column
                    try:
                        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                        m['headers'] = [str(h) if h is not None else '' for h in header]
                    except StopIteration:
                        m['headers'] = []
                    # 数据列采集：表头为空但数据行有值的列（如 O/Q/R/S），避免被误判为不存在
                    data_cols, empty_cols, data_samples = _detect_headerless_data_columns(ws, m.get('headers') or [])
                    m['data_columns'] = data_cols
                    m['empty_columns'] = empty_cols
                    m['data_samples'] = data_samples
                    wb.close()
                elif t == 'csv':
                    with open(path, encoding='utf-8-sig', errors='replace') as fh:
                        first = fh.readline()
                    m['headers'] = [c.strip() for c in first.split(',')] if first else []
                    m['columns'] = len(m['headers'])
                    m['data_columns'] = []
                    m['empty_columns'] = []
                    m['data_samples'] = {}
                else:
                    m['headers'] = []
                    m['data_columns'] = []
                    m['empty_columns'] = []
                    m['data_samples'] = {}
                metas.append(m)
                logs.append(f"inspect: {m['basename']} type={t} rows={m.get('rows')} cols={m.get('columns')}"
                            + (f" 数据列:{','.join(m['data_columns'])}" if m.get('data_columns') else ''))
            except Exception as e:
                metas.append({'path': path, 'basename': os.path.basename(path), 'type': 'unknown', 'error': str(e)})
                logs.append(f"inspect失败 {os.path.basename(path)}: {str(e)}")
        return {'file_metadata': metas, 'logs': logs}

    def call_dify_intent(state):
        logs = list(state.get('logs') or [])
        api_url = app.agent_api_url_entry.get().strip()
        api_key = app.agent_api_key_entry.get().strip()
        col_struct = _column_structure_text(state.get('file_metadata') or [])
        history_section = _format_history_section(state.get('conversation') or [])
        prompt = INTENT_PROMPT_TEMPLATE.format(
            column_structure=col_struct, request=state['user_request'], history_section=history_section)
        answer = app.call_intent_api(api_url, api_key, prompt)
        logs.append(f"意图返回(完整): {answer}")
        meta = app.parse_intent_meta(answer)
        task, params = meta['task'], meta['params']
        if task == 'custom':
            intent = {'task': 'custom', 'params': params, 'can_use_builtin_tool': False,
                      'reason': meta.get('reason', '')}
        else:
            intent = {'task': task, 'params': params, 'can_use_builtin_tool': True}
        # 新增：理解复述与澄清问题（Agent Loop ①/② 的数据来源）
        intent['understanding'] = meta.get('understanding', '')
        intent['clarify_questions'] = meta.get('clarify_questions', [])
        intent['raw_answer'] = answer  # 原始返回存入状态（供交互区/落盘）
        if meta.get('truncated'):
            logs.append("⚠ 意图识别返回疑似被截断（JSON不完整），建议检查 Dify 意图应用 LLM 节点 max_tokens 或回答节点输出")
        logs.append(f"意图识别: {task} params={json.dumps(params, ensure_ascii=False)}")
        return {'intent': intent, 'logs': logs}

    def decide_clarify(state):
        """Loop①条件路由：有澄清问题且尚未回答 → 进入澄清中断；否则继续到理解确认"""
        questions = (state.get('intent') or {}).get('clarify_questions') or []
        rounds = state.get('clarify_rounds') or 0
        if questions and not state.get('clarify_answers') and rounds < 3:
            return 'clarify'
        return 'proceed'

    def clarify_interrupt(state):
        """Loop①中断：暂停图执行，等待用户逐题回答澄清问题；恢复时写入 clarify_answers"""
        questions = (state.get('intent') or {}).get('clarify_questions') or []
        logs = list(state.get('logs') or [])
        logs.append(f"需要澄清 {len(questions)} 个问题，等待用户回答")
        answers = interrupt({'type': 'clarify', 'questions': questions})
        if not isinstance(answers, dict):
            answers = {}
        logs.append("已收到用户澄清回答")
        return {'clarify_answers': answers, 'clarify_rounds': (state.get('clarify_rounds') or 0) + 1,
                'status': 'planning', 'logs': logs}

    def update_context(state):
        """合并澄清答案 / 用户修正进 user_request（作为 Context 的一部分），并清空一次性状态"""
        logs = list(state.get('logs') or [])
        answers = state.get('clarify_answers') or {}
        revision = state.get('user_revision') or ''
        parts = [state.get('user_request', '')]
        if answers:
            parts.append('补充信息：' + '；'.join(f'{q}：{a}' for q, a in answers.items() if a))
        if revision:
            parts.append('用户修正：' + revision)
        merged = '\n'.join(parts).strip()
        logs.append("已更新需求上下文（含澄清/修正信息），重新分析需求")
        return {'user_request': merged, 'clarify_answers': None, 'user_revision': None, 'logs': logs}

    def show_understanding(state):
        """Loop②理解确认：展示 Agent 对需求的理解复述，中断等待用户确认/修正"""
        # 已确认过理解（如澄清第二轮回来）则直接放行，避免重复确认
        if state.get('understanding_ok'):
            return {'status': 'planning'}
        understanding = (state.get('intent') or {}).get('understanding') or ''
        # 兜底：understanding 为空（如 Dify 返回截断/旧格式）时用需求文本展示，避免空白
        if not understanding:
            understanding = f"根据你的需求处理文件：{state.get('user_request', '')[:120]}"
        logs = list(state.get('logs') or [])
        if understanding:
            logs.append(f"我的理解：{understanding}")
        resp = interrupt({'type': 'understanding', 'text': understanding})
        resp = resp or {}
        ok = bool(resp.get('ok', True))
        revision = str(resp.get('revision') or '').strip()
        result = {'understanding': understanding, 'understanding_ok': ok,
                  'status': 'planning', 'logs': logs}
        if not ok and revision:
            result['user_revision'] = revision
        return result

    def decide_understanding(state):
        """Loop②条件路由：理解正确 → planner；用户修正 → 重新理解"""
        if state.get('understanding_ok'):
            return 'proceed'
        return 'revision' if state.get('user_revision') else 'proceed'

    def planner(state):
        logs = list(state.get('logs') or [])
        intent = state['intent']
        task = intent['task']
        if task == 'custom':
            plan = {'goal': intent.get('reason', ''), 'strategy': 'dynamic_code', 'tool': 'custom',
                    'params': {}, 'risk_level': 'medium', 'need_confirmation': True}
        else:
            meta = TOOL_REGISTRY.get(task, {})
            params = intent.get('params') or {}
            ok, missing = validate_tool_params(task, params)
            if not ok:
                # 参数缺失 → 注入澄清问题，路由回澄清循环
                logs.append(f"计划: 参数不完整（{'; '.join(missing)}），需要用户澄清")
                return {
                    'plan': {'goal': meta.get('description', task), 'strategy': 'builtin', 'tool': task,
                             'params': params, 'risk_level': meta.get('risk_level', 'low'),
                             'need_confirmation': True, 'param_missing': True},
                    'intent': {**intent, 'clarify_questions': missing},
                    'logs': logs,
                }
            plan = {'goal': meta.get('description', task), 'strategy': 'builtin', 'tool': task,
                    'params': params, 'risk_level': meta.get('risk_level', 'low'),
                    'need_confirmation': True}
        logs.append(f"计划: strategy={plan['strategy']} tool={plan['tool']}")
        return {'plan': plan, 'logs': logs}

    def route_task(state):
        plan = state.get('plan') or {}
        if plan.get('param_missing'):
            return {'route': 'clarify'}
        return {'route': 'builtin' if plan.get('strategy') == 'builtin' else 'dynamic_code'}

    def execute_builtin(state):
        logs = list(state.get('logs') or [])
        tool = state['plan']['tool']
        files = state['input_files']
        params = state['plan'].get('params') or {}
        out_dir = os.path.dirname(os.path.abspath(files[0]))
        # 工具-文件类型校验：supports 声明的工具只允许匹配的文件类型
        supports = TOOL_REGISTRY.get(tool, {}).get('supports')
        if supports:
            for f in files:
                t = detect_file_type(f)
                if t not in supports:
                    hint = ('；Excel 文件的查找/替换/读取请使用 modify_excel 等 Excel 工具'
                            if t in ('xlsx', 'xls') else '')
                    msg = f"工具 {tool} 不支持文件类型 {t}（支持: {'/'.join(sorted(supports))}）{hint}"
                    logs.append(msg)
                    return {'execution_result': {'ok': False, 'error': msg}, 'error': msg,
                            'logs': logs, 'status': 'failed'}
        outputs = []
        snapshot_path = None
        try:
            if tool == 'retry':
                mode = str(params.get('mode', 'reverse') or 'reverse')
                for f in files:
                    wb, ws = app._load_ws(f)
                    base = os.path.splitext(os.path.basename(f))[0]
                    path, ok, fail = app.retry_failed_rows(ws, mode, out_dir, base)
                    wb.close()
                    outputs.append(path)
                    logs.append(f"重试({mode}): 成功{ok}/失败{fail} -> {os.path.basename(path)}")
            else:
                # 覆盖原文件（save_mode=overwrite）前保存快照：修改后输入=输出同一路径，自检无法自比，
                # 需用快照（修改前内容）与输出对比；快照放系统临时目录，由 OS 回收
                if (tool == 'modify_excel'
                        and str((params or {}).get('save_mode') or '').strip().lower() == 'overwrite'
                        and files and os.path.exists(files[0])):
                    fd, snapshot_path = tempfile.mkstemp(prefix='orig_snapshot_',
                                                         suffix=os.path.splitext(files[0])[1])
                    os.close(fd)
                    shutil.copy2(files[0], snapshot_path)
                outs, summary = TOOL_REGISTRY[tool]['run'](files, params, out_dir)
                outputs.extend(outs)
                logs.append(summary)
            result = {'execution_result': {'ok': True}, 'output_files': outputs, 'logs': logs}
            if snapshot_path:
                result['input_snapshot'] = snapshot_path
            return result
        except Exception as e:
            if snapshot_path and os.path.exists(snapshot_path):
                try:
                    os.remove(snapshot_path)
                except Exception:
                    pass
            logs.append(f"执行失败: {str(e)}")
            return {'execution_result': {'ok': False, 'error': str(e)}, 'error': str(e),
                    'logs': logs, 'status': 'failed'}

    def validate_output_node(state):
        logs = list(state.get('logs') or [])
        for o in state.get('output_files') or []:
            try:
                _, report = validate_output([o], {}, os.path.dirname(o))
                logs.append(f"输出校验: {os.path.basename(o)} -> {report[:120]}")
            except Exception as e:
                logs.append(f"输出校验失败: {os.path.basename(o)}: {str(e)}")
        return {'logs': logs}

    def result_summary(state):
        logs = list(state.get('logs') or [])
        logs.append(f"完成: 输出文件 {len(state.get('output_files') or [])} 个")
        return {'logs': logs, 'status': 'done'}

    def _read_sample_rows(state):
        """读取首个文件前 3 行数据作为样例（供代码生成）"""
        metas = state.get('file_metadata') or []
        if not metas:
            return ''
        path = metas[0].get('path')
        if not path or not os.path.exists(path):
            return ''
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            lines = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True)):
                lines.append(f"行{i + 1}: " + ' | '.join(str(v) if v is not None else '' for v in row))
            wb.close()
            return '\n'.join(lines)
        except Exception as e:
            return f"(样例读取失败: {str(e)})"

    def generate_code(state):
        logs = list(state.get('logs') or [])
        api_url = app.agent_codegen_url_entry.get().strip()
        api_key = app.agent_codegen_key_entry.get().strip()
        if not api_url or not api_key:
            logs.append("代码生成 API 未配置（Agent TAB 代码生成 API配置区）")
            return {'logs': logs, 'status': 'failed', 'error': '代码生成 API 未配置'}
        user_request = state['user_request']
        # 脱敏文件元数据：只留文件名与结构，去掉原始绝对路径，引导代码走 INPUT_DIR
        meta_safe = []
        for m in state.get('file_metadata') or []:
            safe = {k: v for k, v in m.items() if k != 'path'}
            safe['path'] = '<input_dir>/' + os.path.basename(m.get('path') or '')
            meta_safe.append(safe)
        file_metadata = json.dumps(meta_safe, ensure_ascii=False)
        sample_rows = _read_sample_rows(state)
        try:
            result = app.call_code_generator_api(
                api_url, api_key, user_request, file_metadata, sample_rows, EXECUTION_CONSTRAINTS)
            code = str(result.get('code', '') or '')
            if not code:
                keys = list(result.keys()) if isinstance(result, dict) else []
                logs.append(f"代码生成为空: keys={keys} preview={str(result)[:200]}")
                return {'logs': logs, 'status': 'failed', 'error': '代码生成为空'}
            logs.append(f"代码生成成功: {str(result.get('description', ''))[:200]}")
            return {
                'generated_code': code,
                'code_description': str(result.get('description', '') or ''),
                'logs': logs,
            }
        except Exception as e:
            logs.append(f"代码生成失败: {str(e)}")
            return {'logs': logs, 'status': 'failed', 'error': str(e)}

    def validate_code(state):
        logs = list(state.get('logs') or [])
        valid, issues = validate_generated_code(state.get('generated_code', ''))
        logs.append(f"代码校验: {'通过' if valid else '拒绝 ' + '; '.join(issues[:5])}")
        result = {'code_validation': {'valid': valid, 'issues': issues}, 'logs': logs}
        if not valid:
            # 注入拒绝原因（含软提示，供校验拒绝后的自动重生成反馈给 Dify）
            result['repair_feedback'] = '代码校验未通过：' + '；'.join(issues[:5])
        else:
            # 校验通过但存在软提示（如相对路径保存）→ 日志提示
            for h in issues:
                if h.startswith('[提示]'):
                    logs.append(f"提示: {h}")
        return result

    def sandbox_execute(state):
        logs = list(state.get('logs') or [])
        task_id = state.get('task_id', uuid.uuid4().hex)
        out_dir = os.path.dirname(os.path.abspath(state['input_files'][0]))
        ok, log, outputs = run_code_sandbox(
            state.get('generated_code', ''), state['input_files'], out_dir, task_id)
        logs.append(f"沙箱执行: {'成功' if ok else '失败'}")
        if log:
            logs.append(f"沙箱日志: {log[:500]}")
        if ok:
            for o in outputs:
                logs.append(f"输出: {o}")
            return {'execution_result': {'ok': True}, 'output_files': outputs, 'logs': logs}
        # 失败时附带代码摘要（前 8 行），便于人工排查
        code_head = '\n'.join((state.get('generated_code') or '').splitlines()[:8])
        if code_head:
            logs.append(f"代码摘要(前8行): {code_head[:500]}")
        return {'execution_result': {'ok': False, 'log': log}, 'error': log[:500],
                'logs': logs, 'status': 'failed'}

    def code_rejected(state):
        logs = list(state.get('logs') or [])
        issues = state.get('code_validation', {}).get('issues', [])
        logs.append(f"代码被拒绝，不执行: {'; '.join(issues[:5])}")
        return {'logs': logs, 'status': 'failed'}

    def verify_result(state):
        """Loop③结果自检：结构化校验（存在/可打开/非空）+ 语义反思（LLM，失败时降级）"""
        logs = list(state.get('logs') or [])
        checks = []
        outputs = state.get('output_files') or []
        for o in outputs:
            passed = os.path.exists(o)
            detail = '存在'
            if passed:
                try:
                    t = detect_file_type(o)
                    if t in ('xlsx', 'xls'):
                        wb = load_workbook(o, read_only=True)
                        rows = wb.active.max_row
                        wb.close()
                        passed = rows > 0
                        detail = f'可打开, {rows}行'
                    elif t == 'csv':
                        with open(o, encoding='utf-8-sig', errors='replace') as fh:
                            rows = sum(1 for _ in fh)
                        passed = rows > 0
                        detail = f'CSV {rows}行'
                    else:
                        detail = f'{os.path.getsize(o)}字节'
                except Exception as e:
                    passed = False
                    detail = f'打不开: {str(e)}'
            checks.append({'name': os.path.basename(o), 'passed': passed, 'detail': detail})
        if not outputs:
            checks.append({'name': '输出文件', 'passed': False, 'detail': '无输出文件'})
        structural_ok = all(c['passed'] for c in checks)
        # 语义反思：有输出文件才做（无输出时反思无意义，结构化失败已足够触发修复）
        reflection = None
        if not outputs:
            logs.append("无输出文件，跳过语义反思（结构化校验失败即可触发修复）")
        else:
            try:
                api_url = app.agent_api_url_entry.get().strip()
                api_key = app.agent_api_key_entry.get().strip()
                if api_url and api_key:
                    # 覆盖原文件模式：用执行前快照（修改前内容）作为输入样本，与输出（修改后）对比
                    snapshot = state.get('input_snapshot') or ''
                    input_src = ([snapshot] if snapshot and os.path.exists(snapshot)
                                 else (state.get('input_files') or []))
                    reflection = app.call_reflection_api(
                        api_url, api_key,
                        user_request=state.get('user_request', ''),
                        input_meta=json.dumps([{k: v for k, v in m.items() if k != 'path'}
                                               for m in (state.get('file_metadata') or [])], ensure_ascii=False),
                        output_meta=json.dumps(collect_output_meta(outputs), ensure_ascii=False),
                        output_summary='\n'.join(os.path.basename(o) for o in outputs),
                        input_samples=_sample_content(input_src),
                        output_samples=_sample_content(outputs),
                    )
            except Exception as e:
                logs.append(f"反思API调用失败（降级为仅结构化校验）: {str(e)}")
        ok = structural_ok and (reflection is None or reflection.get('satisfied', True))
        reason = '' if ok else (reflection.get('reason', '') if reflection else '结构化校验未通过')
        suggestions = (reflection or {}).get('suggestions', '')
        # 结构化校验失败时拼入沙箱诊断（帮助修复循环定位问题，如"无输出/残留文件"）
        exec_log = (state.get('execution_result') or {}).get('log') or ''
        if not ok and exec_log and '生成任何文件' in exec_log:
            suggestions = (suggestions + ' ' if suggestions else '') + f"沙箱诊断: {exec_log[:500]}"
        # 无输出/无建议时给默认建议，确保修复循环仍可触发
        if not ok and not suggestions:
            if not outputs:
                suggestions = ("代码未生成任何输出文件，请检查读取（os.environ['INPUT_DIR']）与写入"
                               "（os.environ['OUTPUT_DIR']）路径，确保代码执行到保存步骤")
            else:
                suggestions = "输出文件存在但未满足需求，请调整处理逻辑"
        logs.append(f"结果自检: {'通过' if ok else '未通过'} - {reason}")
        # 落盘 verify 记录（每轮修复各记一条）
        try:
            app._agent_write_log(state.get('task_id', ''), 'verify', {
                'round': state.get('repair_rounds') or 0,
                'ok': ok,
                'reason': reason,
                'checks': checks,
                'suggestions': suggestions,
            })
        except Exception:
            pass
        return {
            'verification': {'ok': ok, 'reason': reason, 'suggestions': suggestions, 'checks': checks},
            'repair_feedback': suggestions,
            'logs': logs,
        }

    def decide_verify(state):
        """Loop③条件路由：满足 → 总结；dynamic 未达上限且有建议 → 修复重入；否则 → verify_failed"""
        v = state.get('verification') or {}
        if v.get('ok'):
            return 'satisfied'
        route = state.get('route')
        rounds = state.get('repair_rounds') or 0
        max_r = state.get('max_repair_rounds') or 2
        if route == 'dynamic_code' and rounds < max_r and v.get('suggestions'):
            return 'repair'
        return 'verify_failed'

    def reflect_regenerate(state):
        """Loop③修复节点：携带反思意见 + 上一轮代码重新生成，repair_rounds 递增"""
        logs = list(state.get('logs') or [])
        api_url = app.agent_codegen_url_entry.get().strip()
        api_key = app.agent_codegen_key_entry.get().strip()
        if not api_url or not api_key:
            logs.append("代码生成 API 未配置，无法自动修复")
            return {'logs': logs, 'status': 'failed', 'error': '代码生成 API 未配置'}
        feedback = state.get('repair_feedback') or ''
        user_request = state['user_request']
        req = user_request + f"\n（上一轮结果未满足需求，改进意见：{feedback}）" if feedback else user_request
        meta_safe = []
        for m in state.get('file_metadata') or []:
            safe = {k: v for k, v in m.items() if k != 'path'}
            safe['path'] = '<input_dir>/' + os.path.basename(m.get('path') or '')
            meta_safe.append(safe)
        file_metadata = json.dumps(meta_safe, ensure_ascii=False)
        sample_rows = _read_sample_rows(state)
        try:
            result = app.call_code_generator_api(
                api_url, api_key, req, file_metadata, sample_rows, EXECUTION_CONSTRAINTS,
                previous_code=state.get('generated_code', ''))
            code = str(result.get('code', '') or '')
            if not code:
                logs.append("修复代码生成为空")
                return {'logs': logs, 'status': 'failed', 'error': '修复代码生成为空'}
            logs.append(f"第{(state.get('repair_rounds') or 0) + 1}轮修复代码生成成功")
            return {
                'generated_code': code,
                'code_description': str(result.get('description', '') or ''),
                'repair_rounds': (state.get('repair_rounds') or 0) + 1,
                'logs': logs,
            }
        except Exception as e:
            logs.append(f"修复代码生成失败: {str(e)}")
            return {'logs': logs, 'status': 'failed', 'error': str(e)}

    def verify_failed(state):
        logs = list(state.get('logs') or [])
        v = state.get('verification') or {}
        logs.append(f"结果未通过自检: {v.get('reason', '')}")
        return {'logs': logs, 'status': 'failed'}

    builder = StateGraph(FileAgentState)
    builder.add_node('inspect_files', inspect_files)
    builder.add_node('call_dify_intent', call_dify_intent)
    builder.add_node('clarify_interrupt', clarify_interrupt)
    builder.add_node('update_context', update_context)
    builder.add_node('show_understanding', show_understanding)
    builder.add_node('planner', planner)
    builder.add_node('route_task', route_task)
    builder.add_node('execute_builtin', execute_builtin)
    builder.add_node('validate_output', validate_output_node)
    builder.add_node('result_summary', result_summary)
    builder.add_node('generate_code', generate_code)
    builder.add_node('validate_code', validate_code)
    builder.add_node('sandbox_execute', sandbox_execute)
    builder.add_node('code_rejected', code_rejected)
    builder.add_node('verify_result', verify_result)
    builder.add_node('reflect_regenerate', reflect_regenerate)
    builder.add_node('verify_failed', verify_failed)
    builder.add_edge(START, 'inspect_files')
    builder.add_edge('inspect_files', 'call_dify_intent')
    # Loop①：澄清循环（decide_clarify 为条件路由函数，重入 call_dify_intent）
    builder.add_conditional_edges(
        'call_dify_intent',
        decide_clarify,
        {'clarify': 'clarify_interrupt', 'proceed': 'show_understanding'},
    )
    builder.add_edge('clarify_interrupt', 'update_context')
    builder.add_edge('update_context', 'call_dify_intent')
    # Loop②：理解确认循环（用户修正 → 重入 call_dify_intent）
    builder.add_conditional_edges(
        'show_understanding',
        decide_understanding,
        {'proceed': 'planner', 'revision': 'update_context'},
    )
    builder.add_edge('planner', 'route_task')
    builder.add_conditional_edges(
        'route_task',
        lambda s: s.get('route', 'builtin'),
        {'builtin': 'execute_builtin', 'dynamic_code': 'generate_code', 'clarify': 'clarify_interrupt'},
    )
    builder.add_edge('generate_code', 'validate_code')
    # 校验拒绝 → 自动重生成（repair_rounds 未达上限且有拒绝原因）；达上限/无原因 → code_rejected 安全终止
    builder.add_conditional_edges(
        'validate_code',
        lambda s: (
            'sandbox_execute' if (s.get('code_validation') or {}).get('valid')
            else ('repair' if (s.get('repair_rounds') or 0) < (s.get('max_repair_rounds') or 2)
                  and (s.get('code_validation') or {}).get('issues') else 'rejected')),
        {'sandbox_execute': 'sandbox_execute', 'repair': 'reflect_regenerate', 'rejected': 'code_rejected'},
    )
    builder.add_edge('code_rejected', END)
    builder.add_edge('execute_builtin', 'validate_output')
    builder.add_edge('sandbox_execute', 'validate_output')
    # Loop③：结果自检与修复循环（dynamic 未达上限 → 重新生成代码重入）
    builder.add_edge('validate_output', 'verify_result')
    builder.add_conditional_edges(
        'verify_result',
        decide_verify,
        {'satisfied': 'result_summary', 'repair': 'reflect_regenerate', 'verify_failed': 'verify_failed'},
    )
    builder.add_edge('reflect_regenerate', 'validate_code')
    builder.add_edge('verify_failed', END)
    builder.add_edge('result_summary', END)
    return builder.compile(checkpointer=InMemorySaver())


# ---------- 文件处理 Agent（Phase 5）动态代码分支 ----------

EXECUTION_CONSTRAINTS = (
    "输入文件目录在环境变量 INPUT_DIR，输出目录在环境变量 OUTPUT_DIR，临时目录在环境变量 TEMP_DIR。"
    "代码必须通过 os.environ 读取这些路径，例如："
    "input_dir = os.environ['INPUT_DIR']、output_dir = os.environ['OUTPUT_DIR']。"
    "绝对不要硬编码 /input、/output、/tmp 等路径，也不要读取 file_metadata 中的原始绝对路径。"
    "只能操作系统分配的工作目录；只能读取 INPUT_DIR 中的输入文件；只能写入 OUTPUT_DIR。"
    "必须至少生成 1 个输出文件到 OUTPUT_DIR；禁止只打印统计结果而不落盘。"
    "脚本结束前必须 print 生成的输出文件名或输出路径，便于日志排查。"
    "执行方式（二选一）：要么在顶层直接执行处理逻辑，要么把处理逻辑封装为入口函数并在文件末尾调用它；"
    "推荐入口函数名为 main 或 process_excel，沙箱会自动按 main → process_excel → process → run → execute → handle 的顺序调用第一个存在的入口函数。"
    "禁止既在顶层执行完整逻辑、又定义入口函数后再调用（会重复处理）；如果定义了入口函数，顶层只允许定义语句与 'if __name__ == '__main__': 入口名()'。"
    "禁止网络访问；禁止 subprocess/os.system/shell；禁止删除文件；禁止覆盖原文件；"
    "禁止读取环境变量中的密钥；禁止访问 HOME 目录；禁止 pip install；"
    "禁止 eval/exec/compile 及任何形式的动态执行（exec/eval/compile/__import__）；"
    "必须使用确定性的文件处理逻辑；优先使用 openpyxl/csv/json/pathlib/re/datetime，只有在明显更合适时才使用 pandas。"
)

# AST 校验黑名单（原则：禁危险库/操作，不禁合法文件处理所需的 os/shutil/pathlib 本身）
FORBIDDEN_IMPORTS = {
    'subprocess', 'socket', 'requests', 'urllib', 'http',
    'ctypes', 'sys', 'importlib', 'glob', 'pickle', 'marshal', 'shelve',
}
FORBIDDEN_ATTRS = {
    'system', 'popen', 'unlink', 'rmdir', 'chmod', 'chown', 'rename',
    'symlink', 'home', 'expanduser', 'remove', 'replace',
    'startfile', 'truncate', 'link', 'rmtree', 'move', 'getenv',
}
FORBIDDEN_NAMES = {'eval', 'exec', 'compile', '__import__', 'breakpoint'}
SUSPICIOUS_PATTERNS = (
    'pip install', 'curl ', 'wget ', 'rm -rf', 'chmod +x',
    '~/.ssh', '/etc/passwd', 'authorization',
)


def _call_name(node):
    """取调用表达式的函数名（Name/Attribute）"""
    if isinstance(node, ast.Name):
        return node.id
    if isinstance(node, ast.Attribute):
        return node.attr
    return None


def _has_top_level_execution(tree):
    """判断代码顶层是否有实际执行语句（顶层直接调用 或 if __name__ == '__main__' 块）"""
    for stmt in tree.body:
        if isinstance(stmt, ast.Expr) and isinstance(stmt.value, ast.Call):
            return True
        if isinstance(stmt, ast.If):
            test = stmt.test
            if (isinstance(test, ast.Compare) and len(test.comparators) == 1
                    and isinstance(test.comparators[0], ast.Constant)
                    and test.comparators[0].value == '__main__'):
                return True
    return False


def validate_generated_code(code):
    """AST 静态校验生成代码；返回 (valid, issues列表)。

    - valid: 是否存在硬性问题（导入黑名单/危险调用/危险属性/可疑字符串）
    - issues: 硬性问题 + 软提示（软提示带 [提示] 前缀，不导致拒绝，供修复反馈参考）
    """
    hard_issues = []
    hints = []
    if not code or not code.strip():
        return False, ['代码为空']
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return False, [f"语法错误: {e}"]
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for a in node.names:
                if a.name.split('.')[0] in FORBIDDEN_IMPORTS:
                    hard_issues.append(f"禁止导入: {a.name}")
        elif isinstance(node, ast.ImportFrom):
            if node.module and node.module.split('.')[0] in FORBIDDEN_IMPORTS:
                hard_issues.append(f"禁止导入: {node.module}")
        elif isinstance(node, ast.Call):
            fname = _call_name(node.func)
            if fname in FORBIDDEN_NAMES:
                hard_issues.append(f"禁止调用: {fname}")
            if fname and (fname.endswith('system') or fname.endswith('popen')):
                hard_issues.append(f"禁止调用: {fname}")
            # 软提示：save/to_excel/to_csv 使用字面量相对路径（不拒绝，供修复反馈）
            if fname in ('save', 'to_excel', 'to_csv'):
                target = node.args[0] if node.args else None
                if isinstance(target, ast.Constant) and isinstance(target.value, str):
                    p = target.value
                    if not os.path.isabs(p) and 'OUTPUT_DIR' not in p and 'output' not in p.lower():
                        hints.append(f"[提示] 输出路径疑似未使用OUTPUT_DIR（相对路径保存）: {p}")
        elif isinstance(node, ast.Attribute):
            if node.attr in FORBIDDEN_ATTRS:
                # 精确化：仅拦截 os./shutil./Path 等危险对象上的属性调用，
                # 避免误伤 str.replace / cell.value.replace / list.remove 等安全方法
                obj_src = ast.unparse(node.value) if hasattr(ast, 'unparse') else ''
                root = obj_src.split('.')[0]
                dangerous_obj = (root in ('os', 'shutil')
                                 or obj_src.startswith('Path(')
                                 or obj_src in ('Path', 'pathlib.Path'))
                if dangerous_obj or not obj_src:
                    hard_issues.append(f"禁止属性: {obj_src}.{node.attr}")
        elif isinstance(node, ast.Constant) and isinstance(node.value, str):
            low = node.value.lower()
            for pat in SUSPICIOUS_PATTERNS:
                if pat in low:
                    hard_issues.append(f"可疑字符串: {node.value[:50]}")
                    break
        elif isinstance(node, ast.Try):
            # 空 except 块（仅 pass / 空）→ 吞异常，禁止（异常必须暴露以便诊断）
            for handler in node.handlers:
                body = handler.body or []
                if not body or all(isinstance(st, ast.Pass) for st in body):
                    hard_issues.append("禁止空except吞异常（except 块必须 print 异常或 raise，异常必须暴露）")
    # 软提示：顶层只有函数定义、无实际调用 → 处理逻辑可能不执行
    has_func = any(isinstance(s, (ast.FunctionDef, ast.AsyncFunctionDef)) for s in tree.body)
    if has_func and not _has_top_level_execution(tree):
        hints.append("[提示] 代码仅在顶层定义了函数但未调用，处理逻辑可能不会执行；"
                     "请在末尾添加入口函数调用（如 process_excel()）")
    return len(hard_issues) == 0, hard_issues + hints


def _sandbox_import(name, globals=None, locals=None, fromlist=(), level=0):
    """限制动态代码可导入的模块范围。"""
    root = name.split('.')[0]
    if root in FORBIDDEN_IMPORTS:
        raise ImportError(f"禁止导入: {name}")
    allowed_roots = {
        'os', 'json', 'csv', 're', 'math', 'statistics',
        'pathlib', 'datetime', 'pandas', 'openpyxl',
        'shutil', 'time',
    }
    if root not in allowed_roots:
        raise ImportError(f"未允许的导入: {name}")
    return __import__(name, globals, locals, fromlist, level)


def _install_timeout(timeout):
    """Unix 主线程上安装 SIGALRM 超时；非主线程（signal.signal 受限）或 Windows 无信号支持则跳过。

    返回是否已安装。GUI 后台线程场景返回 False（超时不生效，与旧版行为一致）。
    """
    import signal
    import threading
    if (timeout and timeout > 0 and hasattr(signal, 'SIGALRM')
            and threading.current_thread() is threading.main_thread()):
        def _handler(signum, frame):
            raise TimeoutError(f"沙箱执行超过 {timeout} 秒，已终止")
        signal.signal(signal.SIGALRM, _handler)
        signal.setitimer(signal.ITIMER_REAL, timeout)
        return True
    return False


def _clear_timeout(installed):
    """清除 SIGALRM 定时器并恢复默认处理。"""
    import signal
    if installed:
        try:
            signal.setitimer(signal.ITIMER_REAL, 0)
            signal.signal(signal.SIGALRM, signal.SIG_DFL)
        except Exception:
            pass


def _sandbox_builtins():
    """动态代码可用的最小内建函数集合。"""
    builtin_obj = __builtins__ if isinstance(__builtins__, dict) else __builtins__.__dict__
    allowed = {
        'abs', 'all', 'any', 'bool', 'dict', 'enumerate', 'filter', 'float',
        'int', 'isinstance', 'len', 'list', 'max', 'min', 'print', 'range',
        'round', 'set', 'sorted', 'str', 'sum', 'tuple', 'zip',
        'BaseException', 'Exception', 'ValueError', 'TypeError', 'RuntimeError',
        'KeyError', 'IndexError', 'NameError', 'FileNotFoundError', 'OSError',
    }
    result = {name: builtin_obj[name] for name in allowed if name in builtin_obj}
    result['__import__'] = _sandbox_import
    return result


def _extract_json_string_field(text, field):
    """从 JSON/类 JSON 文本中提取字符串字段。"""
    if not text:
        return ''
    patterns = [
        rf'"{re.escape(field)}"\s*:\s*"((?:\\.|[^"\\])*)"',
        rf"'{re.escape(field)}'\s*:\s*'((?:\\.|[^'\\])*)'",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.DOTALL)
        if not m:
            continue
        raw = m.group(1)
        try:
            if pat.startswith('"'):
                return json.loads(f'"{raw}"')
            return ast.literal_eval(f"'{raw}'")
        except (json.JSONDecodeError, SyntaxError, ValueError):
            return raw.replace('\\n', '\n').replace('\\"', '"').replace("\\'", "'")
    return ''


def _normalize_codegen_payload(payload):
    """将代码生成结果统一整理为 {code, description, ...} 结构。"""
    if isinstance(payload, dict):
        code = str(payload.get('code', '') or '')
        desc = str(payload.get('description', '') or '')
        if code:
            normalized = dict(payload)
            normalized['code'] = code
            normalized['description'] = desc
            return normalized
    return None


# 沙箱工作区根目录：源码运行取项目目录；PyInstaller 打包后 __file__ 指向临时解压目录或
# 包内资源目录（退出即删/只读），故 frozen 模式改用可执行文件所在目录，保证沙箱文件持久保留
if getattr(sys, 'frozen', False):
    exe_dir = os.path.dirname(os.path.abspath(sys.executable))
    # macOS .app 包内可执行文件路径为 dist/App.app/Contents/MacOS/App，
    # 需提升到 dist/ 目录存放沙箱工作区，避免写入 .app 包内部
    if os.path.basename(exe_dir) == 'MacOS' and os.path.basename(os.path.dirname(exe_dir)) == 'Contents':
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(exe_dir)))
    else:
        BASE_DIR = exe_dir
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

APP_NAME = "Excel处理器-Dify答案匹配工具"


def _is_app_translocated(path):
    """macOS 通过 App Translocation 运行时，应用会落到只读的临时挂载目录。"""
    norm = os.path.abspath(path)
    return sys.platform == 'darwin' and '/AppTranslocation/' in norm


def _is_writable_dir(path):
    """检测目录是否可写；不存在时尝试创建并写一个临时探针文件。"""
    try:
        os.makedirs(path, exist_ok=True)
        probe = os.path.join(path, '.write_probe')
        with open(probe, 'a', encoding='utf-8'):
            pass
        os.remove(probe)
        return True
    except Exception:
        return False


def _default_user_data_dir():
    """跨平台用户数据目录：当产物同级目录不可写时，回退到这里。"""
    home = os.path.expanduser('~')
    if sys.platform == 'darwin':
        return os.path.join(home, 'Library', 'Application Support', APP_NAME)
    if os.name == 'nt':
        return os.path.join(os.environ.get('APPDATA', home), APP_NAME)
    return os.path.join(home, f'.{APP_NAME}')


_STORAGE_DIR_CACHE = [None]


def _resolve_storage_dir():
    """解析可写的应用存储目录；优先产物同级，不可写或被 App Translocation 时回退。"""
    if _STORAGE_DIR_CACHE[0] is None:
        candidates = []
        if not _is_app_translocated(BASE_DIR):
            candidates.append(BASE_DIR)
        candidates.append(_default_user_data_dir())
        candidates.append(os.path.expanduser('~'))
        for d in candidates:
            if _is_writable_dir(d):
                _STORAGE_DIR_CACHE[0] = d
                break
        if _STORAGE_DIR_CACHE[0] is None:
            _STORAGE_DIR_CACHE[0] = os.path.expanduser('~')
    return _STORAGE_DIR_CACHE[0]


def _resolve_agent_workspace_root():
    """Agent 沙箱工作区根目录。"""
    return os.path.join(_resolve_storage_dir(), 'agent_workspace')

# 运行日志落盘（exe 同目录；源码运行时在项目目录）。任何异常静默忽略，不影响主流程。
_LOG_LOCK = threading.Lock()


def _resolve_log_dir():
    """确定可用的日志目录：优先产物同级，不可写或被转移运行时回退到用户数据目录。"""
    return _resolve_storage_dir()


def _runtime_log(msg):
    """追加一行到 运行日志.log，格式与界面日志一致 [HH:MM:SS]"""
    try:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n"
        with _LOG_LOCK:
            with open(os.path.join(_resolve_log_dir(), '运行日志.log'), 'a', encoding='utf-8') as fh:
                fh.write(line)
    except Exception:
        pass


# 沙箱工作区内系统文件（残留检测时排除，避免误报为动态代码写入）
_SANDBOX_SYSTEM_FILES = {'script.py', 'execution.log', 'metadata.json'}

# 动态代码常见入口函数名（沙箱按优先级自动调用第一个可调用者；main 优先）
_ENTRY_FUNCTION_NAMES = ('main', 'process_excel', 'process', 'run', 'execute', 'handle')


def run_code_sandbox(code, input_files, out_dir, task_id, timeout=60):
    """受限执行生成代码：隔离工作区 + 环境白名单；返回 (ok, log, output_files)"""
    ws_dir = os.path.join(_resolve_agent_workspace_root(), task_id)
    in_dir = os.path.join(ws_dir, 'input')
    out_ws = os.path.join(ws_dir, 'output')
    tmp_dir = os.path.join(ws_dir, 'temp')
    for d in (in_dir, out_ws, tmp_dir):
        os.makedirs(d, exist_ok=True)

    copied = []
    for f in input_files:
        dest = os.path.join(in_dir, os.path.basename(f))
        shutil.copy2(f, dest)
        copied.append(dest)

    script_path = os.path.join(ws_dir, 'script.py')
    with open(script_path, 'w', encoding='utf-8') as fh:
        fh.write(code)

    env = {
        'PATH': '/usr/bin:/bin:/usr/local/bin',
        'TASK_ID': task_id,
        'INPUT_DIR': in_dir,
        'OUTPUT_DIR': out_ws,
        'TEMP_DIR': tmp_dir,
        'INPUT_FILES': json.dumps(copied),
        'PYTHONPATH': '',
        'LANG': 'en_US.UTF-8',
    }
    timeout_active = False
    try:
        old_cwd = os.getcwd()
        old_env = {k: os.environ.get(k) for k in env}
        stdout_buffer = io.StringIO()
        sandbox_globals = {
            '__name__': '__sandbox__',
            '__builtins__': _sandbox_builtins(),
        }
        compiled = compile(code, script_path, 'exec')
        try:
            os.chdir(ws_dir)
            os.environ.update(env)
            timeout_active = _install_timeout(timeout)
            with contextlib.redirect_stdout(stdout_buffer), contextlib.redirect_stderr(stdout_buffer):
                exec(compiled, sandbox_globals, sandbox_globals)
                # 自动调用入口函数（main 优先，其次常见入口名；LLM 常把逻辑包在函数里只定义不调用）
                entry_name = None
                for name in _ENTRY_FUNCTION_NAMES:
                    candidate = sandbox_globals.get(name)
                    if callable(candidate):
                        entry_name = name
                        break
                if entry_name:
                    sandbox_globals[entry_name]()
        finally:
            _clear_timeout(timeout_active)
            os.chdir(old_cwd)
            for key, value in old_env.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
        log = stdout_buffer.getvalue()[-4000:]
        outputs = [os.path.join(out_ws, n) for n in os.listdir(out_ws)
                   if os.path.isfile(os.path.join(out_ws, n))]
        if not outputs:
            detail = log or "动态代码执行完成，但没有在 OUTPUT_DIR 中生成任何文件"
            # 诊断：检查沙箱工作目录（chdir 目标）是否有代码写入的残留文件（相对路径保存的证据），排除系统文件
            leftovers = [n for n in os.listdir(ws_dir)
                         if os.path.isfile(os.path.join(ws_dir, n)) and n not in _SANDBOX_SYSTEM_FILES]
            if leftovers:
                detail += (f"；检测到代码可能写入了工作目录而非 OUTPUT_DIR（残留文件: {', '.join(leftovers[:5])}）。"
                           "请通过 os.environ['OUTPUT_DIR'] 拼接输出路径，例如 "
                           "output_path = os.path.join(os.environ['OUTPUT_DIR'], '结果.xlsx')")
            elif not log:
                # 无残留且无 stdout → 代码未执行到写入步骤（如读取输入失败被 try/except 吞掉）
                detail += ("；脚本未执行到写入步骤（无输出文件、无残留文件、无 stdout）。"
                           "常见原因：读取输入文件失败被 try/except 吞掉。"
                           "请用 os.path.join(os.environ['INPUT_DIR'], '输入文件名') 读取输入，"
                           "用 os.path.join(os.environ['OUTPUT_DIR'], '结果.xlsx') 保存输出，"
                           "且代码末尾必须 print 输出文件名、不要用 try/except 吞掉异常")
            elif not log:
                detail += "；脚本未输出任何 stdout，请确保代码末尾 print 输出文件名"
            return False, detail, []
        final_outputs = []
        for o in outputs:
            dest = os.path.join(out_dir, os.path.basename(o))
            shutil.copy2(o, dest)
            final_outputs.append(dest)
        return True, log, final_outputs
    except TimeoutError as e:
        _clear_timeout(timeout_active)
        return False, f"执行超时: {str(e)}", []
    except Exception as e:
        _clear_timeout(timeout_active)
        return False, f"执行异常: {str(e)}", []


def open_directory_in_explorer(path):
    """跨平台打开文件所在目录；path 可为文件或目录。"""
    target = os.path.abspath(path)
    directory = target if os.path.isdir(target) else os.path.dirname(target)
    if not os.path.isdir(directory):
        raise FileNotFoundError(f"目录不存在: {directory}")
    if os.name == 'nt':
        os.startfile(directory)
        return
    if sys.platform == 'darwin':
        subprocess.run(['open', directory], check=True)
        return
    subprocess.run(['xdg-open', directory], check=True)


class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel表格处理工具 - Dify答案匹配")
        self.root.geometry("900x700")

        # 设置窗口图标（如果存在）
        try:
            if os.path.exists('icon.ico'):
                self.root.iconbitmap('icon.ico')
        except:
            pass

        # 必须在隐藏主窗口之前完成居中：withdraw 后 winfo_width/height 在 Windows 上返回 1，
        # 若此时再读尺寸设置 geometry，主窗口会变成 1x1 像素导致界面不可见
        self._center_window()

        self._splash = None
        self._show_splash()  # 先显示启动画面，主窗口暂隐藏

        self.selected_files = []
        self.processing = False
        self.forward_selected_files = []
        self.forward_processing = False

        # 保存调试数据的变量
        self.current_question = ""
        self.current_api_response = ""
        self.latest_output_path = None
        self.log_file_path = os.path.join(_resolve_log_dir(), '运行日志.log')

        # 各 API Key 输入框的查看保护状态：widget -> {"clicks":int, "last":float, "unlocked":bool, "after_id":str|None}
        self._apikey_state = {}

        self.create_widgets()
        self._log_window_state("界面组件创建完成")
        self.root.after(SPLASH_MS, self._close_splash)

    def _log_window_state(self, tag, widget=None):
        """记录窗口关键状态（state/geometry/尺寸/位置/映射/可见性），用于排查界面不显示问题"""
        try:
            w = widget if widget is not None else self.root
            try:
                st = w.state()
            except Exception:
                st = 'n/a'
            try:
                geo = w.geometry()
            except Exception:
                geo = 'n/a'
            _runtime_log(f"[窗口] {tag}: state={st} geometry={geo} "
                         f"size={w.winfo_width()}x{w.winfo_height()} "
                         f"pos=({w.winfo_x()},{w.winfo_y()}) "
                         f"mapped={w.winfo_ismapped()} viewable={w.winfo_viewable()}")
        except Exception:
            pass

    def _center_window(self):
        """将主窗口居中（须在 withdraw 显示启动画面之前调用）"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        # 兜底：未取到真实尺寸时使用默认 900x700
        if width < 100 or height < 100:
            width, height = 900, 700
        x = (self.root.winfo_screenwidth() - width) // 2
        y = (self.root.winfo_screenheight() - height) // 2
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        self._log_window_state("窗口居中完成")

    def _show_splash(self):
        """显示无边框启动画面，主窗口暂隐藏"""
        self.root.withdraw()
        self._log_window_state("主窗口已隐藏(启动画面阶段)")
        splash = tk.Toplevel(self.root)
        splash.overrideredirect(True)
        splash.configure(bg="#2b579a")
        w, h = 420, 220
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        splash.geometry(f"{w}x{h}+{x}+{y}")
        tk.Label(splash, text="Excel表格处理工具", font=("PingFang SC", 18, "bold"),
                 bg="#2b579a", fg="white").pack(pady=(45, 5))
        tk.Label(splash, text="Dify答案匹配 - 正在启动...", font=("PingFang SC", 11),
                 bg="#2b579a", fg="#d5e3ff").pack()
        tk.Label(splash, text="v1.0", font=("PingFang SC", 9),
                 bg="#2b579a", fg="#a8c4f0").pack(side=tk.BOTTOM, pady=10)
        self._splash = splash
        self.root.update()  # 立即绘制启动画面
        self._log_window_state("启动画面已显示", splash)

    def _close_splash(self):
        """关闭启动画面，显示主窗口"""
        if self._splash is not None:
            try:
                self._splash.destroy()
            except tk.TclError:
                pass
            self._splash = None
        try:
            self.root.deiconify()
            self.root.update_idletasks()
        except tk.TclError:
            pass
        self._log_window_state("主窗口已显示(启动画面关闭)")
        # 延迟复核：确认主窗口显示后仍保持可见（排查“闪一下又消失”的情况）
        self.root.after(1500, lambda: self._log_window_state("主窗口显示1.5秒后"))
        
    def create_widgets(self):
        """创建界面组件"""
        # 状态栏（先创建，固定在底部）
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_frame = ttk.Frame(self.root, relief=tk.SUNKEN)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        status_bar = ttk.Label(status_frame, textvariable=self.status_var, anchor=tk.W)
        status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
        ttk.Button(status_frame, text="打开最近输出目录", command=self.open_latest_output_dir).pack(side=tk.RIGHT, padx=4, pady=2)
        ttk.Button(status_frame, text="打开日志目录", command=self.open_log_dir).pack(side=tk.RIGHT, padx=4, pady=2)

        # 主容器：Notebook 双TAB（两种模式）
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # TAB1【反向匹配（知识库溯源）】：现有功能，放最前面
        self.reverse_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.reverse_tab, text="反向匹配（知识库溯源）")

        # TAB2【正向问答（工作流溯源）】：新增框架，后续逐步增加工作流调用接口
        self.forward_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.forward_tab, text="正向问答（工作流溯源）")

        # TAB3【文件处理 Agent】：自然语言驱动文件处理
        self.agent_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.agent_tab, text="文件处理 Agent")

        self.create_reverse_match_tab(self.reverse_tab)
        self.create_forward_qa_tab(self.forward_tab)
        self.create_agent_tab(self.agent_tab)

        # TAB切换时强制刷新内容，修复Notebook内嵌Canvas不渲染（空白）问题
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # 统一绑定鼠标滚轮（指向当前激活TAB的Canvas，避免bind_all互相覆盖）
        self._active_canvas = self.reverse_canvas
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)

    def _remember_output_file(self, path):
        """记录当前会话最近一次修改或生成的输出文件。"""
        if path:
            self.latest_output_path = os.path.abspath(path)

    def _show_open_dir_error(self, title, err):
        messagebox.showerror(title, f"打开目录失败：{str(err)}")

    def open_latest_output_dir(self):
        """打开最近一次输出文件所在目录；若当前尚无输出，则退回日志目录。"""
        target = self.latest_output_path or self.log_file_path
        try:
            open_directory_in_explorer(target)
        except Exception as e:
            self._show_open_dir_error("打开最近输出目录失败", e)

    def open_log_dir(self):
        """打开运行日志所在目录。"""
        try:
            self.log_file_path = os.path.join(_resolve_log_dir(), '运行日志.log')
            open_directory_in_explorer(self.log_file_path)
        except Exception as e:
            self._show_open_dir_error("打开日志目录失败", e)

    def _setup_apikey_entry(self, entry):
        """为 API Key 输入框统一配置：默认值掩码保护、用户自填值明文；连点5次+密码查看默认值"""
        entry.configure(show="*")
        # default 记录硬编码默认值（调用方在 insert 后调用本方法），值==默认值时才受保护
        self._apikey_state[entry] = {"clicks": 0, "last": 0.0, "unlocked": False,
                                     "after_id": None, "default": entry.get()}

        # 默认值锁定状态下禁止复制/剪切快捷键（Windows/Linux: Ctrl，macOS: Command）
        for seq in ("<Control-c>", "<Control-x>", "<Command-c>", "<Command-x>"):
            entry.bind(seq, lambda e, w=entry: self._on_apikey_copy_attempt(w))

        # 默认值锁定状态下禁止右键/控制键点击弹出系统上下文菜单（含复制项）
        for seq in ("<Button-2>", "<Button-3>", "<Control-Button-1>"):
            entry.bind(seq, lambda e, w=entry: self._on_apikey_menu_attempt(w))

        # 连续点击计数（add="+" 保留默认聚焦/选中行为）
        entry.bind("<Button-1>", lambda e, w=entry: self._on_apikey_click(w), add="+")

        # 值变化（输入/粘贴/删除）时按值刷新掩码保护：用户自填值明文显示
        entry.bind("<KeyRelease>", lambda e, w=entry: self._refresh_apikey_protection(w), add="+")
        entry.bind("<<Paste>>", lambda e, w=entry: self._refresh_apikey_protection(w), add="+")

    def _refresh_apikey_protection(self, entry):
        """按当前值刷新保护状态：值为默认值时掩码保护；用户自填值明文放行"""
        st = self._apikey_state[entry]
        if _apikey_should_protect(entry.get(), st["default"], st["unlocked"]):
            entry.configure(show="*")  # 值为默认值且未解锁：掩码保护
        elif entry.get() == st["default"]:
            return  # 解锁查看期间，值未变：保持明文，等待自动重锁
        else:
            # 用户自己填入的值：明文显示、可复制；取消未执行的自动重锁
            if st["after_id"]:
                try:
                    self.root.after_cancel(st["after_id"])
                except tk.TclError:
                    pass
                st["after_id"] = None
            st["unlocked"] = False
            entry.configure(show="")

    def _on_apikey_click(self, entry):
        """连续点击计数：默认值在2秒窗口内累计5次触发密码查看弹窗"""
        st = self._apikey_state[entry]
        if st["unlocked"] or entry.get() != st["default"]:
            return  # 已解锁或用户自填值，无需查看保护
        if _apikey_click_tick(st, time.monotonic()):
            self.status_var.set("正在验证查看密码…")
            # 延迟到本次点击事件（press/release）处理完毕后再打开模态对话框，
            # 避免在 Button-1 回调内直接嵌套事件循环导致 macOS 下弹窗按钮无响应
            self.root.after_idle(lambda: self._prompt_apikey_unlock(entry))
        else:
            remaining = APIKEY_CLICK_REQUIRED - st["clicks"]
            self.status_var.set(f"连续点击 API Key 输入框 {remaining} 次可查看明文")

    def _on_apikey_copy_attempt(self, entry):
        """默认值锁定状态禁止复制/剪切；用户自填值或已解锁放行"""
        st = self._apikey_state[entry]
        if not _apikey_should_protect(entry.get(), st["default"], st["unlocked"]):
            return  # 放行（返回 None 让默认行为生效）
        return "break"

    def _on_apikey_menu_attempt(self, entry):
        """默认值锁定状态禁止系统上下文菜单；用户自填值或已解锁放行"""
        st = self._apikey_state[entry]
        if not _apikey_should_protect(entry.get(), st["default"], st["unlocked"]):
            return  # 放行
        return "break"

    def _prompt_apikey_unlock(self, entry):
        """弹出密码输入对话框，校验通过后解锁显示明文并允许复制"""
        if self._apikey_state[entry]["unlocked"]:
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("查看 API Key")
        dialog.resizable(False, False)
        dialog.transient(self.root)

        # 基于主窗口居中显示
        dialog.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - dialog.winfo_reqwidth()) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - dialog.winfo_reqheight()) // 3
        dialog.geometry(f"+{x}+{y}")

        ttk.Label(dialog, text="请输入查看密码：").pack(padx=24, pady=(18, 6))
        pwd_var = tk.StringVar()
        pwd_entry = ttk.Entry(dialog, textvariable=pwd_var, show="*", width=26)
        pwd_entry.pack(padx=24, pady=6)

        result = {"ok": False}

        def on_ok():
            if pwd_var.get() == APIKEY_VIEW_PASSWORD:
                result["ok"] = True
                dialog.destroy()
            else:
                messagebox.showerror("密码错误", "密码不正确，无法查看 API Key。", parent=dialog)
                pwd_var.set("")
                pwd_entry.focus_set()

        def on_cancel():
            dialog.destroy()

        pwd_entry.bind("<Return>", lambda e: on_ok())
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=(6, 18))
        ttk.Button(btn_frame, text="确定", command=on_ok, width=8).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frame, text="取消", command=on_cancel, width=8).pack(side=tk.LEFT, padx=6)

        # 标准模态模式（与 tkinter.simpledialog.Dialog 一致）：窗口映射可见后再设置
        # grab 与焦点，避免 macOS 下 grab_set 作用于未映射窗口导致按钮事件被吞掉
        dialog.wait_visibility()
        dialog.grab_set()
        pwd_entry.focus_set()

        dialog.wait_window()

        if result["ok"]:
            self._unlock_apikey_entry(entry)

    def _unlock_apikey_entry(self, entry):
        """解锁：显示明文 API Key 并允许复制，定时自动重新掩码"""
        st = self._apikey_state[entry]
        st["unlocked"] = True
        entry.configure(show="")  # 显示明文
        self.status_var.set(f"API Key 已解锁显示，{APIKEY_UNLOCK_SECONDS} 秒后自动隐藏")
        # 先取消可能存在的旧定时器，再安排自动重锁
        if st["after_id"]:
            try:
                self.root.after_cancel(st["after_id"])
            except tk.TclError:
                pass
        st["after_id"] = self.root.after(APIKEY_UNLOCK_SECONDS * 1000, lambda: self._lock_apikey_entry(entry))

    def _lock_apikey_entry(self, entry):
        """自动重锁：掩码判定委托给 _refresh_apikey_protection（用户已自填值则不再掩码）"""
        st = self._apikey_state[entry]
        if not st["unlocked"]:
            return
        try:
            if not entry.winfo_exists():
                return
        except tk.TclError:
            return
        st["unlocked"] = False
        st["after_id"] = None
        self._refresh_apikey_protection(entry)  # 值==默认值才掩码，用户自填值保持明文

    def create_reverse_match_tab(self, parent):
        """创建反向匹配（知识库溯源）TAB页 —— 现有功能：从答案反向在知识库中匹配查找系统来源"""
        # 创建主滚动区域
        main_canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        # 当scrollable_frame内容变化时，更新滚动区域
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        # 创建窗口并设置宽度自适应
        canvas_window = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        self.reverse_canvas = main_canvas
        self.reverse_canvas_window = canvas_window

        # 当canvas大小变化时，调整scrollable_frame的宽度
        def configure_scroll_region(event):
            # 设置scrollable_frame的宽度等于canvas的宽度
            main_canvas.itemconfig(canvas_window, width=event.width)

        main_canvas.bind("<Configure>", configure_scroll_region)
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # 将主滚动区域添加到窗口（在状态栏之后）
        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 顶部配置区域 - 主API
        config_frame = ttk.LabelFrame(scrollable_frame, text="Dify API配置（知识库检索）", padding=10)
        config_frame.pack(fill=tk.X, padx=10, pady=5)

        # 配置grid列权重，让第二列能够自适应宽度
        config_frame.columnconfigure(1, weight=1)

        # API地址
        ttk.Label(config_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.api_url_entry = ttk.Entry(config_frame, width=50)
        self.api_url_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        self.api_url_entry.insert(0, "http://10.133.175.249/v1")

        # API Key
        ttk.Label(config_frame, text="API Key:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.api_key_entry = ttk.Entry(config_frame, width=50)
        self.api_key_entry.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.api_key_entry.insert(0, "app-FF8kXf1p3R0FVYD0mzw5KOO5")
        self._setup_apikey_entry(self.api_key_entry)

        # Code参数（必填）
        ttk.Label(config_frame, text="Code (必填):").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.code_entry = ttk.Entry(config_frame, width=50)
        self.code_entry.grid(row=2, column=1, padx=5, pady=2, sticky="ew")
        self.code_entry.insert(0, "CR039")

        # Channel参数（可选）
        ttk.Label(config_frame, text="Channel (可选):").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.channel_entry = ttk.Entry(config_frame, width=50)
        self.channel_entry.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

        # 第二个API配置区域 - 相似度判断
        similarity_frame = ttk.LabelFrame(scrollable_frame, text="Dify API配置（相似度判断）", padding=10)
        similarity_frame.pack(fill=tk.X, padx=10, pady=5)

        # 配置grid列权重
        similarity_frame.columnconfigure(1, weight=1)

        # 相似度API地址
        ttk.Label(similarity_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.similarity_api_url_entry = ttk.Entry(similarity_frame, width=50)
        self.similarity_api_url_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        self.similarity_api_url_entry.insert(0, "http://10.0.173.34/v1")

        # 相似度API Key
        ttk.Label(similarity_frame, text="API Key:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.similarity_api_key_entry = ttk.Entry(similarity_frame, width=50)
        self.similarity_api_key_entry.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.similarity_api_key_entry.insert(0, "app-GrUikNKKtyRIEmQXn4F7QRID")
        self._setup_apikey_entry(self.similarity_api_key_entry)

        # 调试模式开关
        self.debug_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(similarity_frame, text="调试模式（显示详细API调用信息）", variable=self.debug_var).grid(row=2, column=0, columnspan=2, pady=2, sticky=tk.W)

        # 测试按钮
        test_btn_frame = ttk.Frame(similarity_frame)
        test_btn_frame.grid(row=3, column=0, columnspan=2, pady=5)
        self.test_btn = ttk.Button(test_btn_frame, text="测试主API连接", command=self.test_api_connection)
        self.test_btn.pack(side=tk.LEFT, padx=5)
        self.test_similarity_btn = ttk.Button(test_btn_frame, text="测试相似度API", command=self.test_similarity_api_connection)
        self.test_similarity_btn.pack(side=tk.LEFT, padx=5)

        # 文件选择区域
        file_frame = ttk.LabelFrame(scrollable_frame, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        self.select_btn = ttk.Button(file_frame, text="选择Excel文件（可多选）", command=self.select_files)
        self.select_btn.pack(side=tk.LEFT, padx=5)

        self.clear_btn = ttk.Button(file_frame, text="清空文件列表", command=self.clear_files)
        self.clear_btn.pack(side=tk.LEFT, padx=5)

        # 文件列表
        self.file_listbox = tk.Listbox(file_frame, height=5, selectmode=tk.EXTENDED)
        self.file_listbox.pack(fill=tk.X, pady=5)

        # 处理按钮区域
        process_frame = ttk.Frame(scrollable_frame, padding=5)
        process_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # 并发数选择
        ttk.Label(process_frame, text="并发数:").pack(side=tk.LEFT, padx=5)
        self.concurrent_var = tk.StringVar(value="2")
        concurrent_options = ["1", "2", "3", "4", "5", "6", "8", "10"]
        self.concurrent_combo = ttk.Combobox(process_frame, textvariable=self.concurrent_var, 
                                              values=concurrent_options, width=5, state="readonly")
        self.concurrent_combo.pack(side=tk.LEFT, padx=5)
        
        # 请求间隔选择
        ttk.Label(process_frame, text="请求间隔:").pack(side=tk.LEFT, padx=5)
        self.interval_var = tk.StringVar(value="0.5秒")
        interval_options = ["0秒", "0.5秒", "1秒", "2秒", "3秒"]
        self.interval_combo = ttk.Combobox(process_frame, textvariable=self.interval_var,
                                            values=interval_options, width=6, state="readonly")
        self.interval_combo.pack(side=tk.LEFT, padx=5)
        
        self.process_btn = ttk.Button(process_frame, text="开始处理", command=self.start_processing)
        self.process_btn.pack(side=tk.LEFT, padx=10)
        
        self.stop_btn = ttk.Button(process_frame, text="停止处理", command=self.stop_processing, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(process_frame, variable=self.progress_var, maximum=100, length=400)
        self.progress_bar.pack(side=tk.LEFT, padx=10)
        
        self.progress_label = ttk.Label(process_frame, text="进度: 0%")
        self.progress_label.pack(side=tk.LEFT, padx=5)

        # 日志与调试区域 - 使用PanedWindow支持拖拽调整大小
        log_debug_pane = ttk.PanedWindow(scrollable_frame, orient=tk.VERTICAL)
        log_debug_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 处理日志区域
        log_frame = ttk.LabelFrame(log_debug_pane, text="处理日志", padding=10)
        log_debug_pane.add(log_frame, weight=1)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 调试信息区域
        debug_frame = ttk.LabelFrame(log_debug_pane, text="调试信息（可复制）", padding=10)
        log_debug_pane.add(debug_frame, weight=1)

        # 添加按钮区域
        debug_btn_frame = ttk.Frame(debug_frame)
        debug_btn_frame.pack(fill=tk.X)

        ttk.Button(debug_btn_frame, text="清空调试信息", command=self.clear_debug_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(debug_btn_frame, text="复制Question参数", command=self.copy_question).pack(side=tk.LEFT, padx=5)
        ttk.Button(debug_btn_frame, text="复制API返回数据", command=self.copy_api_response).pack(side=tk.LEFT, padx=5)

        self.debug_text = scrolledtext.ScrolledText(debug_frame, height=8, wrap=tk.WORD)
        self.debug_text.pack(fill=tk.BOTH, expand=True)

    def create_forward_qa_tab(self, parent):
        """创建正向问答（工作流溯源）TAB页 —— 正向调用知识问答工作流，从工作流中获取知识系统来源；
        后续将在此页逐步增加调用知识问答工作流的接口"""
        # 创建主滚动区域（与反向匹配TAB结构一致）
        main_canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        # 当scrollable_frame内容变化时，更新滚动区域
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        # 创建窗口并设置宽度自适应
        canvas_window = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        self.forward_canvas = main_canvas
        self.forward_canvas_window = canvas_window

        # 当canvas大小变化时，调整scrollable_frame的宽度
        def configure_scroll_region(event):
            # 设置scrollable_frame的宽度等于canvas的宽度
            main_canvas.itemconfig(canvas_window, width=event.width)

        main_canvas.bind("<Configure>", configure_scroll_region)
        main_canvas.configure(yscrollcommand=scrollbar.set)
        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 1. 模式说明区
        desc_frame = ttk.LabelFrame(scrollable_frame, text="模式说明", padding=10)
        desc_frame.pack(fill=tk.X, padx=10, pady=5)
        desc_text = ("本页为【正向问答】模式：调用知识问答完整问答流(/chat-messages)，\n"
                     "将H列问题传入，从返回结果中解析系统来源(systemfrom)并写入O列（带(正向)标识）。")
        ttk.Label(desc_frame, text=desc_text, wraplength=760, justify=tk.LEFT).pack(anchor=tk.W)

        # 2. 工作流 API 配置区
        forward_config_frame = ttk.LabelFrame(scrollable_frame, text="知识问答工作流 API配置", padding=10)
        forward_config_frame.pack(fill=tk.X, padx=10, pady=5)

        # 配置grid列权重，让第二列能够自适应宽度
        forward_config_frame.columnconfigure(1, weight=1)

        ttk.Label(forward_config_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.forward_api_url_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_api_url_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        self.forward_api_url_entry.insert(0, "http://10.133.175.249/v1")

        ttk.Label(forward_config_frame, text="API Key:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.forward_api_key_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_api_key_entry.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.forward_api_key_entry.insert(0, "app-lm80B7N9ZAAq22Wd70c4lZF7")
        self._setup_apikey_entry(self.forward_api_key_entry)

        ttk.Label(forward_config_frame, text="Code:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.forward_code_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_code_entry.grid(row=2, column=1, padx=5, pady=2, sticky="ew")
        self.forward_code_entry.insert(0, "CR039")

        ttk.Label(forward_config_frame, text="Channel (可选):").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.forward_channel_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_channel_entry.grid(row=3, column=1, padx=5, pady=2, sticky="ew")

        ttk.Label(forward_config_frame, text="InternetFlag (必填):").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.forward_internet_flag_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_internet_flag_entry.grid(row=4, column=1, padx=5, pady=2, sticky="ew")
        self.forward_internet_flag_entry.insert(0, "1")

        ttk.Label(forward_config_frame, text="SystemSource:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.forward_system_source_entry = ttk.Entry(forward_config_frame, width=50)
        self.forward_system_source_entry.grid(row=5, column=1, padx=5, pady=2, sticky="ew")

        # 2.5 文件选择区（正向批量处理）
        forward_file_frame = ttk.LabelFrame(scrollable_frame, text="文件选择（正向批量处理）", padding=10)
        forward_file_frame.pack(fill=tk.X, padx=10, pady=5)

        self.forward_select_btn = ttk.Button(forward_file_frame, text="选择Excel文件（可多选）", command=self.forward_select_files)
        self.forward_select_btn.pack(side=tk.LEFT, padx=5)
        self.forward_clear_btn = ttk.Button(forward_file_frame, text="清空文件列表", command=self.forward_clear_files)
        self.forward_clear_btn.pack(side=tk.LEFT, padx=5)

        self.forward_file_listbox = tk.Listbox(forward_file_frame, height=5, selectmode=tk.EXTENDED)
        self.forward_file_listbox.pack(fill=tk.X, pady=5)

        # 2.6 处理参数区（正向批量处理）
        forward_process_frame = ttk.Frame(scrollable_frame, padding=5)
        forward_process_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(forward_process_frame, text="并发数:").pack(side=tk.LEFT, padx=5)
        self.forward_concurrent_var = tk.StringVar(value="2")
        forward_concurrent_options = ["1", "2", "3", "4", "5", "6", "8", "10"]
        ttk.Combobox(forward_process_frame, textvariable=self.forward_concurrent_var,
                     values=forward_concurrent_options, width=5, state="readonly").pack(side=tk.LEFT, padx=5)

        ttk.Label(forward_process_frame, text="请求间隔:").pack(side=tk.LEFT, padx=5)
        self.forward_interval_var = tk.StringVar(value="0.5秒")
        forward_interval_options = ["0秒", "0.5秒", "1秒", "2秒", "3秒"]
        ttk.Combobox(forward_process_frame, textvariable=self.forward_interval_var,
                     values=forward_interval_options, width=6, state="readonly").pack(side=tk.LEFT, padx=5)

        self.forward_process_btn = ttk.Button(forward_process_frame, text="开始处理", command=self.forward_start_processing)
        self.forward_process_btn.pack(side=tk.LEFT, padx=10)
        self.forward_stop_btn = ttk.Button(forward_process_frame, text="停止处理", command=self.forward_stop_processing, state=tk.DISABLED)
        self.forward_stop_btn.pack(side=tk.LEFT, padx=5)

        self.forward_progress_var = tk.DoubleVar()
        ttk.Progressbar(forward_process_frame, variable=self.forward_progress_var, maximum=100, length=400).pack(side=tk.LEFT, padx=10)
        self.forward_progress_label = ttk.Label(forward_process_frame, text="进度: 0%")
        self.forward_progress_label.pack(side=tk.LEFT, padx=5)

        # 3. 工作流返回结果日志区（放大显示）
        result_frame = ttk.LabelFrame(scrollable_frame, text="工作流返回结果", padding=10)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.forward_log_text = scrolledtext.ScrolledText(result_frame, height=24, wrap=tk.WORD)
        self.forward_log_text.pack(fill=tk.BOTH, expand=True)

    def _on_tab_changed(self, event):
        """TAB切换后强制刷新当前TAB的Canvas内容，修复空白不渲染问题"""
        selected = self.notebook.select()
        if not selected:
            return
        tab = self.notebook.nametowidget(selected)

        canvas = None
        canvas_window = None
        if tab is self.forward_tab:
            canvas = getattr(self, 'forward_canvas', None)
            canvas_window = getattr(self, 'forward_canvas_window', None)
        elif tab is self.agent_tab:
            canvas = getattr(self, 'agent_canvas', None)
            canvas_window = getattr(self, 'agent_canvas_window', None)
        else:
            canvas = getattr(self, 'reverse_canvas', None)
            canvas_window = getattr(self, 'reverse_canvas_window', None)

        self._active_canvas = canvas

        if canvas is None:
            return

        # 延迟到事件循环空闲时刷新，避免阻塞TAB切换（解决切换慢）
        self.root.after_idle(self._refresh_tab_canvas, canvas, canvas_window)

    def _refresh_tab_canvas(self, canvas, canvas_window):
        """刷新指定Canvas：强制嵌入容器宽度=canvas宽度（横向平铺）并更新滚动区域"""
        try:
            width = canvas.winfo_width()
            if canvas_window is not None and width > 1:
                # 强制滚动容器宽度与canvas一致，实现横向平铺
                canvas.itemconfig(canvas_window, width=width)
            canvas.configure(scrollregion=canvas.bbox("all"))
            # 若布局尚未完成（宽度未就绪），再排一次刷新兜底
            if width <= 1:
                self.root.after_idle(self._refresh_tab_canvas, canvas, canvas_window)
        except tk.TclError:
            pass

    def _on_mousewheel(self, event):
        """鼠标滚轮滚动当前激活TAB的滚动区域（兼容 macOS delta=±1/±2 与 Windows delta=±120）"""
        canvas = self._active_canvas
        if canvas is None or not event.delta:
            return
        delta = event.delta
        # macOS 滚轮 delta 为 ±1/±2（非 Windows 的 ±120），直接按单位滚动；Windows 换算为 1 单位
        units = -delta if abs(delta) < 120 else int(-1 * (delta / 120))
        canvas.yview_scroll(units, "units")

    def forward_log(self, message):
        """向正向问答TAB页的日志区写入日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.forward_log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.forward_log_text.see(tk.END)
        self.root.update()
        _runtime_log(message)

    def forward_select_files(self):
        """正向模式：选择文件"""
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.forward_selected_files:
                    self.forward_selected_files.append(file)
                    self.forward_file_listbox.insert(tk.END, file)
            self.forward_log(f"已选择 {len(files)} 个文件")

    def forward_clear_files(self):
        """正向模式：清空文件列表"""
        self.forward_selected_files = []
        self.forward_file_listbox.delete(0, tk.END)
        self.forward_log("已清空文件列表")

    # ==================== 文件处理 Agent（Phase 1） ====================

    def create_agent_tab(self, parent):
        """创建【文件处理 Agent】TAB —— 自然语言驱动文件处理（识别需求→确认→执行）"""
        main_canvas = tk.Canvas(parent)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        canvas_window = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        self.agent_canvas = main_canvas
        self.agent_canvas_window = canvas_window

        def configure_scroll_region(event):
            main_canvas.itemconfig(canvas_window, width=event.width)

        main_canvas.bind("<Configure>", configure_scroll_region)
        main_canvas.configure(yscrollcommand=scrollbar.set)
        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 状态初始化
        self.agent_selected_files = []
        self.agent_running = False
        self.agent_pending_plan = None
        self.agent_log_lock = threading.Lock()
        self.agent_thread_config = None
        # 懒加载 Agent 图：避免 GUI 启动阶段被 LangGraph 依赖链阻塞
        self.agent_graph = None
        self.agent_graph_error = None
        # Agent Loop 交互状态（澄清/理解确认/结果反馈）
        self.agent_phase = 'idle'          # idle / analyze / clarify / understanding / confirm / executing / feedback
        self.agent_conversation = []       # 多轮对话历史（跨轮 Context）
        self.agent_clarify_questions = []  # 当前待澄清问题
        self.agent_last_summary = ''       # 上一轮结果摘要（供反馈循环）

        # 1. 文件选择区
        file_frame = ttk.LabelFrame(scrollable_frame, text="文件选择", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        self.agent_select_btn = ttk.Button(file_frame, text="选择Excel文件（可多选）", command=self.agent_select_files)
        self.agent_select_btn.pack(side=tk.LEFT, padx=5)
        self.agent_clear_btn = ttk.Button(file_frame, text="清空文件列表", command=self.agent_clear_files)
        self.agent_clear_btn.pack(side=tk.LEFT, padx=5)
        self.agent_file_listbox = tk.Listbox(file_frame, height=5, selectmode=tk.EXTENDED)
        self.agent_file_listbox.pack(fill=tk.X, pady=5)

        # 2. 意图识别 API 配置区
        config_frame = ttk.LabelFrame(scrollable_frame, text="意图识别 API配置（Dify File Agent Intent 应用）", padding=10)
        config_frame.pack(fill=tk.X, padx=10, pady=5)
        config_frame.columnconfigure(1, weight=1)
        ttk.Label(config_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.agent_api_url_entry = ttk.Entry(config_frame, width=50)
        self.agent_api_url_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        self.agent_api_url_entry.insert(0, "http://10.0.173.34/v1")
        ttk.Label(config_frame, text="API Key:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.agent_api_key_entry = ttk.Entry(config_frame, width=50)
        self.agent_api_key_entry.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.agent_api_key_entry.insert(0, "app-AQPiQv2xY82RFFzKvXnUOGbY")
        self._setup_apikey_entry(self.agent_api_key_entry)

        # 2.5 代码生成 API 配置区（Phase 5 动态代码）
        codegen_frame = ttk.LabelFrame(scrollable_frame, text="代码生成 API配置（Dify File Agent Code Generator 应用，动态代码用）", padding=10)
        codegen_frame.pack(fill=tk.X, padx=10, pady=5)
        codegen_frame.columnconfigure(1, weight=1)
        ttk.Label(codegen_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.agent_codegen_url_entry = ttk.Entry(codegen_frame, width=50)
        self.agent_codegen_url_entry.grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        self.agent_codegen_url_entry.insert(0, "http://10.0.173.34/v1")
        ttk.Label(codegen_frame, text="API Key:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.agent_codegen_key_entry = ttk.Entry(codegen_frame, width=50)
        self.agent_codegen_key_entry.grid(row=1, column=1, padx=5, pady=2, sticky="ew")
        self.agent_codegen_key_entry.insert(0, "app-5sQe8aFxxLGqSIsValIiB7tY")
        self._setup_apikey_entry(self.agent_codegen_key_entry)

        # 3. 需求描述区
        desc_frame = ttk.LabelFrame(scrollable_frame, text="用户需求描述（自然语言）", padding=10)
        desc_frame.pack(fill=tk.X, padx=10, pady=5)
        self.agent_request_text = tk.Text(desc_frame, height=4, wrap=tk.WORD)
        self.agent_request_text.pack(fill=tk.X)

        # 4. 操作区
        action_frame = ttk.Frame(scrollable_frame, padding=5)
        action_frame.pack(fill=tk.X, padx=10, pady=5)
        self.agent_analyze_btn = ttk.Button(action_frame, text="分析需求", command=self.agent_analyze)
        self.agent_analyze_btn.pack(side=tk.LEFT, padx=5)
        self.agent_confirm_btn = tk.Button(action_frame, text="确认执行", command=self.agent_confirm_execute,
                                           state=tk.DISABLED, bg='#e0e0e0', activebackground='#d0d0d0',
                                           relief=tk.RAISED, padx=10)
        self.agent_confirm_btn.pack(side=tk.LEFT, padx=5)
        self.agent_cancel_btn = ttk.Button(action_frame, text="取消", command=self.agent_cancel_plan, state=tk.DISABLED)
        self.agent_cancel_btn.pack(side=tk.LEFT, padx=5)

        # 4.5 交互区（澄清回答 / 理解确认 / 符号校正 / 结果反馈，按需显隐）
        interact_frame = ttk.LabelFrame(scrollable_frame, text="Agent 交互", padding=10)
        interact_frame.pack(fill=tk.X, padx=10, pady=5)
        # 澄清行
        self.agent_clarify_label = ttk.Label(interact_frame, text="", wraplength=680)
        self.agent_clarify_entry = ttk.Entry(interact_frame, width=80)
        self.agent_clarify_submit_btn = ttk.Button(interact_frame, text="好的，提交", command=self._submit_clarify)
        # 理解确认行
        self.agent_understand_label = ttk.Label(interact_frame, text="", wraplength=680)
        self.agent_understand_ok_btn = ttk.Button(interact_frame, text="没问题，继续", command=self._on_understand_ok)
        self.agent_revise_btn = ttk.Button(interact_frame, text="修改一下", command=self._on_revise)
        # 符号校正确认行（find_check）
        self.agent_find_label = ttk.Label(interact_frame, text="", wraplength=680)
        self.agent_fix_confirm_btn = ttk.Button(interact_frame, text="就用这个，继续", command=self._on_find_fix_confirm)
        # 结果反馈行
        self.agent_satisfied_btn = ttk.Button(interact_frame, text="结果满意", command=self._on_satisfied)
        self.agent_feedback_btn = ttk.Button(interact_frame, text="提出修改意见", command=self._on_feedback_submit)
        self._hide_interaction()

        # 5. 方案预览区
        plan_frame = ttk.LabelFrame(scrollable_frame, text="Agent 执行方案", padding=10)
        plan_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.agent_plan_text = scrolledtext.ScrolledText(plan_frame, height=8, wrap=tk.WORD, state=tk.DISABLED)
        self.agent_plan_text.pack(fill=tk.BOTH, expand=True)

        # 6. 日志区
        log_frame = ttk.LabelFrame(scrollable_frame, text="执行日志", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.agent_log_text = scrolledtext.ScrolledText(log_frame, height=12, wrap=tk.WORD)
        self.agent_log_text.pack(fill=tk.BOTH, expand=True)

        # 处理间隙动态 Loading（LLM 调用/程序处理期间显示，横向波浪点 + 阶段文案）
        self.agent_loading_frame = ttk.Frame(log_frame)
        try:
            # 尽量透明（macOS Tk 支持 systemTransparent）；不支持则回退系统控件底色与容器融合
            self.agent_loading_canvas = tk.Canvas(self.agent_loading_frame, width=64, height=18,
                                                  bg='systemTransparent', highlightthickness=0)
        except tk.TclError:
            self.agent_loading_canvas = tk.Canvas(self.agent_loading_frame, width=64, height=18,
                                                  bg='SystemButtonFace', highlightthickness=0)
        self.agent_loading_canvas.pack(side=tk.LEFT, padx=(0, 10))
        self.agent_loading_label = ttk.Label(self.agent_loading_frame, text="正在处理，请稍候...",
                                             foreground='#666666')
        self.agent_loading_label.pack(side=tk.LEFT)
        self.agent_loading_frame.pack(fill=tk.X, pady=(4, 0))   # 先 pack 再隐藏，保持布局一致
        self.agent_loading_frame.pack_forget()

    def agent_select_files(self):
        """Agent：选择文件"""
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.csv"), ("所有文件", "*.*")]
        )
        if files:
            for f in files:
                if f not in self.agent_selected_files:
                    self.agent_selected_files.append(f)
                    self.agent_file_listbox.insert(tk.END, f)
            self.agent_log(f"已选择 {len(files)} 个文件")

    def agent_clear_files(self):
        """Agent：清空文件列表"""
        self.agent_selected_files = []
        self.agent_file_listbox.delete(0, tk.END)
        self.agent_log("已清空文件列表")

    def agent_log(self, message):
        """Agent：写日志（线程安全）"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        with self.agent_log_lock:
            self.agent_log_text.insert(tk.END, f"[{timestamp}] {message}\n")
            self.agent_log_text.see(tk.END)
            self.root.update()

    # ---- 处理间隙动态 Loading（LLM 调用/程序处理期间显示） ----

    def _start_loading(self, text='正在处理，请稍候...'):
        """显示 Loading 并启动波浪动画（须在主线程调用）"""
        self.agent_loading_label.config(text=text)
        self.agent_loading_frame.pack(fill=tk.X, pady=(4, 0))
        self._loading_on = True
        self._loading_phase = 0.0
        self._loading_job = None
        self._draw_loading_wave()

    def _draw_loading_wave(self):
        """Canvas 横向波浪点：5 点横向排列、正弦起伏且大小同步变化（柔和打字式波浪）"""
        if not getattr(self, '_loading_on', False):
            return
        import math
        c = self.agent_loading_canvas
        c.delete('all')
        n, spacing, cy = 5, 10, 9
        start_x = (64 - spacing * (n - 1)) / 2
        for i in range(n):
            phase = self._loading_phase + i * (2 * math.pi / n)
            wave = math.sin(phase)
            y = cy - wave * 4                                # 上下起伏（幅度 4px）
            r = 2.2 + 1.2 * ((wave + 1) / 2)                 # 半径随起伏同步变化（2.2~3.4）
            x = start_x + i * spacing
            c.create_oval(x - r, y - r, x + r, y + r, fill='#4a90ff', outline='')
        self._loading_phase += 0.3
        self._loading_job = self.root.after(100, self._draw_loading_wave)

    def _stop_loading(self):
        """停止动画并隐藏（须在主线程调用）"""
        self._loading_on = False
        job = getattr(self, '_loading_job', None)
        if job is not None:
            try:
                self.root.after_cancel(job)
            except Exception:
                pass
            self._loading_job = None
        self.agent_loading_canvas.delete('all')
        self.agent_loading_frame.pack_forget()

    def _ensure_agent_graph(self):
        """按需初始化 LangGraph，避免将其放在主窗口启动路径上。"""
        if self.agent_graph is not None:
            return True
        try:
            self.agent_log("正在初始化文件处理 Agent 引擎...")
            self.agent_graph = build_agent_graph(self)
            self.agent_graph_error = None
            self.agent_log("文件处理 Agent 引擎初始化完成")
            return True
        except Exception as e:
            self.agent_graph = None
            self.agent_graph_error = str(e)
            _runtime_log(f"Agent 图初始化失败: {str(e)}")
            self.agent_log(f"✗ 文件处理 Agent 引擎初始化失败: {str(e)}")
            messagebox.showerror(
                "Agent 初始化失败",
                "文件处理 Agent 初始化失败，主界面仍可继续使用。\n\n"
                f"详情：{str(e)}\n\n"
                "请查看 运行日志.log 或重新打包后重试。"
            )
            return False

    def _agent_write_log(self, task_id, section, data):
        """执行日志落盘：追加 JSONL 行到 agent_workspace/<task_id>/execution.log"""
        if not task_id:
            return
        ws_dir = os.path.join(_resolve_agent_workspace_root(), task_id)
        try:
            os.makedirs(ws_dir, exist_ok=True)
            entry = {
                'ts': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'section': section,
                'data': data,
            }
            with open(os.path.join(ws_dir, 'execution.log'), 'a', encoding='utf-8') as fh:
                fh.write(json.dumps(entry, ensure_ascii=False) + '\n')
        except Exception:
            pass

    def _agent_set_plan(self, text, buttons_enabled):
        """Agent：更新方案预览区与按钮状态"""
        self.agent_plan_text.config(state=tk.NORMAL)
        self.agent_plan_text.delete('1.0', tk.END)
        self.agent_plan_text.insert(tk.END, text)
        self.agent_plan_text.config(state=tk.DISABLED)
        state = tk.NORMAL if buttons_enabled else tk.DISABLED
        self.agent_confirm_btn.config(state=state)
        self.agent_cancel_btn.config(state=state)
        if buttons_enabled:
            self._start_confirm_pulse()
        else:
            self._stop_confirm_pulse()

    # ---- Agent Loop 交互辅助（澄清 / 理解确认 / 结果反馈） ----

    def _hide_interaction(self):
        """隐藏交互区全部控件"""
        for w in (self.agent_clarify_label, self.agent_clarify_entry, self.agent_clarify_submit_btn,
                  self.agent_understand_label, self.agent_understand_ok_btn, self.agent_revise_btn,
                  self.agent_find_label, self.agent_fix_confirm_btn,
                  self.agent_satisfied_btn, self.agent_feedback_btn):
            w.pack_forget()

    def _pack_interaction(self, widgets, side=tk.TOP):
        for w in widgets:
            w.pack(side=side, padx=5, pady=2, fill=tk.X)

    def _agent_task_id(self):
        return (self.agent_thread_config or {}).get('configurable', {}).get('thread_id', '')

    def _resume_graph(self, command=None):
        """恢复图执行（澄清/理解确认用 Command(resume=...)，执行确认用 None），并继续处理阶段1"""
        if not self._ensure_agent_graph():
            return
        self.agent_running = True

        def worker():
            try:
                self.root.after(0, lambda: self._start_loading("正在继续分析，请稍候..."))
                partial = self.agent_graph.invoke(
                    command, config=self.agent_thread_config,
                    interrupt_before=["execute_builtin", "sandbox_execute"])
                self._handle_stage1(partial)
            except Exception as e:
                self.agent_log(f"✗ 处理失败: {str(e)}")
                self._agent_set_plan(f"处理失败: {str(e)}", False)
            finally:
                self.root.after(0, self._stop_loading)
                self.agent_running = False
                self.agent_analyze_btn.config(state=tk.NORMAL)

        threading.Thread(target=worker, daemon=True).start()

    def _file_structure_text(self, partial):
        """生成文件列结构文本（供澄清/理解确认展示），无元数据时返回空串"""
        metas = partial.get('file_metadata') or []
        if not metas:
            return ''
        return _column_structure_text(metas)

    def _describe_plan_plain(self, task, params, metas):
        """把任务+参数翻译成客服式口语（兜底：friendly_summary 缺失/截断时用）"""
        fname = (metas[0].get('basename') if metas else '文件')
        if task == 'modify_excel':
            ops_desc = []
            for op in params.get('ops') or []:
                if not isinstance(op, dict):
                    continue
                col = op.get('column', '')
                if op.get('type') == 'replace':
                    find, repl = op.get('find', ''), op.get('replace', '')
                    ops_desc.append(f"把 {col} 列里出现的『{find}』这几个字{'去掉' if repl == '' else '换成『' + repl + '』'}")
                elif op.get('type') == 'delete_rows':
                    ops_desc.append(f"删除 {col} 列等于『{op.get('equals', '')}』的行")
            if ops_desc:
                tail = ('改好的内容会直接覆盖保存回原文件，原来的内容将无法恢复，建议先备份原件。'
                        if str(params.get('save_mode') or '').strip().lower() == 'overwrite'
                        else '改好的内容会保存成一个新文件，您原来的文件不会被动。')
                return ('我准备这样处理「' + fname + '」：\n' +
                        '\n'.join('  · ' + d for d in ops_desc) +
                        '\n' + tail)
        if task == 'filter_excel':
            op = params.get('operator', 'equals')
            op_text = {'equals': '等于', 'not_equals': '不等于', 'contains': '包含',
                       'gte': '大于等于', 'lte': '小于等于', 'gt': '大于', 'lt': '小于'}.get(op, op)
            return f"我会筛选出 {params.get('column', '')} 列{op_text}『{params.get('value', '')}』的行，单独存成一个新文件。"
        if task == 'statistics':
            return "我会帮您统计这份文件的匹配占比、成功率、渠道分布等数据，生成统计结果。"
        if task == 'export_failed':
            return "我会把失败项整理导出成一个新文件，方便您单独查看。"
        if task == 'export_csv':
            return f"我会把指定列（{params.get('columns', '')}）导出成一个 CSV 文件。"
        if task == 'diff':
            return "我会对比两个文件的 O 列渠道差异。"
        if task == 'sort_excel':
            return f"我会按 {params.get('column', '')} 列{'升序' if not params.get('desc') else '降序'}重新排列数据。"
        if task == 'group_excel':
            return f"我会按 {params.get('column', '')} 列把数据分组到不同的 Sheet 里。"
        if task == 'split_excel':
            return f"我会把文件按每 {params.get('rows_per_sheet', 500)} 行拆分成多个 Sheet。"
        # 兜底：未翻译的任务保留原 JSON
        return f"我会按识别出的参数来处理：{json.dumps(params, ensure_ascii=False)}"

    def _intent_friendly_text(self, partial):
        """将意图识别结果整理为客服式口语（LLM friendly_summary 优先，代码兜底）"""
        intent = partial.get('intent') or {}
        task = intent.get('task', '')
        params = intent.get('params') or {}
        understanding = intent.get('understanding') or ''
        friendly = intent.get('friendly_summary') or ''
        reason = intent.get('reason') or ''
        questions = intent.get('clarify_questions') or []
        metas = partial.get('file_metadata') or []
        lines = []
        # 需求理解（LLM 生成；兜底：custom 用 reason，否则用需求文本）
        if understanding:
            lines.append(understanding)
        elif task == 'custom':
            lines.append(reason or '这个需求比较特殊，内置工具处理不了，我需要写一段代码来完成。')
        else:
            lines.append('根据您的需求，我来帮您处理文件。')
        # 方案口语（LLM friendly_summary 优先，代码兜底）
        if task == 'custom':
            lines.append('（这个需求需要动态编写代码来执行）')
        elif friendly:
            lines.append(friendly)
        else:
            lines.append(self._describe_plan_plain(task, params, metas))
        # 待澄清问题
        if questions:
            lines.append('不过有几个地方想先跟您确认一下：')
            for i, q in enumerate(questions, 1):
                lines.append(f"{i}. {q}")
        return '\n'.join(lines)

    def _show_clarify(self, partial, questions):
        """Loop①澄清交互：展示友好意图结果 + 文件列结构 + 澄清问题，等待用户逐题回答（多题用 | 分隔）"""
        self.agent_phase = 'clarify'
        self.agent_clarify_questions = questions
        lines = [self._intent_friendly_text(partial), '']
        col_text = self._file_structure_text(partial)
        if col_text:
            lines += ['文件列结构：', col_text, '']
        lines.append('为了更准确地帮您处理，想跟您确认几个小问题：')
        for i, q in enumerate(questions, 1):
            lines.append(f"{i}. {q}")
        lines.append('')
        lines.append('直接在输入框里回答就行；如果有多个问题，用 | 分隔，然后点「好的，提交」。')
        self._agent_set_plan('\n'.join(lines), False)
        self.agent_clarify_label.config(text='您的回答（多个问题用 | 分隔）：')
        self.agent_clarify_entry.delete(0, tk.END)
        self._pack_interaction([self.agent_clarify_label, self.agent_clarify_entry, self.agent_clarify_submit_btn])

    def _submit_clarify(self):
        """用户提交澄清回答 → 带值恢复图执行"""
        from langgraph.types import Command
        if self.agent_phase != 'clarify':
            return
        raw = self.agent_clarify_entry.get().strip()
        questions = self.agent_clarify_questions
        if not questions:
            return
        if len(questions) > 1 and '|' not in raw:
            messagebox.showwarning("警告", f"有 {len(questions)} 个问题需要回答，请用 | 分隔多个回答")
            return
        parts = [p.strip() for p in raw.split('|')] if raw else ['']
        while len(parts) < len(questions):
            parts.append('')
        answers = {q: (parts[i] if i < len(parts) else '') for i, q in enumerate(questions)}
        self.agent_log("用户提交澄清回答，重新分析需求")
        self._agent_write_log(self._agent_task_id(), 'clarify', {'questions': questions, 'answers': answers})
        self._hide_interaction()
        self.agent_phase = 'analyze'
        self._resume_graph(Command(resume=answers))

    def _show_understanding_wait(self, partial, understanding):
        """Loop②理解确认交互：展示友好意图结果 + 文件列结构，等待用户确认/修正"""
        self.agent_phase = 'understanding'
        self.agent_log("请用户确认对需求的理解")
        text = self._intent_friendly_text(partial)
        col_text = self._file_structure_text(partial)
        if col_text:
            text += f"\n\n文件列结构：\n{col_text}"
        text += "\n\n麻烦您看看我理解得对不对：如果没问题，点「没问题，继续」；要是哪里不对，点「修改一下」，我马上调整。"
        self._agent_set_plan(text, False)
        self.agent_understand_label.config(text=f"我的理解：{understanding}")
        self._pack_interaction([self.agent_understand_label, self.agent_understand_ok_btn, self.agent_revise_btn])

    def _on_understand_ok(self):
        """用户确认理解正确 → 继续规划"""
        from langgraph.types import Command
        if self.agent_phase != 'understanding':
            return
        self.agent_log("用户确认理解正确")
        self._agent_write_log(self._agent_task_id(), 'understanding', {'ok': True})
        self._hide_interaction()
        self.agent_phase = 'analyze'
        self._resume_graph(Command(resume={'ok': True}))

    def _ask_multiline(self, title, prompt):
        """宽大多行输入对话框（替代 simpledialog 单行小框）；返回输入文本，取消返回 None"""
        result = {'value': None}
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(True, True)
        dlg.geometry('620x340')
        dlg.minsize(480, 260)

        frame = ttk.Frame(dlg, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=prompt, wraplength=560).pack(anchor=tk.W, pady=(0, 6))
        text = scrolledtext.ScrolledText(frame, width=72, height=10, wrap=tk.WORD)
        text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        text.focus_set()

        def on_ok():
            result['value'] = text.get('1.0', tk.END).strip()
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="确定", command=on_ok).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btn_row, text="取消", command=on_cancel).pack(side=tk.RIGHT)
        dlg.bind('<Escape>', lambda e: on_cancel())
        dlg.bind('<Control-Return>', lambda e: on_ok())   # Ctrl+Enter 快捷提交
        dlg.protocol('WM_DELETE_WINDOW', on_cancel)

        self.root.wait_window(dlg)
        return result['value']

    def _confirm_overwrite(self, fname):
        """覆盖原文件前的红色警醒确认（覆盖不可逆、提示备份）；返回 True 继续 / False 取消"""
        result = {'ok': False}
        dlg = tk.Toplevel(self.root)
        dlg.title("⚠ 覆盖原文件确认")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.geometry('500x280')
        dlg.minsize(460, 260)

        head = tk.Label(dlg, text="⚠ 即将覆盖原文件", fg='#cc0000', bg='#fff0f0',
                        font=('', 14, 'bold'), pady=8)
        head.pack(fill=tk.X)

        body = ttk.Frame(dlg, padding=14)
        body.pack(fill=tk.BOTH, expand=True)
        msg = (f"您选择将修改结果直接保存回原文件：\n\n「{fname}」\n\n"
               f"⚠ 此操作【不可逆】：原文件内容将被覆盖，无法恢复！\n"
               f"请确认已提前备份好原件，再继续操作。")
        ttk.Label(body, text=msg, wraplength=440, foreground='#cc0000').pack(anchor=tk.W, pady=(0, 14))

        btn_row = ttk.Frame(body)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="我已知晓并已备份，继续覆盖",
                   command=lambda: (result.update(ok=True), dlg.destroy())).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btn_row, text="取消", command=dlg.destroy).pack(side=tk.RIGHT)
        dlg.protocol('WM_DELETE_WINDOW', dlg.destroy)
        dlg.bind('<Escape>', lambda e: dlg.destroy())

        self.root.wait_window(dlg)
        return result['ok']

    def _on_revise(self):
        """用户修正需求：understanding 阶段带着修正重新理解（Loop②重入）；find_check 阶段重新分析"""
        from langgraph.types import Command
        if self.agent_phase == 'understanding':
            revision = self._ask_multiline("修改一下", "请补充或修正您的需求：")
            if revision is None:
                return
            revision = revision.strip()
            self._hide_interaction()
            self.agent_log(f"用户修正需求: {revision[:120]}")
            self._agent_write_log(self._agent_task_id(), 'understanding', {'ok': False, 'revision': revision})
            if not revision:
                # 修正为空 → 视为确认理解，按原需求继续
                self.agent_phase = 'analyze'
                self._resume_graph(Command(resume={'ok': True}))
                return
            self.agent_phase = 'analyze'
            self._resume_graph(Command(resume={'ok': False, 'revision': revision}))
            return
        if self.agent_phase == 'find_check':
            # 查找内容未命中：请用户提供文件中查找内容的实际写法，然后重新分析
            revision = self._ask_multiline("修改一下", "请告诉我文件里查找内容的实际写法：")
            if revision is None:
                return
            revision = revision.strip()
            if not revision:
                return
            self._hide_interaction()
            self.agent_log(f"用户修正需求: {revision[:120]}，重新分析")
            current = self.agent_request_text.get('1.0', tk.END).strip()
            self.agent_request_text.delete('1.0', tk.END)
            self.agent_request_text.insert(tk.END, current + f"\n（修正：{revision}）")
            self.agent_phase = 'idle'
            self.agent_analyze()

    def _show_feedback(self, final):
        """Loop④结果反馈：展示结果摘要 + 自检结论，等待用户满意/提出修改意见"""
        self.agent_phase = 'feedback'
        status = final.get('status', 'done')
        lines = []
        if status == 'failed':
            lines.append("✗ 执行未完成")
            lines.append(f"错误: {final.get('error', '未知错误')}")
            lines.append("可点击「提出修改意见」重新分析，或修改需求后重试。")
        else:
            lines.append("✓ 执行完成")
            for o in final.get('output_files') or []:
                lines.append(f"输出: {o}")
            v = final.get('verification') or {}
            if v:
                lines.append(f"结果自检: {'通过' if v.get('ok') else '未通过'} - {v.get('reason', '')}")
                if not v.get('ok'):
                    lines.append("若结果不符合预期，可点击「提出修改意见」，我将带着意见重新处理。")
        self.agent_last_summary = '\n'.join(lines)
        self._agent_set_plan('\n'.join(lines), False)
        self._pack_interaction([self.agent_satisfied_btn, self.agent_feedback_btn], side=tk.LEFT)

    def _on_satisfied(self):
        """用户确认结果满意 → 结束本轮"""
        if self.agent_phase != 'feedback':
            return
        self.agent_log("用户确认结果满意")
        self._agent_write_log(self._agent_task_id(), 'feedback', {'satisfied': True})
        self._hide_interaction()
        self.agent_phase = 'idle'

    def _on_feedback_submit(self):
        """用户提出修改意见 → 带上下文启动新一轮分析（Loop④重入）"""
        if self.agent_phase != 'feedback':
            return
        feedback = self._ask_multiline("提出修改意见", "请描述需要修改的地方：")
        if feedback is None:
            return
        feedback = feedback.strip()
        if not feedback:
            return
        self._hide_interaction()
        self.agent_log(f"用户反馈: {feedback[:120]}，携带上下文重新分析")
        self._agent_write_log(self._agent_task_id(), 'feedback', {'satisfied': False, 'feedback': feedback})
        # 记录多轮上下文（供下一轮 prompt 注入）
        self.agent_conversation.append({'role': 'assistant', 'content': self.agent_last_summary})
        self.agent_conversation.append({'role': 'user', 'content': f"修改意见：{feedback}"})
        self.agent_conversation = self.agent_conversation[-6:]  # 截断
        # 需求框追加修改意见，供用户查看/编辑
        current = self.agent_request_text.get('1.0', tk.END).strip()
        self.agent_request_text.delete('1.0', tk.END)
        self.agent_request_text.insert(tk.END, current + f"\n（修改意见：{feedback}）")
        self.agent_phase = 'idle'
        self.agent_analyze()

    # ---- 确认按钮呼吸灯闪烁 ----

    def _start_confirm_pulse(self):
        """启动确认按钮呼吸灯动画（正弦渐变 浅灰↔浅红）"""
        self._pulse_on = True
        self._pulse_phase = 0.0
        self._pulse_job = self.root.after(60, self._confirm_pulse_tick)

    def _confirm_pulse_tick(self):
        if not getattr(self, '_pulse_on', False):
            return
        import math
        self._pulse_phase += 0.15
        brightness = (math.sin(self._pulse_phase) + 1) / 2  # 0..1
        base = (0.95, 0.95, 0.95)   # 浅灰
        hl = (1.0, 0.35, 0.35)      # 浅红
        r = int((base[0] + (hl[0] - base[0]) * brightness) * 255)
        g = int((base[1] + (hl[1] - base[1]) * brightness) * 255)
        b = int((base[2] + (hl[2] - base[2]) * brightness) * 255)
        color = f'#{r:02x}{g:02x}{b:02x}'
        try:
            self.agent_confirm_btn.config(bg=color, activebackground=color)
            self._pulse_job = self.root.after(60, self._confirm_pulse_tick)
        except tk.TclError:
            pass

    def _stop_confirm_pulse(self):
        """停止呼吸灯动画并恢复默认背景色"""
        self._pulse_on = False
        job = getattr(self, '_pulse_job', None)
        if job is not None:
            try:
                self.root.after_cancel(job)
            except Exception:
                pass
            self._pulse_job = None
        if hasattr(self, 'agent_confirm_btn'):
            try:
                self.agent_confirm_btn.config(bg='#e0e0e0', activebackground='#d0d0d0')
            except tk.TclError:
                pass

    def agent_analyze(self):
        """Agent：阶段1 —— LangGraph 运行 inspect→intent→planner→route，中断于执行前，展示方案"""
        if self.agent_running:
            return
        if not self.agent_selected_files:
            messagebox.showwarning("警告", "请先选择要处理的文件！")
            return
        api_url = self.agent_api_url_entry.get().strip()
        api_key = self.agent_api_key_entry.get().strip()
        if not api_url or not api_key:
            messagebox.showwarning("警告", "请填写意图识别应用的 API 配置！")
            return
        request = self.agent_request_text.get('1.0', tk.END).strip()
        if not request:
            messagebox.showwarning("警告", "请描述你的需求！")
            return
        if not self._ensure_agent_graph():
            return

        self.agent_running = True
        self.agent_analyze_btn.config(state=tk.DISABLED)
        self._stop_confirm_pulse()
        self.agent_log("正在识别需求（inspect → 意图识别 → 规划）...")
        thread_id = str(uuid.uuid4())
        self.agent_thread_config = {"configurable": {"thread_id": thread_id}}
        self._agent_write_log(thread_id, 'start', {
            'user_request': request,
            'input_files': list(self.agent_selected_files),
            'start_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
        initial = {
            'user_request': request,
            'input_files': list(self.agent_selected_files),
            'logs': [],
            'status': 'planning',
            'task_id': thread_id,
            'conversation': list(self.agent_conversation),  # 多轮上下文（Loop④）
        }

        def worker():
            try:
                self.root.after(0, lambda: self._start_loading("正在识别需求并制定方案，请稍候..."))
                partial = self.agent_graph.invoke(
                    initial, config=self.agent_thread_config,
                    interrupt_before=["execute_builtin", "sandbox_execute"])
                self._handle_stage1(partial)
            except Exception as e:
                self.agent_log(f"✗ 分析失败: {str(e)}")
                self._agent_set_plan(f"分析失败: {str(e)}", False)
            finally:
                self.root.after(0, self._stop_loading)
                self.agent_running = False
                self.agent_analyze_btn.config(state=tk.NORMAL)

        threading.Thread(target=worker, daemon=True).start()

    def _handle_stage1(self, partial):
        """处理阶段1结果：循环处理澄清/理解确认中断，最终展示方案或等待用户交互"""
        for lg in partial.get('logs') or []:
            self.agent_log(lg)
        snapshot = self.agent_graph.get_state(self.agent_thread_config)
        next_nodes = tuple(snapshot.next or ())
        intent = partial.get('intent') or {}
        questions = intent.get('clarify_questions') or []
        task_id = self._agent_task_id()
        self._agent_write_log(task_id, 'stage1', {
            'intent': intent,
            'plan': partial.get('plan'),
            'route': partial.get('route'),
            'file_metadata': partial.get('file_metadata'),
            'next': list(next_nodes),
        })
        if 'clarify_interrupt' in next_nodes:
            # Loop①：需要澄清
            self._show_clarify(partial, questions)
            return
        if 'show_understanding' in next_nodes:
            # Loop②：理解确认 —— 始终展示需求确认视图，等待用户确认/修正
            understanding = intent.get('understanding', '')
            if not understanding:
                # 兜底：截断/旧格式导致 understanding 缺失时用 reason 或需求文本
                understanding = (intent.get('reason') or
                                 f"根据你的需求处理文件：{(partial.get('user_request') or '')[:120]}")
            self.agent_log(f"我的理解：{understanding}")
            self._show_understanding_wait(partial, understanding)
            return
        # 已通过澄清/理解，停在执行前中断点（interrupt_before）→ 展示方案
        self._show_plan_or_dynamic(partial)

    def _show_plan_or_dynamic(self, partial):
        """展示最终执行方案（builtin 或 dynamic_code），等待用户确认执行"""
        plan = partial.get('plan') or {}
        route = partial.get('route') or 'builtin'
        tool = plan.get('tool') or ''
        params = plan.get('params') or {}
        if route == 'dynamic_code' or tool == 'custom':
            self._show_dynamic_plan(partial)
            return
        self.agent_pending_plan = {'task': tool, 'params': params, 'files': list(self.agent_selected_files)}
        # 符号预检：replace 类 find 未命中（含变体/无变体）→ 先与用户确认，再展示方案
        issues = self._check_replace_finds(partial)
        if issues:
            self._show_find_check(partial, issues)
            return
        self._render_plan_view(partial)

    def _render_plan_view(self, partial):
        """口语化渲染 builtin 方案预览（LLM friendly_summary 优先，代码兜底），并启用「确认执行」"""
        plan = partial.get('plan') or {}
        tool = plan.get('tool') or ''
        params = plan.get('params') or {}
        intent = partial.get('intent') or {}
        understanding = intent.get('understanding', '')
        friendly = intent.get('friendly_summary') or ''
        lines = []
        if understanding:
            lines.append(understanding)
        lines.append(friendly or self._describe_plan_plain(tool, params, partial.get('file_metadata') or []))
        for m in partial.get('file_metadata') or []:
            lines.append(f"文件: {m.get('basename')} | {m.get('type')} | 行{m.get('rows')} 列{m.get('columns')}")
        lines.append(f"风险等级: {plan.get('risk_level', 'low')}")
        self.agent_log(f"识别成功: task={tool}, params={json.dumps(params, ensure_ascii=False)}")
        self._agent_set_plan('\n'.join(lines), True)

    def _check_replace_finds(self, partial):
        """对 plan 中 replace 类 op 做 find 存在性预检；返回未命中的 issues（含变体候选）"""
        ops = ((partial.get('plan') or {}).get('params') or {}).get('ops')
        if not isinstance(ops, list):
            return []
        files = list(self.agent_selected_files)
        if not files or detect_file_type(files[0]) not in ('xlsx', 'xls'):
            return []
        issues = []
        try:
            for i, op in enumerate(ops):
                if not isinstance(op, dict) or op.get('type') != 'replace':
                    continue
                find = str(op.get('find', '') or '')
                if not find:
                    continue
                result = check_find_in_column(files[0], op.get('column', 1), find)
                if not result['found']:
                    issues.append({'op_index': i, 'column': op.get('column', 1),
                                   'find': find, 'variants': result['variants']})
        except Exception as e:
            self.agent_log(f"方案预检（符号检测）跳过: {str(e)}")
            return []
        return issues

    def _show_find_check(self, partial, issues):
        """符号差异确认交互：展示检测详情 + 沟通话术，等待用户校正或修改需求"""
        self.agent_phase = 'find_check'
        self.agent_find_issues = issues
        self.agent_find_partial = partial
        issue = issues[0]
        find = issue['find']
        col = issue['column']
        variants = issue['variants']
        detail = f"我在文件里检查了一下，您要找的『{find}』（{col} 列）没有直接找到。\n"
        if variants:
            v = variants[0]
            detail += f"不过文件里有写法很接近的『{v['find']}』，一共出现 {v['count']} 处。\n"
            talk = (f"我检查了一下文件，没找到『{find}』，不过我注意到文件里其实有"
                    f"『{v['find']}』（出现 {v['count']} 处），看起来很像您要找的内容。\n"
                    f"要不要我按『{v['find']}』来处理？")
        else:
            detail += "而且也没找到写法相近的内容，可能需要您确认一下查找内容。\n"
            talk = (f"我在文件里没有找到『{find}』，可能是查找的内容写法不对，"
                    f"麻烦点「修改一下」告诉我文件里的实际写法。")
        self._agent_set_plan(detail, False)
        self.agent_find_label.config(text=talk)
        if variants:
            self._pack_interaction([self.agent_find_label, self.agent_fix_confirm_btn, self.agent_revise_btn])
        else:
            self._pack_interaction([self.agent_find_label, self.agent_revise_btn])

    def _on_find_fix_confirm(self):
        """用户确认按文件中的实际写法校正 → 更新 plan 并渲染方案，等待确认执行"""
        if self.agent_phase != 'find_check':
            return
        issues = getattr(self, 'agent_find_issues', None) or []
        partial = getattr(self, 'agent_find_partial', None)
        if not issues or partial is None:
            return
        issue = issues[0]
        variants = issue['variants']
        if not variants:
            return
        v = variants[0]
        ops = (self.agent_pending_plan or {}).get('params', {}).get('ops')
        if isinstance(ops, list) and 0 <= issue['op_index'] < len(ops):
            ops[issue['op_index']]['find'] = v['find']
        self.agent_log(f"已按文件中的实际写法『{v['find']}』校正查找内容（原『{issue['find']}』未在文件中出现）")
        self._agent_write_log(self._agent_task_id(), 'find_fix', {
            'original_find': issue['find'], 'corrected_find': v['find'], 'count': v['count']})
        self._hide_interaction()
        self.agent_phase = 'analyze'
        self._render_plan_view(partial)

    def _show_dynamic_plan(self, partial):
        """阶段1 动态代码方案展示：策略/说明/校验/代码预览 + 风险提示"""
        code = partial.get('generated_code', '')
        validation = partial.get('code_validation') or {}
        if not code or not validation.get('valid'):
            issues = validation.get('issues', [])
            self.agent_log(f"代码生成/校验未通过: {'; '.join(issues[:5]) if issues else '未生成代码'}")
            self._agent_set_plan(f"任务: custom（动态代码）\n状态: 未通过校验\n{'; '.join(issues[:5])}", False)
            return
        self.agent_pending_plan = {'task': 'dynamic_code', 'params': {}, 'files': list(self.agent_selected_files)}
        lines = [
            f"策略: 动态代码（内置工具无法满足）",
            f"代码说明: {partial.get('code_description', '')}",
            f"代码校验: 通过",
            f"风险等级: medium",
            '-' * 40,
            "代码预览（前800字符）:",
            code[:800],
            '-' * 40,
            "确认后将在一个隔离工作区中执行，仅允许读写分配目录。",
        ]
        self.agent_log("识别成功: 需求需要动态代码，等待确认")
        self._agent_set_plan('\n'.join(lines), True)

    def _agent_plan_summary(self, plan):
        """生成执行方案预览文本"""
        lines = [f"任务: {plan['task']}", f"参数: {json.dumps(plan['params'], ensure_ascii=False)}"]
        lines.append("输入文件:")
        for i, f in enumerate(plan['files']):
            lines.append(f"  [{i}] {os.path.basename(f)}")
        return '\n'.join(lines)

    def agent_confirm_execute(self):
        """Agent：阶段2 —— 恢复 LangGraph 执行 execute→validate→summary"""
        plan = self.agent_pending_plan
        if not plan:
            messagebox.showwarning("警告", "请先分析需求！")
            return
        if not self._ensure_agent_graph():
            return
        # 覆盖原文件（save_mode=overwrite）→ 先红色警醒确认（覆盖不可逆、提示备份）
        params = plan.get('params') or {}
        if str(params.get('save_mode') or '').strip().lower() == 'overwrite':
            fname = os.path.basename((plan.get('files') or [''])[0]) or '原文件'
            if not self._confirm_overwrite(fname):
                self.agent_log("用户取消覆盖原文件")
                return
        task_id = (self.agent_thread_config or {}).get('configurable', {}).get('thread_id', '')
        summary = self._agent_plan_summary(plan)
        if not messagebox.askyesno("确认执行方案", f"即将执行以下方案，确认继续？\n\n{summary}"):
            self.agent_log("用户取消执行")
            self._agent_write_log(task_id, 'cancelled', {'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            return
        self.agent_running = True
        self.agent_confirm_btn.config(state=tk.DISABLED)
        self.agent_cancel_btn.config(state=tk.DISABLED)
        self.agent_analyze_btn.config(state=tk.DISABLED)
        self.agent_log("开始执行（LangGraph 恢复执行）...")
        self._stop_confirm_pulse()
        self._agent_write_log(task_id, 'confirmed', {'plan': plan})

        def worker():
            error = None
            try:
                self.root.after(0, lambda: self._start_loading("正在执行处理任务，请稍候..."))
                final = self.agent_graph.invoke(None, config=self.agent_thread_config)
                for lg in final.get('logs') or []:
                    self.agent_log(lg)
                for o in final.get('output_files') or []:
                    self._remember_output_file(o)
                    self.agent_log(f"✓ 输出: {o}")
                self.agent_log(f"状态: {final.get('status', 'done')}")
                self.agent_log("✓ 全部执行完成")
                self._agent_write_log(task_id, 'done', {
                    'execution_result': final.get('execution_result'),
                    'output_files': final.get('output_files'),
                    'error': final.get('error'),
                    'status': final.get('status'),
                    'verification': final.get('verification'),
                    'repair_rounds': final.get('repair_rounds'),
                    'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
                # Loop④：展示结果摘要与自检结论，等待用户满意/反馈
                self._show_feedback(final)
            except Exception as e:
                self.agent_log(f"✗ 执行失败: {str(e)}")
                error = str(e)
                self._agent_write_log(task_id, 'failed', {
                    'error': error,
                    'end_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                })
            finally:
                self.root.after(0, self._stop_loading)
                self.agent_running = False
                self.agent_analyze_btn.config(state=tk.NORMAL)
                self.agent_confirm_btn.config(state=tk.DISABLED)
                self.agent_cancel_btn.config(state=tk.DISABLED)
                self.agent_pending_plan = None
                self.agent_thread_config = None

        threading.Thread(target=worker, daemon=True).start()

    def agent_cancel_plan(self):
        """Agent：取消待执行方案"""
        self.agent_pending_plan = None
        self._hide_interaction()
        self._agent_set_plan("已取消", False)
        self.agent_log("已取消执行方案")

    def call_intent_api(self, api_url, api_key, prompt):
        """调用 Dify 意图识别 chatflow（/chat-messages blocking），返回 answer 文本"""
        url = f"{api_url}/chat-messages"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "inputs": {},
            "query": prompt,
            "response_mode": "blocking",
            "conversation_id": "",
            "user": "excel-agent-intent"
        }
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        if response.status_code != 200:
            raise Exception(f"意图识别API返回错误 {response.status_code}: {response.text[:300]}")
        result = response.json()
        return str(result.get('answer', '') or '')

    def call_reflection_api(self, api_url, api_key, user_request, input_meta, output_meta, output_summary,
                            input_samples='', output_samples=''):
        """调用反思质检（复用 /chat-messages 通道，不新增 Dify 应用），返回 {satisfied, reason, suggestions}"""
        prompt = REFLECTION_PROMPT_TEMPLATE.format(
            request=user_request,
            input_meta=input_meta,
            output_meta=output_meta,
            output_summary=output_summary,
            input_samples=input_samples,
            output_samples=output_samples,
        )
        answer = self.call_intent_api(api_url, api_key, prompt)
        parsed = _parse_json_loose(answer)
        if not isinstance(parsed, dict):
            return {'satisfied': True, 'reason': f'反思返回无法解析（原文: {answer[:120]}）', 'suggestions': ''}
        return {
            'satisfied': bool(parsed.get('satisfied', True)),
            'reason': str(parsed.get('reason') or ''),
            'suggestions': str(parsed.get('suggestions') or ''),
        }

    def call_code_generator_api(self, api_url, api_key, user_request, file_metadata, sample_rows, constraints,
                                previous_code=''):
        """调用 Dify 代码生成 workflow（/workflows/run blocking），解析 data.outputs.result JSON"""
        url = f"{api_url}/workflows/run"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        inputs = {
            "user_request": user_request,
            "file_metadata": file_metadata,
            "sample_rows": sample_rows,
            "execution_constraints": constraints,
        }
        if previous_code:
            inputs['previous_code'] = previous_code  # 修复循环：携带上一轮代码供参考修改
        payload = {
            "inputs": inputs,
            "response_mode": "blocking",
            "user": "excel-agent-codegen"
        }
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        if response.status_code != 200:
            raise Exception(f"代码生成API返回错误 {response.status_code}: {response.text[:300]}")
        result = response.json()
        data = result.get('data') or {}
        status = data.get('status', '')
        outputs = data.get('outputs') or {}
        if status and str(status).lower() not in ('succeeded', 'success', 'stopped'):
            raise Exception(f"代码生成工作流未成功: status={status}, {json.dumps(result, ensure_ascii=False)[:300]}")
        normalized = _normalize_codegen_payload(outputs)
        if normalized is not None:
            return normalized
        result_str = outputs.get('result', '')
        if not result_str:
            raise Exception(
                f"代码生成工作流返回为空（status={status}），请检查 Dify 中 File Agent Code Generator 工作流："
                f"LLM节点模型是否可用、End节点输出变量是否为 result、应用是否已发布。原始返回: "
                f"{json.dumps(result, ensure_ascii=False)[:300]}")
        if isinstance(result_str, str):
            parsed = _parse_json_loose(result_str)
            normalized = _normalize_codegen_payload(parsed)
            if normalized is not None:
                return normalized
            fallback = {
                'code': _extract_json_string_field(result_str, 'code'),
                'description': _extract_json_string_field(result_str, 'description'),
            }
            normalized = _normalize_codegen_payload(fallback)
            if normalized is not None:
                return normalized
        return outputs

    def parse_intent_meta(self, text):
        """解析意图识别返回文本为完整 dict（task/params/understanding/clarify_questions 等）"""
        if not text:
            raise ValueError("意图识别返回为空")
        parsed = _parse_json_loose(text)
        truncated = False
        if parsed is None:
            # 截断场景：先尝试提取平衡子串（如 {...} 前半部分完整 JSON）
            sub = _extract_json_substring((text or '').strip())
            if sub:
                parsed = _parse_json_loose(sub)
                truncated = True
            else:
                # 无平衡子串（对象未闭合被截断）：用正则尽力提取 task 字段
                m = re.search(r'"task"\s*:\s*"([^"]+)"', text or '')
                if m:
                    parsed = {'task': m.group(1), 'params': {}}
                    truncated = True
        if parsed is None:
            preview = text.strip()[:200]
            if '{{#' in preview:
                raise ValueError(
                    f"意图识别应用的回答节点变量引用未解析（原文: {preview}）。"
                    f"请在 Dify 中打开该应用的「回答」节点，用变量选择器重新插入 LLM 节点输出，或重新导入修复后的 DSL 文件。")
            raise ValueError(f"意图识别返回不是有效JSON（原文: {preview}）")
        if not isinstance(parsed, dict):
            raise ValueError("意图识别返回必须是JSON对象")
        task = str(parsed.get('task') or parsed.get('intent') or '').strip()
        params = parsed.get('params')
        if params is None:
            params = {}
        elif not isinstance(params, dict):
            raise ValueError("params 必须为JSON对象")
        understanding = str(parsed.get('understanding') or '').strip()
        fs_raw = parsed.get('friendly_summary')
        friendly_summary = str(fs_raw).strip() if isinstance(fs_raw, str) else ''
        questions = parsed.get('clarify_questions')
        if not isinstance(questions, list):
            questions = []
        questions = [str(q).strip() for q in questions if str(q).strip()]
        # 截断检测：can_use_builtin_tool 非 bool（如截断成 'f'/'tru'）或原文未以 '}' 结尾
        cb_raw = parsed.get('can_use_builtin_tool')
        if cb_raw is not None and not isinstance(cb_raw, bool):
            truncated = True
            cb_raw = None
        raw_stripped = (text or '').strip()
        if not truncated and raw_stripped and not raw_stripped.endswith('}'):
            # 尾部非 '}'：可能是截断（半截 JSON），也可能是带说明文字（如 "...} 完毕"）。
            # 仅当尾部为 ASCII 字母/引号/冒号/逗号时视为截断特征
            tail = raw_stripped[-1]
            if tail in ('"', ':', ',') or (tail.isascii() and tail.isalpha()):
                truncated = True
        if cb_raw is None:
            can_builtin = (task != 'custom')
        else:
            can_builtin = bool(cb_raw)
        confidence = parsed.get('confidence')
        meta = {
            'task': task,
            'params': params,
            'understanding': understanding,
            'friendly_summary': friendly_summary,
            'clarify_questions': questions,
            'can_use_builtin_tool': can_builtin,
            'confidence': float(confidence) if isinstance(confidence, (int, float)) else None,
            'truncated': truncated,
        }
        if task == 'custom':
            meta['reason'] = str(params.get('reason') or parsed.get('reason') or '')
            return meta
        if task not in TOOL_REGISTRY:
            raise ValueError(f"未知/不支持的任务: {task}（支持: {','.join(sorted(TOOL_REGISTRY.keys()))}）")
        return meta

    def parse_intent_response(self, text):
        """解析意图识别返回文本为 (task, params)；宽松 JSON + 白名单校验，custom 任务单独处理（兼容旧调用）"""
        meta = self.parse_intent_meta(text)
        return meta['task'], meta['params']

    def execute_agent_task(self, task, params, files, out_dir=None):
        """Agent 任务分发（白名单，Phase 2 迁入 Tool Registry）"""
        if out_dir is None:
            out_dir = os.path.dirname(os.path.abspath(files[0])) if files else os.getcwd()
        if task == 'statistics':
            self._agent_task_statistics(files, out_dir)
        elif task == 'export_failed':
            self._agent_task_export_failed(files, params, out_dir)
        elif task == 'retry':
            self._agent_task_retry(files, params, out_dir)
        elif task == 'diff':
            self._agent_task_diff(files, params, out_dir)
        elif task == 'export_csv':
            self._agent_task_export_csv(files, params, out_dir)
        else:
            raise ValueError(f"未知任务: {task}")

    def _load_ws(self, path):
        if str(path).lower().endswith('.csv'):
            import csv as _csv
            wb = Workbook()
            ws = wb.active
            with open(path, 'r', encoding='utf-8-sig') as f:
                for row in _csv.reader(f):
                    ws.append(row)
            return wb, ws
        wb = load_workbook(path)
        return wb, wb.active

    def _agent_task_statistics(self, files, out_dir):
        """统计报表：逐文件生成文本报告并保存 txt"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        for f in files:
            wb, ws = self._load_ws(f)
            report = generate_statistics_report(ws)
            self.agent_log('\n' + report)
            base = os.path.splitext(os.path.basename(f))[0]
            txt_path = os.path.join(out_dir, f"{base}_agent_statistics_{timestamp}.txt")
            with open(txt_path, 'w', encoding='utf-8') as fh:
                fh.write(report)
            self._remember_output_file(txt_path)
            self.agent_log(f"✓ 统计报告已保存: {txt_path}")
            wb.close()

    def _agent_task_export_failed(self, files, params, out_dir):
        """导出失败项"""
        scope = str(params.get('scope', 'all') or 'all')
        fmt = str(params.get('format', 'excel') or 'excel')
        for f in files:
            wb, ws = self._load_ws(f)
            base = os.path.splitext(os.path.basename(f))[0]
            path, cnt = export_failed_rows(ws, out_dir, base, scope=scope, fmt=fmt)
            self._remember_output_file(path)
            self.agent_log(f"✓ 导出失败项 {cnt} 条: {path}")
            wb.close()

    def _agent_task_retry(self, files, params, out_dir):
        """失败重试：复用现有接口调用逻辑"""
        mode = str(params.get('mode', 'reverse') or 'reverse')
        for f in files:
            wb, ws = self._load_ws(f)
            base = os.path.splitext(os.path.basename(f))[0]
            path, ok_cnt, fail_cnt = self.retry_failed_rows(ws, mode, out_dir, base)
            self._remember_output_file(path)
            self.agent_log(f"✓ 重试完成: 成功 {ok_cnt} / 失败 {fail_cnt} -> {path}")
            wb.close()

    def _agent_task_diff(self, files, params, out_dir):
        """两文件 O 列渠道差异对比"""
        if len(files) < 2:
            raise ValueError("diff 任务需要至少选择两个文件！")
        src_idx = int(params.get('file_index', params.get('source_file_index', 0)) or 0)
        tgt_idx = int(params.get('other_file_index', params.get('target_file_index', 1)) or 1)
        if src_idx < 0 or tgt_idx < 0 or src_idx >= len(files) or tgt_idx >= len(files) or src_idx == tgt_idx:
            raise ValueError(f"非法文件索引: source={src_idx}, target={tgt_idx}（共 {len(files)} 个文件）")
        wb_a, ws_a = self._load_ws(files[src_idx])
        wb_b, ws_b = self._load_ws(files[tgt_idx])
        name_a = os.path.splitext(os.path.basename(files[src_idx]))[0]
        name_b = os.path.splitext(os.path.basename(files[tgt_idx]))[0]
        path, report = diff_worksheets(ws_a, ws_b, out_dir, name_a, name_b)
        self._remember_output_file(path)
        self.agent_log('\n' + report)
        self.agent_log(f"✓ 差异报告已保存: {path}")
        wb_a.close()
        wb_b.close()

    def _agent_task_export_csv(self, files, params, out_dir):
        """通用列导出 CSV"""
        columns = str(params.get('columns', '') or '')
        keyword = str(params.get('filter', '') or '')
        for f in files:
            wb, ws = self._load_ws(f)
            base = os.path.splitext(os.path.basename(f))[0]
            path, cnt = export_columns_csv(ws, out_dir, base, columns=columns, keyword=keyword)
            self._remember_output_file(path)
            self.agent_log(f"✓ 已导出 {cnt} 行: {path}")
            wb.close()

    def retry_failed_rows(self, ws, mode, out_dir, source_name):
        """对失败行重新调用接口处理，保存为新文件；返回 (输出路径, 成功数, 失败数)"""
        from openpyxl.styles import PatternFill
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(out_dir, f"{source_name}_agent_retry_{timestamp}.xlsx")

        if mode == 'forward':
            api_url = self.forward_api_url_entry.get().strip()
            api_key = self.forward_api_key_entry.get().strip()
            code = self.forward_code_entry.get().strip() or "CR039"
            channel = self.forward_channel_entry.get().strip()
            internet_flag = self.forward_internet_flag_entry.get().strip()
            system_source = self.forward_system_source_entry.get().strip()
        else:
            api_url = self.api_url_entry.get().strip()
            api_key = self.api_key_entry.get().strip()
            code = self.code_entry.get().strip()
            channel = self.channel_entry.get().strip()
            similarity_api_url = self.similarity_api_url_entry.get().strip()
            similarity_api_key = self.similarity_api_key_entry.get().strip()
        if not api_url or not api_key:
            raise ValueError(f"{'正向' if mode == 'forward' else '反向'} TAB 的 API 配置不完整，请先填写！")

        failed_rows = []
        for row in ws.iter_rows(min_row=2):
            p_val = _cell_str(row, COL_FAIL)
            s_val = _cell_str(row, COL_MATCH_TYPE)
            if not p_val.startswith('失败'):
                continue
            is_forward_fail = s_val.startswith('正向失败')
            if mode == 'forward' and not is_forward_fail:
                continue
            if mode != 'forward' and is_forward_fail:
                continue
            failed_rows.append(row)

        self.agent_log(f"待重试行数: {len(failed_rows)}（模式: {mode}）")
        ok_cnt = 0
        fail_cnt = 0
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row in failed_rows:
            row_idx = row[0].row
            if mode == 'forward':
                question = _cell_str(row, COL_QUESTION)
                if not question:
                    fail_cnt += 1
                    continue
                try:
                    api_result = self.call_forward_workflow_api(api_url, api_key, question, code, channel,
                                                                internet_flag, system_source)
                    systemfrom = api_result.get('systemfrom', '')
                    if systemfrom:
                        existing = _cell_str(row, COL_CHANNEL)
                        ws.cell(row=row_idx, column=COL_CHANNEL).value = self.merge_o_column(
                            existing, f"{systemfrom}(正向)")
                        ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = "正向工作流获取成功"
                        ws.cell(row=row_idx, column=COL_FAIL).value = None
                        ws.cell(row=row_idx, column=COL_FAIL).fill = green_fill
                        ok_cnt += 1
                    else:
                        ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = "正向失败: 工作流未返回systemfrom"
                        fail_cnt += 1
                except Exception as e:
                    ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = f"正向失败: {str(e)[:40]}"
                    fail_cnt += 1
            else:
                answer = _cell_str(row, COL_ANSWER)
                question = _cell_str(row, COL_QUESTION)
                if not answer:
                    fail_cnt += 1
                    continue
                try:
                    clean_answer = self.clean_answer_content(answer)
                    api_result = self.call_dify_api(api_url, api_key, clean_answer, code, channel)
                    channel_name = ''
                    match_type = '匹配失败'
                    if api_result and 'arr' in api_result:
                        arr_data = api_result['arr']
                        match_result = self.find_best_match(clean_answer, arr_data, original_question=question)
                        if match_result['matched']:
                            channel_name = match_result['channel']
                            match_type = '关键字匹配成功'
                        elif similarity_api_url and similarity_api_key:
                            sim = self.similarity_match(clean_answer, arr_data, similarity_api_url, similarity_api_key)
                            if sim and sim.get('matched'):
                                channel_name = sim.get('channel', '')
                                match_type = '相似度匹配成功'
                    if channel_name:
                        existing = _cell_str(row, COL_CHANNEL)
                        ws.cell(row=row_idx, column=COL_CHANNEL).value = self.merge_o_column(
                            existing, f"{channel_name}(反向)")
                        ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = match_type
                        ws.cell(row=row_idx, column=COL_FAIL).value = None
                        ws.cell(row=row_idx, column=COL_FAIL).fill = green_fill
                        ok_cnt += 1
                    else:
                        ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = match_type
                        fail_cnt += 1
                except Exception as e:
                    ws.cell(row=row_idx, column=COL_MATCH_TYPE).value = f"异常: {str(e)[:40]}"
                    fail_cnt += 1
            self.agent_log(f"  行 {row_idx}: {'成功' if ok_cnt + fail_cnt else ''}")

        wb = ws.parent
        wb.save(path)
        return path, ok_cnt, fail_cnt

    def call_forward_workflow_api(self, api_url, api_key, question, code, channel, internet_flag='', system_source=''):
        """调用知识问答完整问答流(/chat-messages)并解析 systemfrom（系统来源）

        返回: {'systemfrom': str, 'outputs': dict/list, 'raw': dict}
        """
        url = f"{api_url}/chat-messages"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        inputs = {}
        if code:
            inputs["code"] = code
        if channel:
            inputs["channel"] = channel
        if internet_flag:
            inputs["internetFlag"] = internet_flag
        if system_source:
            inputs["systemSource"] = system_source
        payload = {
            "inputs": inputs,
            "query": question,
            "response_mode": "streaming",
            "conversation_id": "",
            "user": "excel-processor-forward"
        }

        response = requests.post(url, headers=headers, json=payload, timeout=120, stream=True)
        if response.status_code != 200:
            raise Exception(f"问答流API返回错误 {response.status_code}: {response.text[:500]}")

        # streaming 模式下返回 SSE 事件流，逐行解析并构造为类 blocking 的 result
        answer_parts = []
        node_outputs = {}
        metadata = {}
        for line in response.iter_lines(decode_unicode=True):
            if not line:
                continue
            line = line.strip()
            if not line.startswith('data:'):
                continue
            data_str = line[5:].strip()
            if data_str == '[DONE]':
                break
            try:
                event = json.loads(data_str)
            except (json.JSONDecodeError, ValueError):
                continue
            event_name = event.get('event', '')
            if event_name == 'message':
                part = event.get('answer', '') or ''
                if part:
                    answer_parts.append(part)
            elif event_name in ('node_finished', 'workflow_finished'):
                ev_data = event.get('data') or {}
                outputs = ev_data.get('outputs')
                if isinstance(outputs, dict) and outputs:
                    # 后到的节点覆盖先到的（Func 节点在末尾，systemfrom 取最后值）
                    node_outputs.update(outputs)
            elif event_name == 'message_end':
                metadata = event.get('metadata') or {}
                break

        result = {
            'answer': ''.join(answer_parts),
            'outputs': node_outputs if node_outputs else None,
            'metadata': metadata,
        }

        outputs = result
        systemfrom = ''
        answer_text = ''

        # 解析优先级：顶层 -> data.outputs -> outputs -> retriever_resources -> answer JSON/文本 -> 深度递归
        if isinstance(result, dict):
            # 1. 顶层
            systemfrom = str(result.get('systemfrom', '') or '').strip()
            # 2. workflows/run 标准返回：data.outputs
            if not systemfrom and isinstance(result.get('data'), dict):
                data_outputs = result['data'].get('outputs')
                if isinstance(data_outputs, dict):
                    outputs = data_outputs
                    systemfrom = str(data_outputs.get('systemfrom', '') or '').strip()
            # 3. chatflow 节点输出
            if not systemfrom and isinstance(result.get('outputs'), dict):
                outputs = result['outputs']
                systemfrom = str(outputs.get('systemfrom', '') or '').strip()
            # 4. metadata.retriever_resources（知识库检索资源，来源字段）
            if not systemfrom and isinstance(result.get('metadata'), dict):
                resources = result['metadata'].get('retriever_resources')
                if isinstance(resources, list) and resources:
                    names = []
                    for res in resources:
                        name = str(res.get('dataset_name') or res.get('document_name') or '').strip()
                        if name and name not in names:
                            names.append(name)
                    if names:
                        systemfrom = ','.join(names)
            # 5. answer 字段（可能是 JSON 字符串，兼容代码围栏/前后空白）
            if not systemfrom:
                answer = result.get('answer', '')
                if isinstance(answer, str) and answer.strip():
                    stripped = answer.strip()
                    answer_text = answer
                    if stripped.startswith('```'):
                        stripped = stripped.strip('`').strip()
                        if stripped.lower().startswith('json'):
                            stripped = stripped[4:].strip()
                    try:
                        parsed = json.loads(stripped)
                        outputs = parsed
                        if isinstance(parsed, dict):
                            systemfrom = str(parsed.get('systemfrom', '') or '').strip()
                    except (json.JSONDecodeError, ValueError):
                        pass
                    # 5.5 文本正则提取（配合回答节点标记，如【系统来源】xxx）
                    if not systemfrom:
                        systemfrom = _extract_systemfrom_from_text(answer_text)
            # 6. 深度递归兜底
            if not systemfrom:
                found = _deep_find(result, 'systemfrom')
                if found is not None:
                    systemfrom = str(found).strip()

        raw_preview = ''
        if not systemfrom:
            # 排除超长 answer，展示其余结构（含 metadata），便于排查来源字段
            preview_obj = {k: v for k, v in result.items() if k != 'answer'}
            raw_preview = json.dumps(preview_obj, ensure_ascii=False)[:400]

        return {'systemfrom': systemfrom, 'outputs': outputs, 'raw': result,
                'raw_preview': raw_preview, 'answer_full': answer_text}

    def forward_start_processing(self):
        """正向模式：开始批量处理"""
        if not self.forward_selected_files:
            messagebox.showwarning("警告", "请先选择要处理的Excel文件！")
            return
        api_url = self.forward_api_url_entry.get().strip()
        api_key = self.forward_api_key_entry.get().strip()
        code = self.forward_code_entry.get().strip() or "CR039"
        channel = self.forward_channel_entry.get().strip()
        internet_flag = self.forward_internet_flag_entry.get().strip()
        system_source = self.forward_system_source_entry.get().strip()
        if not api_url or not api_key:
            messagebox.showwarning("警告", "请填写完整的工作流API配置信息！")
            return
        try:
            concurrent_count = int(self.forward_concurrent_var.get())
        except:
            concurrent_count = 2
        try:
            interval_str = self.forward_interval_var.get().replace('秒', '').strip()
            request_interval = float(interval_str)
        except:
            request_interval = 0.5

        self.forward_processing = True
        self.forward_process_btn.config(state=tk.DISABLED)
        self.forward_stop_btn.config(state=tk.NORMAL)
        self.forward_select_btn.config(state=tk.DISABLED)
        self.forward_clear_btn.config(state=tk.DISABLED)

        thread = threading.Thread(
            target=self.forward_process_files,
            args=(api_url, api_key, code, channel, internet_flag, system_source, concurrent_count, request_interval)
        )
        thread.daemon = True
        thread.start()

    def forward_stop_processing(self):
        """正向模式：停止处理"""
        self.forward_processing = False
        self.forward_log("正在停止正向处理...")

    def forward_process_files(self, api_url, api_key, code, channel, internet_flag, system_source, concurrent_count, request_interval=0.5):
        """正向模式：处理文件（在线程中执行）"""
        total_files = len(self.forward_selected_files)
        for file_idx, file_path in enumerate(self.forward_selected_files):
            if not self.forward_processing:
                break
            try:
                self.forward_log(f"\n{'='*60}")
                self.forward_log(f"开始正向处理文件 [{file_idx+1}/{total_files}]: {os.path.basename(file_path)}")
                self.forward_log(f"{'='*60}")
                self.forward_process_single_file(file_path, api_url, api_key, code, channel, internet_flag, system_source, concurrent_count, request_interval)
                self.forward_log(f"✓ 文件处理完成: {os.path.basename(file_path)}")
            except Exception as e:
                self.forward_log(f"✗ 处理文件失败: {os.path.basename(file_path)} - {str(e)}")
            progress = ((file_idx + 1) / total_files) * 100
            self.forward_progress_var.set(progress)
            self.forward_progress_label.config(text=f"进度: {progress:.1f}%")
            self.root.update()

        self.forward_processing = False
        self.forward_process_btn.config(state=tk.NORMAL)
        self.forward_stop_btn.config(state=tk.DISABLED)
        self.forward_select_btn.config(state=tk.NORMAL)
        self.forward_clear_btn.config(state=tk.NORMAL)
        self.forward_log("\n" + "="*60)
        self.forward_log("正向处理全部完成！")
        self.forward_log("="*60)
        messagebox.showinfo("完成", "正向处理全部完成！")

    def forward_process_single_file(self, file_path, api_url, api_key, code, channel, internet_flag, system_source, concurrent_count=4, request_interval=0.5):
        """正向模式：处理单个Excel文件（以H列问题调用工作流，获取systemfrom写入O列）"""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        try:
            if not self.check_file_writable(file_path):
                self.forward_log(f"⚠ 警告: 文件可能被占用（请关闭Excel后再试）")
                self.forward_log(f"  处理完成后将自动保存到新文件")

            self.forward_log("正在加载Excel文件...")
            wb = load_workbook(file_path)
            ws = wb.active
            max_row = ws.max_row
            self.forward_log(f"文件总行数: {max_row}")
            self.forward_log(f"使用参数 - Code: {code}, Channel: {channel if channel else '(空)'}, InternetFlag: {internet_flag if internet_flag else '(空)'}, SystemSource: {system_source if system_source else '(空)'}")
            self.forward_log(f"并发数: {concurrent_count}, 请求间隔: {request_interval}秒")

            processed_count = 0
            matched_count = 0
            error_count = 0

            # 收集需要处理的行：H列（第8列）问题非空
            rows_to_process = []
            for row_idx in range(2, max_row + 1):
                question = ws.cell(row=row_idx, column=8).value
                if question and str(question).strip():
                    rows_to_process.append({
                        'row_idx': row_idx,
                        'question': str(question).strip(),
                    })

            self.forward_log(f"需要处理的行数: {len(rows_to_process)}")

            log_lock = threading.Lock()

            def safe_log(msg):
                with log_lock:
                    self.forward_log(msg)

            def process_forward_row(row_info):
                if not self.forward_processing:
                    return None
                row_idx = row_info['row_idx']
                question = row_info['question']
                result = {
                    'row_idx': row_idx,
                    'success': False,
                    'systemfrom': None,
                    'outputs': None,
                    'error': None,
                }
                try:
                    if request_interval > 0:
                        import random
                        time.sleep(request_interval * random.uniform(0.8, 1.2))
                    api_result = self.call_forward_workflow_api(api_url, api_key, question, code, channel, internet_flag, system_source)
                    result['outputs'] = api_result.get('outputs')
                    result['raw_preview'] = api_result.get('raw_preview', '')
                    result['answer_full'] = api_result.get('answer_full', '')
                    systemfrom = api_result.get('systemfrom', '')
                    if systemfrom:
                        result['success'] = True
                        result['systemfrom'] = systemfrom
                    else:
                        result['error'] = "工作流未返回systemfrom"
                except Exception as e:
                    result['error'] = str(e)
                return result

            completed = 0
            total = len(rows_to_process)
            results = []

            with ThreadPoolExecutor(max_workers=concurrent_count) as executor:
                future_to_row = {
                    executor.submit(process_forward_row, row_info): row_info
                    for row_info in rows_to_process
                }
                for future in as_completed(future_to_row):
                    if not self.forward_processing:
                        break
                    result = future.result()
                    if result:
                        results.append(result)
                        completed += 1
                        progress = (completed / total) * 100 if total > 0 else 100
                        self.forward_progress_var.set(progress)
                        self.forward_progress_label.config(text=f"进度: {progress:.1f}%")
                        self.status_var.set(f"正向处理中: {completed}/{total} 行")
                        self.root.update_idletasks()

                        row_idx = result['row_idx']
                        if result['success']:
                            safe_log(f"[{completed}/{total}] 行 {row_idx} ✓ 正向获取成功 | 系统来源: {result['systemfrom']}")
                            matched_count += 1
                        else:
                            err_msg = result.get('error') or '未知原因'
                            safe_log(f"[{completed}/{total}] 行 {row_idx} ✗ 正向失败: {err_msg}")
                            raw_preview = result.get('raw_preview', '')
                            if raw_preview:
                                safe_log(f"    ↑ 问答流原始返回: {raw_preview}")
                            answer_full = result.get('answer_full', '')
                            if answer_full:
                                safe_log(f"    ↑ 完整answer文本: {answer_full[:2000]}")
                            error_count += 1
                    processed_count = completed

            # 写回结果到Excel
            self.forward_log("正在写入结果到Excel...")
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for result in results:
                if not result:
                    continue
                row_idx = result['row_idx']
                if result['success']:
                    # O列：合并写入 systemfrom(正向)，保留已有反向数据
                    existing_channel = ws.cell(row=row_idx, column=15).value
                    ws.cell(row=row_idx, column=15).value = self.merge_o_column(
                        existing_channel, f"{result['systemfrom']}(正向)")
                    ws.cell(row=row_idx, column=19).value = "正向工作流获取成功"
                    # 清除P列（第16列）历史失败标记，避免上一轮失败残留
                    ws.cell(row=row_idx, column=16).value = None
                    ws.cell(row=row_idx, column=16).fill = PatternFill()
                else:
                    ws.cell(row=row_idx, column=16).value = "失败(正向)"
                    ws.cell(row=row_idx, column=16).fill = red_fill
                    reason = result.get('error') or '未知原因'
                    ws.cell(row=row_idx, column=19).value = f"正向失败: {reason[:50]}"

            # 保存文件
            self.forward_log("正在保存文件...")
            saved_file_path = file_path
            try:
                wb.save(file_path)
                self._remember_output_file(file_path)
                self.forward_log(f"✓ 文件已保存: {file_path}")
            except PermissionError:
                base_name = os.path.splitext(file_path)[0]
                ext = os.path.splitext(file_path)[1]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_file_path = f"{base_name}_正向processed_{timestamp}{ext}"
                self.forward_log(f"⚠ 原文件被占用，保存到新文件: {new_file_path}")
                wb.save(new_file_path)
                saved_file_path = new_file_path
                self._remember_output_file(new_file_path)
                self.forward_log(f"✓ 文件已保存: {new_file_path}")
                messagebox.showwarning("保存提示", f"原文件被占用，已保存到新文件:\n{new_file_path}")
            except Exception as e:
                self.forward_log(f"✗ 保存文件失败: {str(e)}")
                raise

            # 导出正向失败记录
            output_dir = os.path.dirname(os.path.abspath(saved_file_path))
            source_filename = os.path.splitext(os.path.basename(saved_file_path))[0]
            self._save_forward_failed_records(ws, results, output_dir, source_filename)

            self.forward_log(f"\n正向处理统计:")
            self.forward_log(f"  总行数: {max_row}")
            self.forward_log(f"  需处理行数: {total}")
            self.forward_log(f"  已处理: {processed_count}")
            self.forward_log(f"  获取成功: {matched_count}")
            self.forward_log(f"  处理失败: {error_count}")

        except Exception as e:
            raise Exception(f"正向处理文件失败: {str(e)}")

    def _save_forward_failed_records(self, ws, results, output_dir, source_filename):
        """将正向处理失败记录导出为CSV文件，方便排查"""
        import csv

        failed_records = [r for r in results if r and not r.get('success', False)]
        if not failed_records:
            self.forward_log("没有正向失败记录，跳过导出")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"{source_filename}_正向失败记录_{timestamp}.csv"
        csv_path = os.path.join(output_dir, csv_filename)

        try:
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['行号', '消息ID', '问题', 'systemfrom', '原始返回', '错误'])
                for record in failed_records:
                    row_idx = record['row_idx']
                    message_id = str(ws.cell(row=row_idx, column=1).value or '')
                    question = str(ws.cell(row=row_idx, column=8).value or '')
                    systemfrom = str(record.get('systemfrom', '') or '')
                    outputs = record.get('outputs')
                    outputs_str = json.dumps(outputs, ensure_ascii=False) if outputs else ''
                    error = str(record.get('error', '') or '')
                    writer.writerow([row_idx, message_id, question, systemfrom, outputs_str[:3000], error])
                self._remember_output_file(csv_path)
            self.forward_log(f"✓ 正向失败记录已保存: {csv_path} ({len(failed_records)}条)")
        except Exception as e:
            self.forward_log(f"✗ 导出正向失败记录时出错: {str(e)}")

    def select_files(self):
        """选择文件"""
        files = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if files:
            for file in files:
                if file not in self.selected_files:
                    self.selected_files.append(file)
                    self.file_listbox.insert(tk.END, file)
            self.log(f"已选择 {len(files)} 个文件")
            
    def clear_files(self):
        """清空文件列表"""
        self.selected_files = []
        self.file_listbox.delete(0, tk.END)
        self.log("已清空文件列表")

    def clear_debug_log(self):
        """清空调试信息"""
        self.debug_text.delete(1.0, tk.END)
        self.current_question = ""
        self.current_api_response = ""
        self.log("已清空调试信息")

    def copy_question(self):
        """复制Question参数"""
        if self.current_question:
            self.root.clipboard_clear()
            self.root.clipboard_append(self.current_question)
            self.log("已复制Question参数到剪贴板")
        else:
            messagebox.showinfo("提示", "暂无Question参数可复制")

    def copy_api_response(self):
        """复制API返回数据"""
        if self.current_api_response:
            self.root.clipboard_clear()
            self.root.clipboard_append(self.current_api_response)
            self.log("已复制API返回数据到剪贴板")
        else:
            messagebox.showinfo("提示", "暂无API返回数据可复制")

    def debug_log(self, message):
        """添加调试日志"""
        if self.debug_var.get():
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.debug_text.insert(tk.END, f"[{timestamp}] {message}\n")
            # 不自动滚动，方便用户复制内容
            self.root.update()

    def test_api_connection(self):
        """测试API连接"""
        api_url = self.api_url_entry.get().strip()
        api_key = self.api_key_entry.get().strip()
        code = self.code_entry.get().strip() or "CR039"
        channel = self.channel_entry.get().strip()

        if not api_url or not api_key:
            messagebox.showwarning("警告", "请填写完整的API配置信息！")
            return

        self.log("开始测试API连接...")

        try:
            # 测试查询
            test_question = "测试查询"

            url = f"{api_url}/workflows/run"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "inputs": {
                    "question": test_question,
                    "code": code,
                },
                "response_mode": "blocking",
                "user": "test-user"
            }

            if channel:
                payload["inputs"]["channel"] = channel

            response = requests.post(url, headers=headers, json=payload, timeout=10)

            if response.status_code == 200:
                result = response.json()
                self.log("✓ API连接成功！")
                self.log(f"返回数据: {json.dumps(result, ensure_ascii=False, indent=2)[:500]}...")

                # 尝试解析返回数据结构
                if 'arr' in result:
                    self.log("检测到直接包含arr字段的数据结构")
                elif 'data' in result and 'outputs' in result.get('data', {}):
                    self.log("检测到标准Workflow数据结构")
                elif 'records' in result:
                    self.log("检测到知识库检索数据结构")

                messagebox.showinfo("成功", "API连接测试成功！\n请查看日志了解返回数据结构。")
            else:
                error_msg = f"API返回错误: {response.status_code}\n"
                try:
                    error_detail = response.json()
                    error_msg += f"错误详情: {json.dumps(error_detail, ensure_ascii=False)}"
                except:
                    error_msg += f"响应内容: {response.text}"
                self.log(f"✗ {error_msg}")
                messagebox.showerror("失败", error_msg)

        except Exception as e:
            error_msg = f"API连接失败: {str(e)}"
            self.log(f"✗ {error_msg}")
            messagebox.showerror("错误", error_msg)

    def test_similarity_api_connection(self):
        """测试相似度API连接"""
        api_url = self.similarity_api_url_entry.get().strip()
        api_key = self.similarity_api_key_entry.get().strip()

        if not api_url or not api_key:
            messagebox.showwarning("警告", "请填写完整的相似度API配置信息！")
            return

        self.log("开始测试相似度API连接...")

        try:
            # 测试查询
            test_original = "测试原始内容"
            test_deduplicated = "测试去重后内容"

            url = f"{api_url}/workflows/run"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }
            payload = {
                "inputs": {
                    "original": test_original,
                    "deduplicated": test_deduplicated
                },
                "response_mode": "blocking",
                "user": "test-user"
            }

            response = requests.post(url, headers=headers, json=payload, timeout=10)

            if response.status_code == 200:
                result = response.json()
                self.log("✓ 相似度API连接成功！")
                self.log(f"返回数据: {json.dumps(result, ensure_ascii=False)}")
                messagebox.showinfo("成功", "相似度API连接测试成功！\n请查看日志了解返回数据结构。")
            else:
                error_msg = f"相似度API返回错误: {response.status_code}\n"
                try:
                    error_detail = response.json()
                    error_msg += f"错误详情: {json.dumps(error_detail, ensure_ascii=False)}"
                except:
                    error_msg += f"响应内容: {response.text}"
                self.log(f"✗ {error_msg}")
                messagebox.showerror("失败", error_msg)

        except Exception as e:
            error_msg = f"相似度API连接失败: {str(e)}"
            self.log(f"✗ {error_msg}")
            messagebox.showerror("错误", error_msg)

    def log(self, message):
        """添加日志（界面 + 落盘 运行日志.log）"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        _runtime_log(message)
        
    def start_processing(self):
        """开始处理"""
        if not self.selected_files:
            messagebox.showwarning("警告", "请先选择要处理的Excel文件！")
            return

        api_url = self.api_url_entry.get().strip()
        api_key = self.api_key_entry.get().strip()
        code = self.code_entry.get().strip()
        channel = self.channel_entry.get().strip()

        # 相似度API配置
        similarity_api_url = self.similarity_api_url_entry.get().strip()
        similarity_api_key = self.similarity_api_key_entry.get().strip()
        
        # 获取并发数
        try:
            concurrent_count = int(self.concurrent_var.get())
        except:
            concurrent_count = 2

        # 获取请求间隔
        try:
            interval_str = self.interval_var.get().replace('秒', '').strip()
            request_interval = float(interval_str)
        except:
            request_interval = 0.5

        if not api_url or not api_key:
            messagebox.showwarning("警告", "请填写完整的API配置信息！")
            return

        # 验证code是否为空
        if not code:
            messagebox.showwarning("警告", "Code参数为必填项，请输入Code值！")
            return

        self.processing = True
        self.process_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.select_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)

        # 在新线程中处理
        thread = threading.Thread(target=self.process_files, args=(api_url, api_key, code, channel, similarity_api_url, similarity_api_key, concurrent_count, request_interval))
        thread.daemon = True
        thread.start()
        
    def stop_processing(self):
        """停止处理"""
        self.processing = False
        self.log("正在停止处理...")
        
    def process_files(self, api_url, api_key, code, channel, similarity_api_url, similarity_api_key, concurrent_count, request_interval=0.5):
        """处理文件（在线程中执行）"""
        total_files = len(self.selected_files)

        for file_idx, file_path in enumerate(self.selected_files):
            if not self.processing:
                break

            try:
                self.log(f"\n{'='*60}")
                self.log(f"开始处理文件 [{file_idx+1}/{total_files}]: {os.path.basename(file_path)}")
                self.log(f"{'='*60}")

                # 处理单个文件
                self.process_single_file(file_path, api_url, api_key, code, channel, similarity_api_url, similarity_api_key, concurrent_count, request_interval)

                self.log(f"✓ 文件处理完成: {os.path.basename(file_path)}")

            except Exception as e:
                self.log(f"✗ 处理文件失败: {os.path.basename(file_path)} - {str(e)}")

            # 更新总体进度
            progress = ((file_idx + 1) / total_files) * 100
            self.progress_var.set(progress)
            self.progress_label.config(text=f"进度: {progress:.1f}%")
            self.root.update()

        self.processing = False
        self.process_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.select_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)

        self.log("\n" + "="*60)
        self.log("所有文件处理完成！")
        self.log("="*60)
        self.status_var.set("处理完成")
        messagebox.showinfo("完成", "所有文件处理完成！")

    def check_file_writable(self, file_path):
        """检查文件是否可写入（未被占用）"""
        try:
            # 尝试以追加模式打开文件
            with open(file_path, 'a'):
                pass
            return True
        except PermissionError:
            return False
        except Exception:
            return True  # 其他错误不阻止处理

    def process_single_file(self, file_path, api_url, api_key, code, channel, similarity_api_url, similarity_api_key, concurrent_count=4, request_interval=0.5):
        """处理单个Excel文件（支持并发处理）"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        
        try:
            # 检查文件是否被占用
            if not self.check_file_writable(file_path):
                self.log(f"⚠ 警告: 文件可能被占用（请关闭Excel后再试）")
                self.log(f"  处理完成后将自动保存到新文件")

            # 加载工作簿
            self.log("正在加载Excel文件...")
            wb = load_workbook(file_path)
            ws = wb.active

            # 获取总行数
            max_row = ws.max_row
            self.log(f"文件总行数: {max_row}")
            self.log(f"使用参数 - Code: {code}, Channel: {channel if channel else '(空)'}")
            self.log(f"并发数: {concurrent_count}, 请求间隔: {request_interval}秒")
            if similarity_api_url:
                self.log(f"相似度API已配置: {similarity_api_url}")

            # 统计数据
            processed_count = 0
            matched_count = 0
            error_count = 0
            similarity_matched_count = 0

            # 收集需要处理的行
            rows_to_process = []
            for row_idx in range(2, max_row + 1):
                hit_type = ws.cell(row=row_idx, column=12).value
                answer = ws.cell(row=row_idx, column=9).value
                message_id = ws.cell(row=row_idx, column=1).value
                question = ws.cell(row=row_idx, column=8).value  # H列是问题
                
                if hit_type in [4, 9, '4', '9'] and answer:
                    rows_to_process.append({
                        'row_idx': row_idx,
                        'hit_type': hit_type,
                        'answer': str(answer),
                        'message_id': message_id,
                        'question': str(question) if question else ''  # 新增：原始问题
                    })

            self.log(f"需要处理的行数: {len(rows_to_process)}")
            
            # 创建线程锁用于日志输出
            log_lock = threading.Lock()
            
            def safe_log(msg):
                with log_lock:
                    self.log(msg)

            # 处理单行的函数
            def process_single_row(row_info):
                if not self.processing:
                    return None
                    
                row_idx = row_info['row_idx']
                answer = row_info['answer']
                message_id = row_info['message_id']
                hit_type = row_info['hit_type']
                original_question = row_info.get('question', '')  # 获取原始问题
                
                result = {
                    'row_idx': row_idx,
                    'success': False,
                    'channel': None,
                    'match_type': None,
                    'clean_answer': None,
                    'arr_data': None,
                    'error': None,
                    'keyword_match_info': None,  # 关键字匹配统计信息
                    'similarity_request': None,  # 相似度接口提交参数
                    'similarity_response': None   # 相似度接口返回结果
                }
                
                try:
                    # 请求间隔限流（随机抖动避免线程同时请求）
                    if request_interval > 0:
                        import random
                        jitter = random.uniform(0.8, 1.2)
                        time.sleep(request_interval * jitter)

                    # 清理答案内容
                    clean_answer = self.clean_answer_content(answer)
                    result['clean_answer'] = clean_answer
                    
                    # 调用Dify API
                    api_result = self.call_dify_api(api_url, api_key, clean_answer, code, channel)
                    
                    if api_result and 'arr' in api_result:
                        arr_data = api_result['arr']
                        result['arr_data'] = arr_data
                        
                        # 尝试关键字匹配（传入原始问题用于问题原文匹配加分）
                        match_result = self.find_best_match(clean_answer, arr_data, original_question=question)
                        
                        # 记录关键字匹配统计信息
                        result['keyword_match_info'] = (
                            f"匹配数={match_result['best_keyword_count']}/{match_result['total_keywords']}, "
                            f"阈值≥{match_result['min_keyword_count']}"
                        )
                        
                        if match_result['matched']:
                            result['success'] = True
                            result['channel'] = match_result['channel']
                            result['match_type'] = "关键字匹配成功"
                        else:
                            # 关键字匹配失败，尝试相似度API
                            if similarity_api_url and similarity_api_key:
                                similarity_result = self.similarity_match(
                                    clean_answer, arr_data, similarity_api_url, similarity_api_key
                                )
                                
                                # 保存提交参数
                                deduplicated_str = json.dumps(arr_data, ensure_ascii=False)
                                result['similarity_request'] = f"original: {clean_answer}\ndeduplicated: {deduplicated_str[:2000]}..."
                                
                                if similarity_result:
                                    # 【重要】保存原始API响应，而不是解析后的结果
                                    if similarity_result.get('raw_response'):
                                        # 从原始响应中提取 outputs 部分
                                        raw_resp = similarity_result['raw_response']
                                        if 'data' in raw_resp and 'outputs' in raw_resp.get('data', {}):
                                            result['similarity_response'] = json.dumps(raw_resp['data']['outputs'], ensure_ascii=False)
                                        else:
                                            result['similarity_response'] = json.dumps(raw_resp, ensure_ascii=False)
                                    elif similarity_result.get('error'):
                                        result['similarity_response'] = f"错误: {similarity_result.get('error')}"
                                    else:
                                        result['similarity_response'] = json.dumps(similarity_result, ensure_ascii=False)
                                    
                                    if similarity_result.get('matched'):
                                        result['success'] = True
                                        result['channel'] = similarity_result.get('channel', '')
                                        result['match_type'] = "相似度匹配成功"
                                    else:
                                        result['match_type'] = "匹配失败"
                                else:
                                    result['match_type'] = "匹配失败"
                                    result['similarity_response'] = "接口返回为空或请求失败"
                            else:
                                result['match_type'] = "匹配失败"
                                result['similarity_request'] = "未配置相似度API"
                    else:
                        result['match_type'] = "API错误"
                        
                except Exception as e:
                    result['error'] = str(e)
                    result['match_type'] = f"异常: {str(e)[:30]}"
                
                return result

            # 使用线程池并发处理
            completed = 0
            total = len(rows_to_process)
            
            with ThreadPoolExecutor(max_workers=concurrent_count) as executor:
                # 提交所有任务
                future_to_row = {
                    executor.submit(process_single_row, row_info): row_info 
                    for row_info in rows_to_process
                }
                
                # 收集结果
                results = []
                for future in as_completed(future_to_row):
                    if not self.processing:
                        break
                        
                    result = future.result()
                    if result:
                        results.append(result)
                        completed += 1
                        
                        # 更新进度
                        progress = (completed / total) * 100 if total > 0 else 100
                        self.progress_var.set(progress)
                        self.progress_label.config(text=f"进度: {progress:.1f}%")
                        self.status_var.set(f"处理中: {completed}/{total} 行")
                        self.root.update_idletasks()
                        
                        # 输出日志
                        row_idx = result['row_idx']
                        if result['success']:
                            safe_log(f"[{completed}/{total}] 行 {row_idx} ✓ {result['match_type']} | 渠道: {result['channel']}")
                            matched_count += 1
                        else:
                            safe_log(f"[{completed}/{total}] 行 {row_idx} ✗ {result['match_type']}")
                            error_count += 1
                            if result['match_type'] == "相似度匹配成功":
                                similarity_matched_count += 1
                    
                    processed_count = completed
            
            # 写入结果到Excel
            self.log("正在写入结果到Excel...")
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for result in results:
                if result:
                    row_idx = result['row_idx']
                    
                    # Q列（第17列）：对比的原始问题
                    if result['clean_answer']:
                        ws.cell(row=row_idx, column=17).value = result['clean_answer']
                    
                    # R列（第18列）：arr数组数据
                    if result['arr_data']:
                        ws.cell(row=row_idx, column=18).value = json.dumps(result['arr_data'], ensure_ascii=False, indent=2)
                    
                    if result['success']:
                        # O列（第15列）：渠道，带(反向)标识，并保留已有正向数据
                        existing_channel = ws.cell(row=row_idx, column=15).value
                        ws.cell(row=row_idx, column=15).value = self.merge_o_column(existing_channel, f"{result['channel']}(反向)")
                        # S列（第19列）：匹配方式
                        ws.cell(row=row_idx, column=19).value = result['match_type']
                        # 清除P列（第16列）历史失败标记，避免上一轮失败残留
                        ws.cell(row=row_idx, column=16).value = None
                        ws.cell(row=row_idx, column=16).fill = PatternFill()
                    else:
                        # P列（第16列）：失败标记（带反向标识）
                        ws.cell(row=row_idx, column=16).value = "失败(反向)"
                        ws.cell(row=row_idx, column=16).fill = red_fill
                        # S列（第19列）：匹配方式
                        ws.cell(row=row_idx, column=19).value = result['match_type']
                        
                        # T列（第20列）：相似度接口提交参数和返回结果
                        if result.get('similarity_request') or result.get('similarity_response'):
                            t_value = ""
                            if result.get('similarity_request'):
                                t_value += f"提交参数是：\n{result['similarity_request']}\n\n"
                            if result.get('similarity_response'):
                                t_value += f"返回结果是：\n{result['similarity_response']}"
                            ws.cell(row=row_idx, column=20).value = t_value

            # 保存文件
            self.log("正在保存文件...")
            saved_file_path = file_path
            try:
                wb.save(file_path)
                self._remember_output_file(file_path)
                self.log(f"✓ 文件已保存: {file_path}")
            except PermissionError:
                base_name = os.path.splitext(file_path)[0]
                ext = os.path.splitext(file_path)[1]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_file_path = f"{base_name}_processed_{timestamp}{ext}"
                self.log(f"⚠ 原文件被占用，保存到新文件: {new_file_path}")
                wb.save(new_file_path)
                saved_file_path = new_file_path
                self._remember_output_file(new_file_path)
                self.log(f"✓ 文件已保存: {new_file_path}")
                messagebox.showwarning("保存提示", f"原文件被占用，已保存到新文件:\n{new_file_path}")
            except Exception as e:
                self.log(f"✗ 保存文件失败: {str(e)}")
                raise

            # 导出失败记录到CSV
            output_dir = os.path.dirname(os.path.abspath(saved_file_path))
            source_filename = os.path.splitext(os.path.basename(saved_file_path))[0]
            self._save_failed_records(ws, results, output_dir, source_filename)

            self.log(f"\n处理统计:")
            self.log(f"  总行数: {max_row}")
            self.log(f"  需处理行数: {total}")
            self.log(f"  已处理: {processed_count}")
            self.log(f"  匹配成功: {matched_count}")
            self.log(f"  处理失败: {error_count}")
            
        except Exception as e:
            raise Exception(f"处理文件失败: {str(e)}")
            
    def clean_answer_content(self, answer):
        """清理答案内容，移除不需要的文案"""
        import re

        # 需要移除的文案模式
        patterns_to_remove = [
            r'【答案来源说明】本次回答基于本地知识库整理生成。',
            r'【答案来源说明】本次回答基于知识库中相似问题整理而成，仅供参考。',
            r'【答案来源说明】.*?。',
        ]

        clean_answer = answer
        for pattern in patterns_to_remove:
            clean_answer = re.sub(pattern, '', clean_answer)

        # 清理多余的空白字符和换行符
        clean_answer = ' '.join(clean_answer.split())

        return clean_answer.strip()

    def _request_with_retry(self, url, headers, payload, timeout, max_retries=2, api_name="API"):
        """带重试机制的API请求，仅对服务端错误(500/502/503/504)和网络异常重试"""
        retry_delays = [2, 4]  # 每次重试的等待秒数
        last_response = None

        for attempt in range(max_retries + 1):
            # 用户停止时立即退出重试
            if not self.processing:
                return last_response

            try:
                response = requests.post(url, headers=headers, json=payload, timeout=timeout)

                if response.status_code == 200:
                    return response

                last_response = response

                # 服务端错误才重试
                if response.status_code in [500, 502, 503, 504] and attempt < max_retries:
                    delay = retry_delays[attempt]
                    self.log(f"    ⚠ {api_name}返回{response.status_code}，{delay}秒后重试({attempt+1}/{max_retries})...")
                    time.sleep(delay)
                    continue

                # 非服务端错误或重试次数用尽，直接返回
                return response

            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                if attempt < max_retries and self.processing:
                    delay = retry_delays[attempt]
                    self.log(f"    ⚠ {api_name}请求异常({str(e)[:80]})，{delay}秒后重试({attempt+1}/{max_retries})...")
                    time.sleep(delay)
                    continue
                raise

        return last_response

    def call_dify_api(self, api_url, api_key, question, code, channel):
        """调用Dify API"""
        try:
            # 保存当前的question参数
            self.current_question = question

            # 调试模式：记录question
            if self.debug_var.get():
                self.debug_log("\n" + "="*80)
                self.debug_log("【提交API的Question参数】")
                self.debug_log("-" * 80)
                self.debug_log(question)
                self.debug_log("-" * 80)
                self.debug_log(f"参数: code={code}, channel={channel if channel else '(空)'}")
                self.debug_log("="*80)

            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }

            # 构建请求体（根据工作流参数）
            url = f"{api_url}/workflows/run"
            payload = {
                "inputs": {
                    "question": question,
                    "code": code,
                },
                "response_mode": "blocking",
                "user": "excel-processor"
            }

            # 如果channel不为空，添加到inputs中
            if channel:
                payload["inputs"]["channel"] = channel

            # 发送请求（带重试）
            self.log(f"    请求URL: {url}")
            self.log(f"    参数: code={code}, channel={channel if channel else '(空)'}")
            response = self._request_with_retry(url, headers, payload, timeout=30, api_name="知识库API")

            # 打印响应状态码
            self.log(f"    响应状态码: {response.status_code}")

            if response.status_code != 200:
                error_text = response.text[:500]
                raise Exception(f"API返回错误 {response.status_code}: {error_text}")

            # 解析响应
            result = response.json()

            # 调试模式：记录完整返回数据
            if self.debug_var.get():
                api_response_str = json.dumps(result, ensure_ascii=False, indent=2)
                self.debug_log("\n" + "="*80)
                self.debug_log("API完整返回数据:")
                self.debug_log(api_response_str[:2000])
                self.debug_log("="*80 + "\n")

            # 打印原始返回数据（前500字符）
            self.log(f"    原始返回数据: {json.dumps(result, ensure_ascii=False)[:300]}...")

            # 尝试多种数据格式解析
            # 格式1: 直接包含arr字段（您提供的格式）
            if 'arr' in result:
                self.log(f"    检测到直接arr字段格式")
                return result

            # 格式2: 标准workflow返回格式
            if 'data' in result and 'outputs' in result.get('data', {}):
                self.log(f"    检测到标准workflow格式")
                outputs = result['data']['outputs']

                # 优先检查arr字段（包含完整对象的数组，有content/channel/score）
                if 'arr' in outputs:
                    self.log(f"    检测到outputs.arr字段，包含{len(outputs['arr'])}个元素")
                    if self.debug_var.get() and outputs['arr']:
                        self.debug_log(f"arr字段第一个元素: {json.dumps(outputs['arr'][0], ensure_ascii=False)[:200]}...")
                    return {'arr': outputs['arr']}

                if 'result' in outputs:
                    result_str = outputs['result']
                    try:
                        if isinstance(result_str, str):
                            # 尝试解析JSON字符串
                            parsed_result = json.loads(result_str)

                            # 如果解析出的是列表，构建arr结构
                            if isinstance(parsed_result, list):
                                # 调试：检查列表中的元素类型
                                if parsed_result and len(parsed_result) > 0:
                                    first_item_type = type(parsed_result[0]).__name__
                                    self.log(f"    解析到列表，共{len(parsed_result)}个元素，第一个元素类型: {first_item_type}")

                                    if self.debug_var.get():
                                        self.debug_log(f"解析后的列表前3个元素:")
                                        for i in range(min(3, len(parsed_result))):
                                            item = parsed_result[i]
                                            if isinstance(item, dict):
                                                self.debug_log(f"  元素[{i}]: 字典 - keys: {list(item.keys())}")
                                            else:
                                                self.debug_log(f"  元素[{i}]: {type(item).__name__} - {str(item)[:100]}...")

                                return {
                                    'result': result_str,
                                    'arr': parsed_result
                                }
                        elif isinstance(result_str, dict):
                            return result_str
                    except json.JSONDecodeError as e:
                        self.log(f"    警告: 无法解析result字段: {str(e)}")
                        return None

            # 格式3: knowledge检索返回格式
            if 'records' in result:
                self.log(f"    检测到知识库检索格式")
                records = result['records']
                # 转换为arr格式
                arr = []
                for record in records:
                    if isinstance(record, dict) and 'content' in record:
                        arr.append(record)
                return {'arr': arr}

            # 格式4: 检查是否有其他可能包含数组数据的字段
            for key in ['results', 'data', 'items']:
                if key in result and isinstance(result[key], list):
                    self.log(f"    检测到{key}字段格式")
                    return {'arr': result[key]}

            self.log(f"    警告: 无法识别的返回数据格式")
            self.log(f"    完整返回: {json.dumps(result, ensure_ascii=False)[:1000]}")
            return None

        except requests.exceptions.Timeout:
            raise Exception("API请求超时")
        except requests.exceptions.RequestException as e:
            raise Exception(f"API请求失败: {str(e)}")
        except json.JSONDecodeError as e:
            raise Exception(f"API返回数据格式错误: {str(e)}")

    def similarity_match(self, original_text, arr_data, similarity_api_url, similarity_api_key):
        """使用相似度API进行二次匹配（一次性传入整个arr数组）"""
        if not arr_data or not isinstance(arr_data, list):
            self.log(f"    arr_data为空或不是列表")
            return None

        # 检查相似度API配置
        if not similarity_api_url or not similarity_api_key:
            self.log(f"    ✗ 相似度API未配置（URL或Key为空）")
            return None

        self.log(f"    相似度API地址: {similarity_api_url}")
        self.log(f"    开始相似度匹配，共{len(arr_data)}个候选项")

        # 调试模式：记录参数
        if self.debug_var.get():
            self.debug_log(f"\n--- 相似度匹配 ---")
            self.debug_log(f"original (前100字符): {original_text[:100]}...")
            self.debug_log(f"deduplicated (arr数组): {len(arr_data)}个元素")
            for i, item in enumerate(arr_data[:3]):
                if isinstance(item, dict):
                    self.debug_log(f"  元素[{i}]: channel={item.get('channel', 'N/A')}, content前50字符={str(item.get('content', ''))[:50]}...")

        try:
            # 一次性调用相似度API，传入整个arr数组
            result = self.call_similarity_api(
                similarity_api_url,
                similarity_api_key,
                original_text,
                arr_data
            )

            # 解析返回结果
            if result:
                matched = result.get('matched', False)
                channel = result.get('channel', '')
                matched_item = result.get('matched_item', {})

                # 处理 matched 可能是字符串的情况
                if isinstance(matched, str):
                    matched = matched.lower() == 'true'

                self.log(f"    【相似度判断最终结果】matched={matched}, channel='{channel}'")

                if matched:
                    self.log(f"    ✓ 相似度匹配成功 | 渠道: {channel}")
                    if self.debug_var.get():
                        self.debug_log(f"\n{'='*60}")
                        self.debug_log(f"【相似度匹配最终结果】成功")
                        self.debug_log(f"  channel: {channel}")
                        self.debug_log(f"  matched_item: {json.dumps(matched_item, ensure_ascii=False)[:500]}")
                        self.debug_log(f"{'='*60}\n")
                    return {'matched': True, 'channel': channel, 'matched_item': matched_item, 'raw_result': result}
                else:
                    self.log(f"    ✗ 相似度匹配未找到匹配项（matched=false）")
                    return {'matched': False, 'channel': '', 'matched_item': {}, 'raw_result': result}
            else:
                self.log(f"    ✗ 相似度API返回数据为空（result=None）")
                return None

        except Exception as e:
            self.log(f"    ✗ 相似度判断失败: {str(e)}")
            if self.debug_var.get():
                self.debug_log(f"\n【相似度判断异常】")
                self.debug_log(f"  错误: {str(e)}")
                import traceback
                self.debug_log(f"  堆栈跟踪:\n{traceback.format_exc()}")
            return {'matched': False, 'error': str(e)}

    def call_similarity_api(self, api_url, api_key, original, deduplicated):
        """调用相似度判断API（一次性传入整个arr数组，转为JSON字符串）"""
        try:
            url = f"{api_url}/workflows/run"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            }

            # 将arr数组转换为JSON字符串
            deduplicated_str = json.dumps(deduplicated, ensure_ascii=False) if isinstance(deduplicated, list) else deduplicated

            payload = {
                "inputs": {
                    "original": original,
                    "deduplicated": deduplicated_str  # 传入JSON字符串
                },
                "response_mode": "blocking",
                "user": "excel-processor"
            }

            # 详细日志：请求参数
            self.log(f"    【相似度API请求】")
            self.log(f"      URL: {url}")
            self.log(f"      original长度: {len(original)} 字符")
            self.log(f"      deduplicated数组长度: {len(deduplicated) if isinstance(deduplicated, list) else 'N/A'} 个元素")
            if self.debug_var.get():
                self.debug_log(f"\n{'='*60}")
                self.debug_log(f"【相似度API请求参数】")
                self.debug_log(f"original (前200字符): {original[:200]}...")
                self.debug_log(f"deduplicated (前3个元素):")
                if isinstance(deduplicated, list):
                    for i, item in enumerate(deduplicated[:3]):
                        if isinstance(item, dict):
                            self.debug_log(f"  [{i}] channel={item.get('channel', 'N/A')}, content前100字符={str(item.get('content', ''))[:100]}...")
                self.debug_log(f"{'='*60}")

            response = self._request_with_retry(url, headers, payload, timeout=120, api_name="相似度API")

            self.log(f"      响应状态码: {response.status_code}")

            if response.status_code != 200:
                error_text = response.text[:500]
                self.log(f"      错误响应: {error_text}")
                raise Exception(f"相似度API返回错误 {response.status_code}: {error_text}")

            result = response.json()

            # 【重要】先打印完整的原始API响应，用于调试
            self.log(f"      【原始API响应（解析前）】")
            self.log(f"      {json.dumps(result, ensure_ascii=False)[:3000]}")

            if self.debug_var.get():
                self.debug_log(f"\n{'='*80}")
                self.debug_log(f"【相似度API原始响应（完整）】")
                self.debug_log(json.dumps(result, ensure_ascii=False))
                self.debug_log(f"{'='*80}")

            # 解析返回数据 - Dify Workflow 格式
            outputs = None
            if 'data' in result and isinstance(result['data'], dict) and 'outputs' in result['data']:
                outputs = result['data']['outputs']
                self.log(f"      【解析路径】从 result['data']['outputs'] 获取")
            elif 'outputs' in result:
                outputs = result['outputs']
                self.log(f"      【解析路径】从 result['outputs'] 获取")
            else:
                # 直接返回结果（可能是简化格式）
                outputs = result
                self.log(f"      【解析路径】直接使用 result 作为 outputs")

            if outputs:
                self.log(f"      【outputs内容】{json.dumps(outputs, ensure_ascii=False)[:1000]}")

                # 【关键修复】检查 outputs 中是否有 text 字段（Dify 返回格式）
                # text 字段是一个 JSON 字符串，包含真正的 matched、channel 等信息
                if 'text' in outputs and isinstance(outputs['text'], str):
                    self.log(f"      【检测到text字段】需要解析text字符串")
                    try:
                        text_content = outputs['text']
                        parsed_text = json.loads(text_content)
                        self.log(f"      【text解析结果】matched={parsed_text.get('matched')}, channel={parsed_text.get('channel')}")
                        # 使用解析后的内容
                        outputs = parsed_text
                    except json.JSONDecodeError as e:
                        self.log(f"      【text解析失败】{str(e)}")

                matched = outputs.get('matched', False)
                channel = outputs.get('channel', '')
                matched_item = outputs.get('matched_item', {})

                # 处理 matched 可能是字符串的情况
                if isinstance(matched, str):
                    matched = matched.lower() == 'true'

                # 详细日志：最终结果
                self.log(f"      【解析结果】matched={matched}, channel={channel}")
                if self.debug_var.get():
                    self.debug_log(f"【相似度判断最终结果】")
                    self.debug_log(f"  matched: {matched} (类型: {type(matched).__name__})")
                    self.debug_log(f"  channel: {channel}")
                    self.debug_log(f"  matched_item: {json.dumps(matched_item, ensure_ascii=False)[:500]}")

                return {
                    'matched': matched,
                    'channel': channel,
                    'matched_item': matched_item if isinstance(matched_item, dict) else {},
                    'raw_response': result  # 保存原始响应
                }
            else:
                raise Exception(f"无法解析API返回格式: {json.dumps(result, ensure_ascii=False)[:300]}")

        except requests.exceptions.Timeout:
            self.log(f"      ✗ 请求超时（120秒）")
            raise Exception("相似度API请求超时（已等待120秒）")
        except requests.exceptions.RequestException as e:
            self.log(f"      ✗ 请求异常: {str(e)}")
            raise Exception(f"相似度API请求失败: {str(e)}")
        except json.JSONDecodeError as e:
            self.log(f"      ✗ JSON解析错误: {str(e)}")
            raise Exception(f"相似度API返回数据格式错误: {str(e)}")

    def _save_failed_records(self, ws, results, output_dir, source_filename):
        """将失败记录导出为CSV文件，方便排查"""
        import csv
        from datetime import datetime
        
        failed_records = [r for r in results if r and not r.get('success', False)]
        if not failed_records:
            self.log("没有失败记录，跳过导出")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_filename = f"{source_filename}_失败记录_{timestamp}.csv"
        csv_path = os.path.join(output_dir, csv_filename)
        
        try:
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([
                    '行号', '消息ID', '问题', '原始答案(前2000字)', '清理后答案(前2000字)',
                    '命中类型', '匹配方式/失败原因', 'arr候选项数量',
                    'arr候选项详情', '关键字匹配信息',
                    '相似度API请求(前5000字)', '相似度API响应(前5000字)', '异常信息'
                ])
                
                for record in failed_records:
                    row_idx = record['row_idx']
                    message_id = str(ws.cell(row=row_idx, column=1).value or '')
                    question = str(ws.cell(row=row_idx, column=8).value or '')[:2000]
                    raw_answer = str(ws.cell(row=row_idx, column=9).value or '')[:2000]
                    clean_answer = str(record.get('clean_answer', '') or '')[:2000]
                    hit_type = str(record.get('hit_type', '') or '')
                    match_type = str(record.get('match_type', '') or '')
                    arr_data = record.get('arr_data')
                    arr_count = len(arr_data) if arr_data else 0
                    arr_detail = json.dumps(arr_data, ensure_ascii=False) if arr_data else ''
                    keyword_info = str(record.get('keyword_match_info', '') or '')
                    sim_request = str(record.get('similarity_request', '') or '')[:5000]
                    sim_response = str(record.get('similarity_response', '') or '')[:5000]
                    error = str(record.get('error', '') or '')
                    
                    writer.writerow([
                        row_idx, message_id, question, raw_answer, clean_answer,
                        hit_type, match_type, arr_count, arr_detail, keyword_info,
                        sim_request, sim_response, error
                    ])
                self._remember_output_file(csv_path)
            
            self.log(f"✓ 失败记录已保存: {csv_path} ({len(failed_records)}条)")
        except Exception as e:
            self.log(f"✗ 导出失败记录时出错: {str(e)}")

    def merge_o_column(self, existing, new_entry):
        """合并O列渠道值：反向条目在前、正向条目在后，逗号分隔，自动去重。

        参数:
            existing:  O列现有值（可能为空，或含 (反向)/(正向) 标识的历史数据）
            new_entry: 新条目，形如 'RMEET(反向)' 或 'UES(正向)'
        返回:
            合并后的字符串，如 'RMEET(反向),UES(正向)'
        """
        reverse_entries = []
        forward_entries = []

        if existing:
            for part in str(existing).split(','):
                part = part.strip()
                if not part:
                    continue
                if part.endswith('(正向)'):
                    forward_entries.append(part)
                elif part.endswith('(反向)'):
                    reverse_entries.append(part)
                else:
                    # 无标识的历史数据视为反向条目（兼容旧版本），补标识
                    reverse_entries.append(f"{part}(反向)")

        new_entry = new_entry.strip()
        if new_entry.endswith('(反向)'):
            if new_entry not in reverse_entries:
                reverse_entries.append(new_entry)
        else:
            if new_entry not in forward_entries:
                forward_entries.append(new_entry)

        return ','.join(reverse_entries + forward_entries)

    def find_best_match(self, target_answer, arr_data, original_question=None):
        """找到与目标答案最匹配的数据（基于关键字匹配+问题原文匹配）
        
        参数:
            target_answer: 目标答案文本
            arr_data: 候选项列表
            original_question: 原始问题文本，用于问题原文匹配加分（可选）
            
        返回字典: {'matched': bool, 'channel': str, 'item': dict/None,
                   'best_keyword_count': int, 'total_keywords': int, 'min_keyword_count': int,
                   'question_match_bonus': int}
        """
        # 空数据时返回统计信息为0的失败结果
        if not arr_data or not isinstance(arr_data, list):
            self.log(f"    arr_data为空或不是列表")
            return {'matched': False, 'channel': '', 'item': None,
                    'best_keyword_count': 0, 'total_keywords': 0, 'min_keyword_count': 0}

        # 清理目标答案
        clean_target = ' '.join(target_answer.split())

        # 提取关键词（细粒度分词）
        import re

        # 方法1: 提取中文词组（2-4个字）
        chinese_keywords = re.findall(r'[\u4e00-\u9fa5]{2,4}', clean_target)

        # 方法2: 提取英文单词
        english_keywords = re.findall(r'[a-zA-Z]{2,}', clean_target)

        # 方法3: 提取数字
        number_keywords = re.findall(r'[0-9]+', clean_target)

        # 合并所有关键词
        keywords = chinese_keywords + english_keywords + number_keywords

        # 去重
        keywords = list(set(keywords))

        if not keywords:
            self.log(f"    未提取到有效关键词")
            if self.debug_var.get():
                self.debug_log(f"错误: 无法从答案中提取关键词")
                self.debug_log(f"答案内容: {clean_target}")
            return {'matched': False, 'channel': '', 'item': None,
                    'best_keyword_count': 0, 'total_keywords': 0, 'min_keyword_count': 0}

        self.log(f"    提取关键词: {', '.join(keywords[:10])}")  # 显示前10个

        # 关键字匹配阈值：至少匹配超过98%的关键词才算成功（严格要求）
        keyword_threshold = 0.98  # 98%阈值
        min_keyword_count = max(2, int(len(keywords) * keyword_threshold))  # 至少需要2个关键词匹配

        self.log(f"    关键字匹配阈值: ≥{min_keyword_count}/{len(keywords)} ({int(keyword_threshold*100)}%)")

        # 调试模式：记录匹配过程
        if self.debug_var.get():
            self.debug_log(f"\n开始匹配过程 - 提取了{len(keywords)}个关键词")
            self.debug_log(f"关键词列表: {', '.join(keywords)}")
            self.debug_log(f"匹配阈值: ≥{min_keyword_count}/{len(keywords)} ({int(keyword_threshold*100)}%)")

        best_match = None
        best_keyword_count = 0
        best_score = 0
        best_question_bonus = 0  # 最佳匹配的问题原文加分
        question_matched_fully = False  # 是否有问题原文完整匹配

        for idx, item in enumerate(arr_data):
            if not isinstance(item, dict):
                continue

            content = item.get('content', '')
            if not content:
                continue

            # 清理content
            clean_content = ' '.join(content.split())

            # 统计关键词匹配数量
            keyword_count = 0
            matched_keywords = []
            for keyword in keywords:
                if keyword in clean_content:
                    keyword_count += 1
                    matched_keywords.append(keyword)

            # 问题原文匹配（加分项）
            question_match_bonus = 0
            if original_question:
                clean_question = ' '.join(str(original_question).split())
                # 问题原文完整出现在content中，加分
                if clean_question and clean_question in clean_content:
                    question_match_bonus = len(clean_question)  # 问题越长加分越多
                # 短问题（≤10字）部分匹配也加分
                elif clean_question and len(clean_question) <= 10:
                    # 检查问题字符是否大部分出现在content中
                    q_chars = set(clean_question)
                    c_chars = set(clean_content)
                    overlap = len(q_chars & c_chars)
                    if overlap >= len(q_chars) * 0.8:
                        question_match_bonus = overlap

            # 获取该条目的score
            item_score = item.get('score', 0)

            # 调试模式：记录每个条目的匹配详情
            if self.debug_var.get():
                self.debug_log(f"\n条目[{idx}] - channel: {item.get('channel', 'N/A')}")
                self.debug_log(f"  关键字匹配: {keyword_count}/{len(keywords)}")
                self.debug_log(f"  匹配的关键词: {', '.join(matched_keywords) if matched_keywords else '无'}")
                self.debug_log(f"  问题原文加分: {question_match_bonus}")
                self.debug_log(f"  Score: {item_score:.3f}")
                self.debug_log(f"  Content预览: {clean_content[:100]}...")

            # 综合得分 = 关键字匹配数 + 问题原文加分
            total_score = keyword_count + question_match_bonus

            # 选择逻辑：优先选择综合得分最高的
            if total_score > (best_keyword_count + best_question_bonus):
                best_keyword_count = keyword_count
                best_score = item_score
                best_match = item
                best_question_bonus = question_match_bonus
                if question_match_bonus > 0:
                    question_matched_fully = True
            elif total_score == (best_keyword_count + best_question_bonus) and total_score > 0:
                # 综合得分相同，比较score
                if item_score > best_score:
                    best_score = item_score
                    best_match = item
                    best_question_bonus = question_match_bonus

        # 计算有效的匹配阈值（问题原文完整匹配时降低阈值要求）
        effective_min_keyword_count = min_keyword_count
        if best_question_bonus > 0 and original_question:
            # 问题原文有匹配，降低关键字阈值到70%
            effective_min_keyword_count = max(1, int(len(keywords) * 0.7))

        # 至少匹配阈值数量的关键词才算成功
        if best_match and best_keyword_count >= effective_min_keyword_count:
            matched_channel = best_match.get('channel', '')
            match_info = f"匹配数={best_keyword_count}/{len(keywords)}"
            if best_question_bonus > 0:
                match_info += f", 问题原文加分={best_question_bonus}"
            self.log(f"  ✓ 匹配成功: {match_info}, 阈值≥{effective_min_keyword_count}, Score={best_score:.3f}, 渠道={matched_channel}")

            # 调试模式：记录最终匹配结果
            if self.debug_var.get():
                self.debug_log(f"\n最终匹配结果:")
                self.debug_log(f"  渠道: {matched_channel}")
                self.debug_log(f"  关键字匹配数: {best_keyword_count}/{len(keywords)}")
                self.debug_log(f"  问题原文加分: {best_question_bonus}")
                self.debug_log(f"  有效阈值: ≥{effective_min_keyword_count}")
                self.debug_log(f"  Score: {best_score:.3f}")
                self.debug_log(f"  Content: {best_match.get('content', 'N/A')[:200]}...\n")

            return {'matched': True, 'channel': matched_channel, 'item': best_match,
                    'best_keyword_count': best_keyword_count, 'total_keywords': len(keywords),
                    'min_keyword_count': effective_min_keyword_count,
                    'question_match_bonus': best_question_bonus}

        # 匹配未达到阈值，返回统计信息触发相似度匹配
        self.log(f"    匹配未达阈值 (匹配数={best_keyword_count}/{len(keywords)}, 需要≥{effective_min_keyword_count})")

        if self.debug_var.get():
            self.debug_log(f"\n匹配失败:")
            self.debug_log(f"  最佳匹配数: {best_keyword_count}/{len(keywords)}")
            self.debug_log(f"  问题原文加分: {best_question_bonus}")
            self.debug_log(f"  有效阈值: ≥{effective_min_keyword_count}")
            self.debug_log(f"  将触发相似度API进行二次匹配\n")

        return {'matched': False, 'channel': '', 'item': None,
                'best_keyword_count': best_keyword_count, 'total_keywords': len(keywords),
                'min_keyword_count': effective_min_keyword_count,
                'question_match_bonus': best_question_bonus}


def main():
    """主函数：从启动起全程写运行日志；异常同时落盘 启动错误.log 并弹窗提示"""
    import traceback
    _runtime_log("=" * 60)
    _runtime_log("程序启动")
    _runtime_log(f"运行模式: {'打包可执行文件' if getattr(sys, 'frozen', False) else '源码运行'}")
    _runtime_log(f"程序路径: {sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)}")
    _runtime_log(f"工作目录: {os.getcwd()}")
    _runtime_log(f"日志目录: {_resolve_log_dir()}")

    root = None
    try:
        _runtime_log("正在创建主窗口...")
        root = tk.Tk()
        _runtime_log(f"Tk版本: {root.tk.call('info', 'patchlevel')}")
        _runtime_log(f"屏幕尺寸: {root.winfo_screenwidth()}x{root.winfo_screenheight()}")
        _runtime_log("正在初始化应用界面...")
        app = ExcelProcessorApp(root)  # 内部完成窗口居中与启动画面
        _runtime_log("界面初始化完成，进入主循环")
    except Exception:
        _runtime_log("启动异常:\n" + traceback.format_exc())
        try:
            err_path = os.path.join(_resolve_log_dir(), '启动错误.log')
            with open(err_path, 'a', encoding='utf-8') as fh:
                fh.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 启动异常:\n{traceback.format_exc()}\n")
        except Exception:
            pass
        try:
            if root is not None:
                messagebox.showerror("启动失败", f"程序启动时发生异常，详情见 运行日志.log：\n\n{traceback.format_exc()}")
        except Exception:
            pass
        raise

    try:
        root.mainloop()
    except Exception:
        _runtime_log("运行异常:\n" + traceback.format_exc())
        raise
    finally:
        _runtime_log("程序退出")


if __name__ == "__main__":
    main()
